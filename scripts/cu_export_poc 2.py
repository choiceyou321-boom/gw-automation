"""
Computer Use Export PoC — 근태신청현황 페이지에서 엑셀 다운로드 시도

전략:
1) Playwright로 LNB 클릭 시퀀스 (HR → 근태관리 → 근태신청 → 근태신청현황) 진입
2) 그리드가 로드될 때까지 대기
3) ComputerUseGWAgent로 "화면의 엑셀 아이콘을 찾아 클릭" task 호출
4) page.expect_download() 컨텍스트에서 다운로드 캡처
5) 파일 저장 + openpyxl로 헤더+5행 미리보기

실행: .venv/bin/python scripts/cu_export_poc.py
"""
from __future__ import annotations
import json, logging, re, sys, time, os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
# .env 명시적 로드 (ComputerUseGWAgent가 ANTHROPIC_API_KEY를 os.environ에서 직접 읽음)
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context
from src.gw.approval.computer_use_agent import ComputerUseGWAgent

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("cu_poc")

ROOT = Path(__file__).resolve().parent.parent
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)


def click_text(page: Page, text: str, timeout_ms: int = 4000) -> bool:
    """page + iframe 모두에서 텍스트 찾아 클릭"""
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        for sel in [f"text='{text}'", f"li:has-text('{text}')", f"a:has-text('{text}')"]:
            try:
                loc = fr.locator(sel).first
                if loc.count() > 0 and loc.is_visible(timeout=500):
                    loc.click(force=True, timeout=timeout_ms)
                    return True
            except Exception:
                continue
    return False


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def navigate_to_attendance_status(page: Page) -> bool:
    """홈 → HR → 근태관리 → 근태신청 → 근태신청현황 진입"""
    m = re.match(r"(https://[^/]+)", page.url)
    base = m.group(1) if m else "https://gw.glowseoul.co.kr"

    page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(4000)

    # HR 진입
    try:
        page.locator("span.module-link.HR").first.click(force=True, timeout=5000)
        page.wait_for_timeout(5000)
        logger.info("✓ HR 모듈 진입")
    except Exception as e:
        logger.error(f"HR 진입 실패: {e}")
        return False

    # 우선 이미 열린 탭에 근태신청현황이 있는지 확인
    if click_text(page, "근태신청현황", timeout_ms=2000):
        page.wait_for_timeout(3000)
        logger.info("✓ 기존 탭의 근태신청현황 클릭 성공")
        return True

    # 단계별 LNB 펼침
    for step_text, wait_ms in [("근태관리", 2500), ("근태신청", 2500), ("근태신청현황", 4000)]:
        ok = click_text(page, step_text, timeout_ms=4000)
        if ok:
            page.wait_for_timeout(wait_ms)
            logger.info(f"✓ '{step_text}' 클릭")
        else:
            logger.warning(f"✗ '{step_text}' 못 찾음 — 2차 시도")
            page.wait_for_timeout(1500)
            ok = click_text(page, step_text, timeout_ms=4000)
            if ok:
                page.wait_for_timeout(wait_ms)
                logger.info(f"✓ '{step_text}' 2차 클릭")
            else:
                logger.error(f"'{step_text}' 진입 실패")
                # 진단용 스크린샷
                try:
                    page.screenshot(path=str(SCR_DIR / f"cu_fail_{step_text}.png"))
                except Exception:
                    pass
                if step_text == "근태신청현황":
                    return False
    return True


def attempt_excel_download_via_cu(page: Page) -> dict:
    """Computer Use로 엑셀 다운로드 시도"""
    agent = ComputerUseGWAgent(page=page)
    task = (
        "현재 화면은 근태신청현황 페이지입니다. "
        "데이터 그리드가 보입니다. "
        "이 그리드 데이터를 엑셀 파일로 다운로드해야 합니다. "
        "화면에서 '엑셀', '다운로드', '엑셀 다운로드' 같은 텍스트나 아이콘을 찾아 클릭하세요. "
        "그리드 우측 상단의 작은 아이콘들 중에 다운로드 아이콘이 있을 가능성이 높습니다. "
        "확인 다이얼로그가 뜨면 '확인' 또는 '다운로드'를 클릭하세요."
    )

    saved_path = None
    try:
        with page.expect_download(timeout=30000) as dl_info:
            result = agent.fill_form_with_vision(task=task, data={}, max_steps=10)
            logger.info(f"CU 결과: {result}")
        dl = dl_info.value
        suffix = Path(dl.suggested_filename).suffix or ".xlsx"
        saved_path = DL_DIR / f"attendance_status{suffix}"
        dl.save_as(str(saved_path))
        logger.info(f"★ 저장: {saved_path.name} ({saved_path.stat().st_size}B)")
        return {"success": True, "path": str(saved_path), "summary": summarize_xlsx(saved_path)}
    except PWTimeout:
        return {"success": False, "error": "다운로드 타임아웃"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def main():
    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw,
            headless=False,
            user_id="tgjeon",
        )
        # 뷰포트 키우기 (CU가 화면 분석할 때 잘 보이도록)
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass

        if not navigate_to_attendance_status(page):
            logger.error("근태신청현황 진입 실패")
            return

        page.screenshot(path=str(SCR_DIR / "cu_before_download.png"))
        logger.info("Computer Use 다운로드 시도 시작")
        result = attempt_excel_download_via_cu(page)
        logger.info(f"=== 최종 결과 ===")
        logger.info(json.dumps(result, ensure_ascii=False, indent=2)[:1500])

        context.close()
        browser.close()


if __name__ == "__main__":
    main()
