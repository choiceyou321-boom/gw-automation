"""
프로젝트 프로젝트 관리표 xlsx 양식 생성 스크립트
글로우서울 인테리어/건축 시공 회사

생성 파일: data/프로젝트_프로젝트 관리표_양식.xlsx
시트 구성:
  1. 프로젝트 관리표 (메인)
  2. 지출내역
  3. 발주현황
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 공통 스타일 정의 ─────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # 진한 파란
EVEN_FILL    = PatternFill("solid", fgColor="D6E4F0")   # 연한 파란
HEADER_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="맑은 고딕", size=10)
BOLD_FONT    = Font(name="맑은 고딕", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color="AAAAAA")
MED  = Side(style="medium", color="1F4E79")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEAD = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

FMT_NUMBER   = '#,##0'
FMT_PERCENT  = '0.0%'
FMT_DATE     = 'YYYY-MM-DD'
FMT_TEXT     = '@'


def set_header(ws, row, cols):
    """헤더 행 스타일 일괄 적용. cols: [(col_idx, text, width), ...]"""
    for col_idx, text, width in cols:
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 22


def style_body(ws, row, col, value=None, fmt=None, align=CENTER, bold=False):
    """본문 셀 스타일 적용"""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if fmt:
        cell.number_format = fmt
    cell.font      = BOLD_FONT if bold else BODY_FONT
    cell.alignment = align
    cell.border    = BORDER_THIN
    return cell


def apply_even_row(ws, row, max_col):
    """짝수 행 배경색 적용"""
    if row % 2 == 0:
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = EVEN_FILL


# ── 시트1: 프로젝트 관리표 (메인) ──────────────────────────────────────────────────

def build_sheet1(wb):
    ws = wb.create_sheet("프로젝트 관리표")

    # 헤더 정의: (col, 헤더명, 열너비)
    headers = [
        (1,  "순번",               6),
        (2,  "프로젝트코드",       14),
        (3,  "프로젝트명",         28),
        (4,  "발주처/클라이언트", 18),
        (5,  "계약금액(원)",       16),
        (6,  "부가세",             14),
        (7,  "계약금액\n(VAT포함)",16),
        (8,  "계약일",             13),
        (9,  "공사시작일",         13),
        (10, "공사종료일",         13),
        (11, "선급금(원)",         14),
        (12, "기성금1차",          13),
        (13, "기성금2차",          13),
        (14, "기성금3차",          13),
        (15, "잔금",               13),
        (16, "총수령금액",         14),
        (17, "총지출금액",         14),
        (18, "잔액",               13),
        (19, "수익률(%)",          11),
        (20, "비고/메모",          20),
    ]
    set_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    # 예시 데이터
    sample = [
        # row 2
        [
            1, "GS-25-0088", "강남 오피스 인테리어", "(주)글로우파트너스",
            50_000_000, "=E2*0.1", "=E2+F2",
            "2025-03-01", "2025-03-15", "2025-05-30",
            5_000_000, 20_000_000, 15_000_000, 0, 0,
            "=SUM(K2:O2)", 38_000_000, "=P2-Q2", "=IF(P2=0,0,R2/P2)",
            "1차 기성 완료"
        ],
        # row 3
        [
            2, "GS-25-0102", "서초 카페 인테리어", "개인 홍길동",
            28_000_000, "=E3*0.1", "=E3+F3",
            "2025-04-10", "2025-04-20", "2025-06-15",
            8_000_000, 0, 0, 0, 0,
            "=SUM(K3:O3)", 22_000_000, "=P3-Q3", "=IF(P3=0,0,R3/P3)",
            "진행중"
        ],
        # row 4
        [
            3, "GS-25-0115", "마포 음식점 리모델링", "(주)맛나푸드",
            15_000_000, "=E4*0.1", "=E4+F4",
            "2025-05-01", "2025-05-10", "2025-07-10",
            3_000_000, 8_000_000, 4_000_000, 0, 0,
            "=SUM(K4:O4)", 13_500_000, "=P4-Q4", "=IF(P4=0,0,R4/P4)",
            "공사중"
        ],
    ]

    # 컬럼별 숫자 형식 (1-based col index)
    number_cols  = {5, 6, 7, 11, 12, 13, 14, 15, 16, 17, 18}
    date_cols    = {8, 9, 10}
    percent_cols = {19}
    text_cols    = {2, 20}

    for r_idx, row_data in enumerate(sample, start=2):
        apply_even_row(ws, r_idx, 20)
        for c_idx, val in enumerate(row_data, start=1):
            if c_idx in number_cols:
                style_body(ws, r_idx, c_idx, val, FMT_NUMBER, RIGHT)
            elif c_idx in date_cols:
                style_body(ws, r_idx, c_idx, val, FMT_DATE, CENTER)
            elif c_idx in percent_cols:
                style_body(ws, r_idx, c_idx, val, FMT_PERCENT, CENTER)
            elif c_idx == 1:
                style_body(ws, r_idx, c_idx, val, None, CENTER)
            elif c_idx in text_cols:
                style_body(ws, r_idx, c_idx, val, FMT_TEXT, LEFT)
            else:
                style_body(ws, r_idx, c_idx, val, None, LEFT)

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    # 합계 행 (row 5)
    sum_row = 5
    ws.row_dimensions[sum_row].height = 22
    ws.cell(sum_row, 1).value = "합계"
    ws.cell(sum_row, 1).font  = BOLD_FONT
    ws.cell(sum_row, 1).alignment = CENTER
    ws.cell(sum_row, 1).fill  = HEADER_FILL
    ws.cell(sum_row, 1).font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    ws.merge_cells(f"A{sum_row}:D{sum_row}")

    sum_number_cols = {5, 6, 7, 11, 12, 13, 14, 15, 16, 17, 18}
    col_letters = {5:"E", 6:"F", 7:"G", 11:"K", 12:"L", 13:"M", 14:"N", 15:"O",
                   16:"P", 17:"Q", 18:"R"}
    for c in range(1, 21):
        cell = ws.cell(sum_row, c)
        cell.border = BORDER_THIN
        cell.fill   = PatternFill("solid", fgColor="1F4E79")
        cell.font   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        if c in sum_number_cols:
            letter = col_letters[c]
            cell.value = f"=SUM({letter}2:{letter}4)"
            cell.number_format = FMT_NUMBER
            cell.alignment = RIGHT

    return ws


# ── 시트2: 지출내역 ───────────────────────────────────────────────────────────

def build_sheet2(wb):
    ws = wb.create_sheet("지출내역")

    headers = [
        (1,  "순번",         6),
        (2,  "프로젝트코드", 14),
        (3,  "프로젝트명",   24),
        (4,  "지출일",       13),
        (5,  "지출항목",     18),
        (6,  "거래처",       18),
        (7,  "공급가액",     14),
        (8,  "부가세",       12),
        (9,  "합계금액",     14),
        (10, "결재상태",     14),
        (11, "결재문서번호", 16),
        (12, "비고",         20),
    ]
    set_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    # 지출항목 드롭다운 목록 (데이터 유효성 검사)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_item = DataValidation(
        type="list",
        formula1='"외주공사비,재료비,인건비,경비,설계비,감리비,기타"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_item.sqref = "E2:E1000"
    ws.add_data_validation(dv_item)

    dv_status = DataValidation(
        type="list",
        formula1='"미결재,결재완료,임시저장,반려"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.sqref = "J2:J1000"
    ws.add_data_validation(dv_status)

    # 예시 데이터
    sample = [
        [1, "GS-25-0088", "강남 오피스 인테리어",  "2025-04-10", "외주공사비", "(주)ABC인테리어",  15_000_000, "=G2*0.1", "=G2+H2", "결재완료", "EA-2025-0041", ""],
        [2, "GS-25-0088", "강남 오피스 인테리어",  "2025-04-22", "재료비",     "한국자재(주)",       5_500_000, "=G3*0.1", "=G3+H3", "결재완료", "EA-2025-0055", ""],
        [3, "GS-25-0088", "강남 오피스 인테리어",  "2025-05-03", "경비",       "직접지출",              300_000, "=G4*0.1", "=G4+H4", "미결재",   "",             "택시비 포함"],
        [4, "GS-25-0102", "서초 카페 인테리어",    "2025-05-01", "외주공사비", "(주)카페인테리어",  10_000_000, "=G5*0.1", "=G5+H5", "임시저장", "",             ""],
        [5, "GS-25-0115", "마포 음식점 리모델링",  "2025-05-20", "외주공사비", "맛나인테리어",       8_000_000, "=G6*0.1", "=G6+H6", "결재완료", "EA-2025-0071", ""],
    ]

    number_cols  = {7, 8, 9}
    date_cols    = {4}
    status_col   = {10}

    for r_idx, row_data in enumerate(sample, start=2):
        apply_even_row(ws, r_idx, 12)
        for c_idx, val in enumerate(row_data, start=1):
            if c_idx in number_cols:
                style_body(ws, r_idx, c_idx, val, FMT_NUMBER, RIGHT)
            elif c_idx in date_cols:
                style_body(ws, r_idx, c_idx, val, FMT_DATE, CENTER)
            elif c_idx == 1:
                style_body(ws, r_idx, c_idx, val, None, CENTER)
            elif c_idx in {2, 11}:
                style_body(ws, r_idx, c_idx, val, FMT_TEXT, CENTER)
            elif c_idx in status_col:
                cell = style_body(ws, r_idx, c_idx, val, None, CENTER)
                # 결재상태별 색상
                if val == "결재완료":
                    cell.font = Font(name="맑은 고딕", size=10, color="0070C0", bold=True)
                elif val == "미결재":
                    cell.font = Font(name="맑은 고딕", size=10, color="FF0000")
                elif val == "임시저장":
                    cell.font = Font(name="맑은 고딕", size=10, color="FFC000", bold=True)
            else:
                style_body(ws, r_idx, c_idx, val, None, LEFT)
        ws.row_dimensions[r_idx].height = 20

    # 합계 행
    sum_row = len(sample) + 2
    ws.row_dimensions[sum_row].height = 22
    ws.merge_cells(f"A{sum_row}:F{sum_row}")
    cell = ws.cell(sum_row, 1)
    cell.value = "합계"
    cell.fill  = HEADER_FILL
    cell.font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    cell.alignment = CENTER
    for c in range(1, 13):
        ws.cell(sum_row, c).border = BORDER_THIN
        ws.cell(sum_row, c).fill  = HEADER_FILL
        ws.cell(sum_row, c).font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    for c, letter in {7: "G", 8: "H", 9: "I"}.items():
        ws.cell(sum_row, c).value = f"=SUM({letter}2:{letter}{sum_row-1})"
        ws.cell(sum_row, c).number_format = FMT_NUMBER
        ws.cell(sum_row, c).alignment = RIGHT

    return ws


# ── 시트3: 발주현황 ───────────────────────────────────────────────────────────

def build_sheet3(wb):
    ws = wb.create_sheet("발주현황")

    headers = [
        (1,  "순번",       6),
        (2,  "프로젝트코드", 14),
        (3,  "공종",       18),
        (4,  "협력사명",   18),
        (5,  "계약금액",   14),
        (6,  "선급금",     12),
        (7,  "기성1",      12),
        (8,  "기성2",      12),
        (9,  "잔금",       12),
        (10, "지급완료금액", 14),
        (11, "미지급액",   13),
        (12, "계약일",     13),
        (13, "비고",       20),
    ]
    set_header(ws, 1, headers)
    ws.freeze_panes = "A2"

    # 공종 드롭다운
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_type = DataValidation(
        type="list",
        formula1='"음향,경량칸막이,전기,설비,목공,타일,도장,유리,가구,간판,기타"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref = "C2:C1000"
    ws.add_data_validation(dv_type)

    # 예시 데이터
    sample = [
        [1, "GS-25-0088", "경량칸막이", "(주)ABC인테리어",  18_000_000, 3_000_000, 10_000_000, 5_000_000, 0, "=SUM(F2:I2)", "=E2-J2", "2025-03-05", ""],
        [2, "GS-25-0088", "전기",       "동서전기(주)",      8_000_000, 1_000_000,  4_000_000, 3_000_000, 0, "=SUM(F3:I3)", "=E3-J3", "2025-03-05", ""],
        [3, "GS-25-0088", "음향",       "사운드코리아",      5_000_000,         0,  5_000_000,         0, 0, "=SUM(F4:I4)", "=E4-J4", "2025-03-10", ""],
        [4, "GS-25-0102", "목공",       "나무랑(주)",        12_000_000, 2_000_000,         0,         0, 0, "=SUM(F5:I5)", "=E5-J5", "2025-04-15", "진행중"],
        [5, "GS-25-0115", "설비",       "설비전문(주)",       6_000_000, 1_000_000,  5_000_000,         0, 0, "=SUM(F6:I6)", "=E6-J6", "2025-05-05", ""],
    ]

    number_cols = {5, 6, 7, 8, 9, 10, 11}
    date_cols   = {12}

    for r_idx, row_data in enumerate(sample, start=2):
        apply_even_row(ws, r_idx, 13)
        for c_idx, val in enumerate(row_data, start=1):
            if c_idx in number_cols:
                cell = style_body(ws, r_idx, c_idx, val, FMT_NUMBER, RIGHT)
                # 미지급액이 0보다 크면 강조
                if c_idx == 11 and isinstance(val, str) and val.startswith("="):
                    pass  # 수식이므로 조건부 서식은 별도 처리
            elif c_idx in date_cols:
                style_body(ws, r_idx, c_idx, val, FMT_DATE, CENTER)
            elif c_idx == 1:
                style_body(ws, r_idx, c_idx, val, None, CENTER)
            elif c_idx == 2:
                style_body(ws, r_idx, c_idx, val, FMT_TEXT, CENTER)
            else:
                style_body(ws, r_idx, c_idx, val, None, LEFT)
        ws.row_dimensions[r_idx].height = 20

    # 합계 행
    sum_row = len(sample) + 2
    ws.row_dimensions[sum_row].height = 22
    ws.merge_cells(f"A{sum_row}:D{sum_row}")
    for c in range(1, 14):
        ws.cell(sum_row, c).border = BORDER_THIN
        ws.cell(sum_row, c).fill  = HEADER_FILL
        ws.cell(sum_row, c).font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    ws.cell(sum_row, 1).value = "합계"
    ws.cell(sum_row, 1).alignment = CENTER
    for c, letter in {5:"E", 6:"F", 7:"G", 8:"H", 9:"I", 10:"J", 11:"K"}.items():
        ws.cell(sum_row, c).value = f"=SUM({letter}2:{letter}{sum_row-1})"
        ws.cell(sum_row, c).number_format = FMT_NUMBER
        ws.cell(sum_row, c).alignment = RIGHT

    return ws


# ── 메인 실행 ─────────────────────────────────────────────────────────────────

def main():
    base_dir  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path  = os.path.join(base_dir, "data", "프로젝트_프로젝트 관리표_양식.xlsx")

    wb = Workbook()
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)

    wb.save(out_path)
    print(f"✅ 파일 생성 완료: {out_path}")
    print(f"   시트 목록: {wb.sheetnames}")
    return out_path


if __name__ == "__main__":
    main()
