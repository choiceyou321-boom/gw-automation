"""
자원(UK) / 일정(UE) 모듈에서 엑셀 다운로드 자동 실행
- v3에서 'specialLnb=Y' URL 진입 시 '엑셀 다운로드즐겨찾기' 버튼 발견됨
- 해당 URL로 직접 이동 → "엑셀 다운로드" 버튼 클릭 → 다운로드 캡처

대상 URL:
  RM: /#/UK/UKA/UKA0000?specialLnb=Y&moduleCode=UK&menuCode=UKA&pageCode=UKA0000
  CL: /#/UE/UEA/UEA0000?specialLnb=Y&moduleCode=UE&menuCode=UEA&pageCode=UEA0000
"""
from __future__ import annotations
import json, logging, re, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("export_uk_ue")

ROOT = Path(__file__).resolve().parent.parent
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)

TARGETS = [
    # (모듈 코드, 라벨, URL path, span.module-link 클래스, 다운로드 파일명)
    ("UK", "자원예약(RM)", "/#/UK/UKA/UKA0000", "RM", "resource_reservation"),
    ("UE", "일정(CL)", "/#/UE/UEA/UEA0000", "CL", "schedule"),
]


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def try_download_in_all_frames(page: Page, button_text: str, save_path: Path, timeout_ms: int = 20000) -> dict:
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        # 부분 일치 셀렉터 시도 (text/title/aria-label 모두)
        selectors = [
            f"[title*='{button_text}']",       # ★ title 속성 (텍스트 없는 아이콘 대응)
            f"[aria-label*='{button_text}']",
            f"button:has-text('{button_text}')",
            f"a:has-text('{button_text}')",
            f"[role='button']:has-text('{button_text}')",
            f"text='{button_text}'",
        ]
        for sel in selectors:
            try:
                loc = fr.locator(sel).first
                if loc.count() == 0:
                    continue
                if not loc.is_visible(timeout=500):
                    continue
                logger.info(f"      → '{button_text}' 발견(셀렉터={sel}), 클릭")
                try:
                    with page.expect_download(timeout=timeout_ms) as dl_info:
                        loc.click(force=True, timeout=3000)
                        # 확인 모달 대응
                        try:
                            for ok in ["확인", "다운로드", "OK", "저장"]:
                                ok_loc = fr.locator(f"button:has-text('{ok}')").last
                                if ok_loc.count() > 0 and ok_loc.is_visible(timeout=400):
                                    ok_loc.click(timeout=1500)
                                    break
                        except Exception:
                            pass
                    dl: Download = dl_info.value
                    suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                    path = save_path.with_suffix(suffix)
                    dl.save_as(str(path))
                    size = path.stat().st_size if path.exists() else 0
                    logger.info(f"      ★ 저장: {path.name} ({size}B)")
                    return {"saved": str(path), "size": size,
                            "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
                except PWTimeout:
                    logger.warning(f"      다운로드 타임아웃 (sel={sel})")
                    continue
                except Exception as e:
                    logger.warning(f"      다운로드 실패 (sel={sel}): {e}")
                    continue
            except Exception:
                continue
    return {"saved": None, "error": "no_download_triggered"}


def main():
    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw, headless=False, user_id="tgjeon",
        )
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        all_results = []
        for mod_code, label, path, home_icon, fname in TARGETS:
            logger.info(f"\n=== {label} ({path}) ===")
            # 홈 진입 후 module-link 클릭 (이게 v3에서 검증된 진입 방법)
            page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3500)
            try:
                page.locator(f"span.module-link.{home_icon}").first.click(force=True, timeout=5000)
                page.wait_for_timeout(4500)
                logger.info(f"  ✓ {home_icon} 모듈 진입")
            except Exception as e:
                logger.error(f"  {home_icon} 진입 실패: {e}")
                continue

            try:
                page.screenshot(path=str(SCR_DIR / f"export_{home_icon}_before.png"))
            except Exception:
                pass

            # 다운로드 시도 — '엑셀 다운로드', '엑셀다운로드', '엑셀' 순
            save_path = DL_DIR / fname
            result = None
            for btn_text in ["엑셀 다운로드", "엑셀다운로드", "엑셀"]:
                result = try_download_in_all_frames(page, btn_text, save_path)
                if result.get("saved"):
                    break

            try:
                page.screenshot(path=str(SCR_DIR / f"export_{home_icon}_after.png"))
            except Exception:
                pass

            all_results.append({"target": label, "url_path": path, **result})

        context.close()
        browser.close()

    # 결과 요약
    print("\n" + "=" * 70)
    for r in all_results:
        print(f"\n## {r['target']}")
        if r.get("saved"):
            print(f"  ✓ 저장: {r['saved']}")
            print(f"  크기: {r['size']}B")
            summary = r.get("summary")
            if summary and "error" not in summary:
                for sheet, rows in (summary.get("preview") or {}).items():
                    print(f"\n  시트 '{sheet}':")
                    for i, row in enumerate(rows[:6]):
                        print(f"    [{i}] " + " | ".join(c[:25] for c in row[:8]))
        else:
            print(f"  ✗ {r.get('error', 'unknown')}")

    OUT = ROOT / "data" / "export_uk_ue_result.json"
    OUT.write_text(json.dumps(all_results, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    logger.info(f"결과 저장: {OUT}")


if __name__ == "__main__":
    main()
