"""
계약서 입력양식 Excel 생성 스크립트
실행: python scripts/create_contract_template.py
결과: data/계약서_입력양식.xlsx
"""
import pathlib
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(__file__).parent.parent / "data" / "계약서_입력양식.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
EXAMPLE_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def header_style(cell, text):
    cell.value = text
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def example_style(cell, text):
    cell.value = text
    cell.fill = EXAMPLE_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    cell.font = Font(size=9, color="444444")

def data_style(cell):
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.font = Font(size=10)


wb = openpyxl.Workbook()

# ── Sheet 1: 자재납품계약서 ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "자재납품계약서"

headers_mat = [
    ("협력사명", "예: 오티엘"),
    ("협력사_법인명", "예: 주식회사 오티엘 (없으면 협력사명과 동일)"),
    ("대표자", "예: 이 봉 식"),
    ("주소", "예: 서울특별시 도봉구 도봉산3길 46, 지하층(도봉동)"),
    ("사업자번호", "예: 390-86-03578"),
    ("자재유형", "예: 조명, 커튼, 위생기구, 도어락 등"),
    ("자재상세설명", "예: 조명(등기구) 제작 및 납품 (견적서 참조)"),
    ("계약금액", "숫자만 입력 (VAT 별도, 예: 47500000)"),
    ("계약금비율(%)", "숫자만 입력 (예: 40, 잔금은 자동)"),
    ("중도금비율(%)", "숫자만 입력 (예: 50, 잔금은 자동)"),
    ("납품일자", "예: 2026-02-28"),
    ("설치완료일", "예: 2026-03-16 (없으면 납품일자와 동일)"),
    ("계약일", "예: 2026-02-03 (없으면 오늘)"),
]

ws1.row_dimensions[1].height = 32
ws1.row_dimensions[2].height = 24
for i in range(3, 15):
    ws1.row_dimensions[i].height = 22

for col, (h, ex) in enumerate(headers_mat, 1):
    header_style(ws1.cell(1, col), h)
    example_style(ws1.cell(2, col), ex)

# 예시 데이터 행
example_mat = [
    "오티엘", "주식회사 오티엘", "이 봉 식",
    "서울특별시 도봉구 도봉산3길 46, 지하층(도봉동)", "390-86-03578",
    "조명", "조명(등기구) 제작 및 납품 (견적서 참조)",
    "47500000", "40", "50",
    "2026-02-28", "2026-03-16", "2026-02-03",
]
for col, val in enumerate(example_mat, 1):
    c = ws1.cell(3, col)
    c.value = val
    data_style(c)

# 빈 입력 행 5개
for row in range(4, 9):
    for col in range(1, len(headers_mat) + 1):
        data_style(ws1.cell(row, col))

# 컬럼 너비
col_widths_mat = [14, 22, 14, 40, 16, 12, 36, 16, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths_mat, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 안내 텍스트
ws1["A10"].value = "※ 계약금비율 + 중도금비율 = 90 (잔금 10%는 자동 계산됩니다)"
ws1["A10"].font = Font(color="FF0000", italic=True, size=9)

# ── Sheet 2: 공사계약서 ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("공사계약서")

headers_con = [
    ("협력사명", "예: 호림ENG, 아트음향"),
    ("대표자", "예: 이 만 식"),
    ("주소", "예: 서울특별시 강동구 구천면로 361, 102동 601호"),
    ("공사유형", "예: 위생배관공사, 음향공사, 전기공사, 도장공사"),
    ("공급가액", "숫자만 입력 (VAT 제외, 예: 48000000) ← VAT 자동 계산"),
    ("착공일", "예: 2026-02-05"),
    ("준공예정일", "예: 2026-03-16"),
    ("중도금2차조건", "예: 공종 작업 완료 이후 / 시운전 완료 이후 / 설치 완료 이후"),
    ("계약일", "예: 2026-02-05 (없으면 오늘)"),
]

ws2.row_dimensions[1].height = 32
ws2.row_dimensions[2].height = 24
for i in range(3, 12):
    ws2.row_dimensions[i].height = 22

for col, (h, ex) in enumerate(headers_con, 1):
    header_style(ws2.cell(1, col), h)
    example_style(ws2.cell(2, col), ex)

# 예시 데이터 2행
examples_con = [
    ["호림ENG", "이 만 식",
     "서울특별시 강동구 구천면로 361, 102동 601호(암사동, 암사 THE Nest)",
     "위생배관공사", "48000000", "2026-02-05", "2026-03-16",
     "공종 작업 완료 이후", "2026-02-05"],
    ["아트음향", "홍 영 자",
     "인천광역시 서구 건지로 281, 1층(석남동)",
     "음향공사", "8600000", "2026-02-05", "2026-03-16",
     "시운전 완료 이후", "2026-02-05"],
]
for r_offset, ex_row in enumerate(examples_con):
    for col, val in enumerate(ex_row, 1):
        c = ws2.cell(3 + r_offset, col)
        c.value = val
        data_style(c)

# 빈 입력 행 5개
for row in range(5, 10):
    for col in range(1, len(headers_con) + 1):
        data_style(ws2.cell(row, col))

col_widths_con = [14, 14, 46, 18, 16, 14, 14, 28, 14]
for i, w in enumerate(col_widths_con, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2["A11"].value = "※ 대금 지급 구조: 계약금30% + 중도금1차30% + 중도금2차30% + 잔금10% (자동 계산)"
ws2["A11"].font = Font(color="FF0000", italic=True, size=9)

# ── Sheet 3: 작성 가이드 ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("작성가이드")
guide_rows = [
    ("계약서 자동 작성 가이드", ""),
    ("", ""),
    ("📌 사용 방법", ""),
    ("1. 자재납품계약서 또는 공사계약서 시트에 정보를 입력합니다.", ""),
    ("2. 챗봇에게 파일을 첨부하고 '계약서 작성해줘'라고 하면 자동으로 생성됩니다.", ""),
    ("3. 또는 스크립트 직접 실행: python scripts/generate_contracts_from_excel.py data/계약서_입력양식.xlsx", ""),
    ("", ""),
    ("📌 주의사항", ""),
    ("- 계약금비율 + 중도금비율 합계는 90이 되어야 합니다 (잔금 10% 자동)", ""),
    ("- 날짜는 YYYY-MM-DD 형식으로 입력하세요", ""),
    ("- 금액은 쉼표 없이 숫자만 입력하세요 (예: 47500000)", ""),
    ("- 협력사_법인명을 비워두면 협력사명과 동일하게 처리됩니다", ""),
    ("", ""),
    ("📌 공사 유형 예시", ""),
    ("위생배관공사, 음향공사, 전기공사, 냉난방공사, 도장공사, 금속공사, 커튼월공사", ""),
    ("", ""),
    ("📌 중도금2차 조건 예시", ""),
    ("공종 작업 완료 이후 (위생배관, 전기 등)", ""),
    ("시운전 완료 이후 (음향, 냉난방 등)", ""),
    ("설치 완료 이후 (가구, 조명 등)", ""),
]
for row_idx, (col_a, col_b) in enumerate(guide_rows, 1):
    ws3.cell(row_idx, 1).value = col_a
    ws3.cell(row_idx, 2).value = col_b
    if row_idx == 1:
        ws3.cell(row_idx, 1).font = Font(bold=True, size=14, color="1F4E79")
    elif col_a.startswith("📌"):
        ws3.cell(row_idx, 1).font = Font(bold=True, size=11)
ws3.column_dimensions["A"].width = 75
ws3.column_dimensions["B"].width = 20

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print(f"✅ 계약서 입력양식 생성 완료: {OUT}")
