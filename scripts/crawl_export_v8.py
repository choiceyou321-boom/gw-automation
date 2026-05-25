"""
아마란스 v8 — 알려진 URL 직접 진입 + cel_save 다운로드
─────────────────────────────────────────────────
v3~v7 결론: LNB 자동 leaf 클릭은 신뢰성 부족
v8 전략: 이미 매핑된 정확한 페이지 URL로 직접 goto → 다운로드

알려진 URL (GW_PAGES_ANALYSIS.md + v3 발견 기반):
  - 근태신청현황: /#/HP/HPD0122/HRD0220 (600행, v4에서 진입 확인)
  - 지출결의이체현황: /#/HP/APB1020/APB1020 (109건)
  - 실행예산신청: /#/BN/NCB0020/NCB0020
  - 실행예산마감: /#/BN/NCB0025/NCB0025
  - 예산조정신청: /#/BN/NCB0030/NCB0030
  - 프로젝트등록: /#/BN/NCF0090/SYB0060 (200건)
  - 예산일계표: /#/BN/NCC0230/NCC0230
  - 예산월계표: /#/BN/NCC0240/NCC0240
  - 세출총괄표: /#/BN/NCC0430/NCC0430
  - 세입총괄표: /#/BN/NCC0440/NCC0440
  - 예실대비현황: /#/BN/NCC0610/NCC0610
  - 예산구성비현황: /#/BN/NCC0620/NCC0620
  - 예실대비현황(상세): /#/BN/NCC0630/NCC0630
  - 예실대비현황(사업별): /#/BN/NCC0631/NCC0631
  - 예산과목원장: /#/BN/NCC0640/NCC0640
  - 예산과목등록: /#/BN/NCF0030/NCF0030
  - 자원예약: /#/UK/UKA/UKA0000
  - 일정: /#/UE/UEA/UEA0000
  - 메일: /#/UD/UDA/UDA0000
"""
from __future__ import annotations
import argparse, json, logging, re, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("v8")

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "data" / "amaranth_v8.json"
OUT_MD = ROOT / "docs" / "AMARANTH_EXPORT_v8.md"
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)

# (label, url_path, category)
TARGETS = [
    # 근태/HR
    ("근태신청현황", "/#/HP/HPD0122/HRD0220", "HR-근태"),
    ("지출결의이체현황", "/#/HP/APB1020/APB1020", "HR-지출"),
    # 예산관리 (BN)
    ("실행예산신청", "/#/BN/NCB0020/NCB0020", "BN-예산성"),
    ("실행예산마감", "/#/BN/NCB0025/NCB0025", "BN-예산성"),
    ("예산조정신청", "/#/BN/NCB0030/NCB0030", "BN-예산성"),
    ("예산조정마감", "/#/BN/NCB0035/NCB0035", "BN-예산성"),
    ("예산초기이월등록", "/#/BN/NCB0040/NCB0040", "BN-예산성"),
    ("예산마감/이월", "/#/BN/NCB0050/NCB0050", "BN-예산성"),
    ("프로젝트등록", "/#/BN/NCF0090/SYB0060", "BN-기초정보"),
    ("예산과목등록", "/#/BN/NCF0030/NCF0030", "BN-기초정보"),
    ("예산일계표", "/#/BN/NCC0230/NCC0230", "BN-보고서"),
    ("예산월계표", "/#/BN/NCC0240/NCC0240", "BN-보고서"),
    ("세출총괄표", "/#/BN/NCC0430/NCC0430", "BN-보고서"),
    ("세입총괄표", "/#/BN/NCC0440/NCC0440", "BN-보고서"),
    ("예실대비현황", "/#/BN/NCC0610/NCC0610", "BN-보고서"),
    ("예산구성비현황", "/#/BN/NCC0620/NCC0620", "BN-보고서"),
    ("예실대비현황(상세)", "/#/BN/NCC0630/NCC0630", "BN-보고서"),
    ("예실대비현황(사업별)", "/#/BN/NCC0631/NCC0631", "BN-보고서"),
    ("예산과목원장", "/#/BN/NCC0640/NCC0640", "BN-보고서"),
    # CL/RM
    ("자원예약", "/#/UK/UKA/UKA0000", "RM"),
    ("일정", "/#/UE/UEA/UEA0000", "CL"),
    # 메일
    ("메일 환경설정", "/#/UD/UDA/UDA0000", "ML"),
]

JS_PAGE_PROBE = r"""
() => ({
    title: (document.querySelector('[class*="PageTitle"], [class*="pageTitle"], h1, h2')?.innerText || '').trim().slice(0, 80),
    has_excel: !!document.querySelector('img[src*="cel_save"]'),
    has_grid: !!document.querySelector('[class*="OBTDataGrid"], [class*="RealGrid"]'),
    body_sample: (document.body.innerText || '').trim().slice(0, 200),
})
"""


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "row_counts": {}, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
            out["row_counts"][name] = ws.max_row
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def try_excel_download(page: Page, save_label: str) -> dict | None:
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            loc = fr.locator("button:has(img[src*='cel_save'])").first
            if loc.count() == 0:
                loc = fr.locator("img[src*='cel_save']").first
                if loc.count() == 0:
                    continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                with page.expect_download(timeout=40000) as dl_info:  # 40초로 늘림
                    loc.click(force=True, timeout=3000)
                    page.wait_for_timeout(1500)
                    # 확인/조회 옵션 다이얼로그 대응 (모든 frame)
                    for fr2 in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
                        try:
                            for ok in ["확인", "다운로드", "OK", "저장"]:
                                okl = fr2.locator(f"button:has-text('{ok}')").last
                                if okl.count() > 0 and okl.is_visible(timeout=400):
                                    okl.click(timeout=1500)
                                    break
                        except Exception:
                            continue
                dl = dl_info.value
                suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                path = DL_DIR / ("v8_" + safe_filename(save_label) + suffix)
                dl.save_as(str(path))
                size = path.stat().st_size if path.exists() else 0
                return {"saved": str(path), "size": size,
                        "suggested_name": dl.suggested_filename,
                        "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
            except PWTimeout:
                return {"saved": None, "error": "timeout"}
            except Exception as e:
                return {"saved": None, "error": str(e)}
        except Exception:
            continue
    return None


def click_inquiry_first(page: Page) -> bool:
    """일부 페이지는 '조회' 버튼을 먼저 눌러야 데이터 로드 → export 가능"""
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            loc = fr.locator("button:has-text('조회')").first
            if loc.count() > 0 and loc.is_visible(timeout=500):
                loc.click(force=True, timeout=2000)
                page.wait_for_timeout(3000)
                return True
        except Exception:
            continue
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--targets", type=str, default=None, help="라벨 필터(쉼표)")
    args = ap.parse_args()

    only = set(args.targets.split(",")) if args.targets else None

    results = []
    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw, headless=args.headless, user_id="tgjeon",
        )
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        for label, path, category in TARGETS:
            if only and label not in only:
                continue
            full = base + path
            logger.info(f"\n=== [{category}] {label} ===")
            logger.info(f"  goto: {path}")
            try:
                page.goto(full, wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(8000)  # 페이지 로드 충분히 대기
            except Exception as e:
                logger.warning(f"  goto 실패: {e}")
                results.append({"label": label, "path": path, "category": category, "error": str(e)})
                continue

            # 조회 버튼 먼저 클릭 (BN/HR 페이지는 조회 후 데이터 로드되어야 엑셀 활성)
            if click_inquiry_first(page):
                logger.info(f"  '조회' 클릭 → 데이터 로드 대기")
                page.wait_for_timeout(5000)

            # 페이지 진입 확인 (조회 후)
            try:
                page.screenshot(path=str(SCR_DIR / f"v8_{safe_filename(label)}.png"))
            except Exception:
                pass
            try:
                probe = page.evaluate(JS_PAGE_PROBE)
            except Exception:
                probe = {}
            logger.info(f"  title='{probe.get('title','')}' has_excel={probe.get('has_excel')} has_grid={probe.get('has_grid')}")

            # 엑셀 다운로드 시도
            dl_result = None
            if probe.get("has_excel"):
                logger.info(f"  엑셀 아이콘 존재 → 다운로드")
                # 추가 안정화 대기
                page.wait_for_timeout(2000)
                dl_result = try_excel_download(page, label)
                if dl_result and dl_result.get("saved"):
                    logger.info(f"  ✓ 저장: {Path(dl_result['saved']).name} ({dl_result['size']}B)")
                else:
                    logger.warning(f"  ✗ {dl_result}")

            results.append({
                "label": label, "path": path, "category": category,
                "url_final": page.url,
                "page_title": probe.get("title", ""),
                "has_excel": probe.get("has_excel"),
                "has_grid": probe.get("has_grid"),
                "body_sample": probe.get("body_sample", ""),
                "download": dl_result,
            })

        context.close()
        browser.close()

    OUT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON 저장: {OUT_JSON}")

    # MD
    successes = [r for r in results if r.get("download") and r["download"].get("saved")]
    lines = ["# 아마란스 v8 — URL 직접 진입 Export", ""]
    lines.append(f"- 시각: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 시도: {len(results)} / 성공: {len(successes)}")
    lines.append("")

    lines.append(f"## ✅ 다운로드 성공 ({len(successes)})")
    lines.append("| 카테고리 | 페이지 | 파일 | 크기 | 시트(행) |")
    lines.append("|---|---|---|---|---|")
    for r in successes:
        d = r["download"]
        s = d.get("summary") or {}
        sh = ", ".join(f"{n}({s.get('row_counts',{}).get(n,'?')})" for n in s.get("sheets", []))
        lines.append(f"| {r['category']} | {r['label']} | `{Path(d['saved']).name}` | {d['size']}B | {sh} |")
    lines.append("")

    lines.append("## 전체 페이지 진입 결과")
    lines.append("| 카테고리 | 페이지 | URL | 그리드 | 엑셀 | 다운 |")
    lines.append("|---|---|---|---|---|---|")
    for r in results:
        g = "✓" if r.get("has_grid") else ""
        e = "✓" if r.get("has_excel") else ""
        d = r.get("download") or {}
        ds = "✅" if d.get("saved") else ("⚠️" + (d.get("error","")[:20] if d else ""))
        lines.append(f"| {r['category']} | {r['label']} | `{r['path']}` | {g} | {e} | {ds} |")
    lines.append("")

    # 미리보기
    if successes:
        lines.append("## 다운로드 파일 미리보기\n")
        for r in successes:
            d = r["download"]
            s = d.get("summary")
            if not s or "error" in s:
                continue
            lines.append(f"### {r['category']} > {r['label']}")
            for sheet, rows in s.get("preview", {}).items():
                lines.append(f"\n**시트** `{sheet}` ({s['row_counts'].get(sheet, '?')}행)")
                if rows:
                    headers = rows[0]
                    n = min(len(headers), 10)
                    lines.append("| " + " | ".join(str(h)[:25] for h in headers[:n]) + " |")
                    lines.append("|" + "|".join(["---"] * n) + "|")
                    for row in rows[1:6]:
                        padded = list(row) + [""] * (n - len(row)) if len(row) < n else row
                        lines.append("| " + " | ".join(str(c)[:25] for c in padded[:n]) + " |")
            lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"MD 저장: {OUT_MD}")
    logger.info(f"=== 시도 {len(results)} / 다운로드 성공 {len(successes)} ===")


if __name__ == "__main__":
    main()
