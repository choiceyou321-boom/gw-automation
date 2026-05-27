"""
아마란스 export 가능 메뉴에서 샘플 파일을 다운로드하고 내용을 파악한다.
- crawl_amaranth_menus.py 결과(data/amaranth_menus_v2.json)에서
  export_candidates 가 있는 메뉴를 모두 순회한다.
- 페이지 진입 → 조회 버튼 클릭(있으면) → export 버튼 클릭 → 다운로드 캡처.
- 다운로드된 파일은 data/amaranth_exports/<module>_<page>_<label>.xlsx 형태로 저장.
- 엑셀 파일은 openpyxl 로 첫 시트 + 헤더 + 5행 샘플을 콘솔/Markdown 으로 요약.

실행:
    .venv/bin/python scripts/download_amaranth_sample.py [--limit N] [--label "예실대비"]
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from playwright.sync_api import sync_playwright, Page, Download

from src.shared.auth.login import login_and_get_context

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("download_amaranth_sample")

ROOT = Path(__file__).resolve().parent.parent
IN_JSON = ROOT / "data" / "amaranth_menus_v2.json"
OUT_DIR = ROOT / "data" / "amaranth_exports"
OUT_SUMMARY = ROOT / "docs" / "AMARANTH_EXPORTS_SUMMARY.md"


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def try_click_inquiry(page: Page) -> bool:
    """조회/검색 버튼이 있으면 먼저 클릭 (데이터를 로드해야 export 가능한 경우 대비)"""
    for text in ["조회", "검색", "Search"]:
        try:
            loc = page.locator(f"button:has-text('{text}')").first
            if loc.count() > 0 and loc.is_visible():
                loc.click(timeout=2000)
                page.wait_for_timeout(2500)
                logger.info(f"  '{text}' 클릭 완료")
                return True
        except Exception:
            continue
    return False


def try_download(page: Page, button_text: str, timeout_ms: int = 15000) -> Download | None:
    """export 버튼 클릭 → 다운로드 캡처. 팝업/모달이 뜨면 확인 클릭 시도."""
    try:
        with page.expect_download(timeout=timeout_ms) as dl_info:
            # 정확히 일치하는 버튼 우선
            loc = page.locator(f"button:has-text('{button_text}')").first
            if loc.count() == 0:
                loc = page.locator(f"[title*='{button_text}']").first
            if loc.count() == 0:
                logger.warning(f"  버튼 '{button_text}' 찾지 못함")
                return None
            loc.click(force=True, timeout=3000)
            # 일부 메뉴는 다운로드 옵션 다이얼로그가 뜸. '확인' 클릭 시도.
            try:
                for ok_text in ["확인", "다운로드", "저장", "OK"]:
                    ok_loc = page.locator(f"button:has-text('{ok_text}')").last
                    if ok_loc.count() > 0 and ok_loc.is_visible():
                        ok_loc.click(timeout=2000)
                        break
            except Exception:
                pass
        return dl_info.value
    except Exception as e:
        logger.warning(f"  다운로드 실패 ({button_text}): {e}")
        return None


def summarize_xlsx(path: Path) -> dict:
    """엑셀 파일 첫 시트 헤더 + 5행 미리보기"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl 미설치"}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        out = {"sheets": sheets, "preview": {}}
        for name in sheets[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=10)
    ap.add_argument("--label", type=str, default=None, help="라벨 부분일치 필터")
    ap.add_argument("--headless", action="store_true")
    args = ap.parse_args()

    if not IN_JSON.exists():
        logger.error(f"먼저 crawl_amaranth_menus.py 를 실행하세요: {IN_JSON} 없음")
        return

    data = json.loads(IN_JSON.read_text(encoding="utf-8"))
    candidates = [m for m in data["menus"] if m.get("export_candidates")]
    if args.label:
        candidates = [m for m in candidates if args.label in (m.get("label") or "")]
    logger.info(f"Export 후보 메뉴: {len(candidates)}개 (limit={args.limit})")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    results = []

    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw,
            headless=args.headless,
            user_id="tgjeon",
        )

        m = re.match(r"(https://[^/]+)", page.url)
        base_url = m.group(1) if m else "https://gw.glowseoul.co.kr"

        for idx, menu in enumerate(candidates[: args.limit]):
            label = menu.get("label") or "?"
            url = menu.get("url") or ""
            full = url if url.startswith("http") else f"{base_url}/{url.lstrip('/')}"
            logger.info(f"[{idx+1}/{min(len(candidates), args.limit)}] {label} → {url}")
            try:
                page.goto(full, wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(2500)
            except Exception as e:
                logger.warning(f"  goto 실패: {e}")
                continue

            try_click_inquiry(page)

            # export_candidates 의 텍스트 후보 중 "엑셀" 우선
            exports = menu["export_candidates"]
            priorities = ["엑셀", "엑셀다운", "엑셀저장", "엑셀출력", "Excel", "다운로드", "내보내기", "CSV", "PDF"]
            sorted_exports = sorted(
                exports,
                key=lambda e: next((i for i, p in enumerate(priorities) if p in e["text"]), 999),
            )

            saved_path = None
            tried = []
            for ex in sorted_exports[:3]:
                btn = ex["text"]
                tried.append(btn)
                dl = try_download(page, btn)
                if not dl:
                    continue
                fname = safe_filename(
                    f"{menu.get('module','?')}_{menu.get('page_code','?')}_{label}"
                ) + Path(dl.suggested_filename).suffix
                saved_path = OUT_DIR / fname
                try:
                    dl.save_as(str(saved_path))
                    logger.info(f"  ★ 저장: {saved_path.name} ({saved_path.stat().st_size}B)")
                    break
                except Exception as e:
                    logger.warning(f"  저장 실패: {e}")
                    saved_path = None

            entry = {
                "label": label,
                "url": url,
                "module": menu.get("module"),
                "page_code": menu.get("page_code"),
                "tried_buttons": tried,
                "saved": str(saved_path) if saved_path else None,
            }
            if saved_path and saved_path.suffix.lower() in (".xlsx", ".xls"):
                entry["summary"] = summarize_xlsx(saved_path)
            results.append(entry)

        context.close()
        browser.close()

    # Markdown 요약
    lines = ["# 아마란스 Export 샘플 다운로드 결과", ""]
    lines.append(f"- 실행 시각: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 시도한 메뉴: {len(results)}")
    saved = [r for r in results if r["saved"]]
    lines.append(f"- 다운로드 성공: {len(saved)}")
    lines.append("")
    for r in results:
        lines.append(f"## {r['label']}")
        lines.append(f"- URL: `{r['url']}`")
        lines.append(f"- 시도 버튼: {r['tried_buttons']}")
        lines.append(f"- 저장: {r['saved'] or '실패'}")
        s = r.get("summary")
        if s and "error" not in s:
            for sheet, rows in s.get("preview", {}).items():
                lines.append(f"\n**시트**: {sheet}")
                if rows:
                    headers = rows[0]
                    lines.append("| " + " | ".join(headers) + " |")
                    lines.append("|" + "|".join(["---"] * len(headers)) + "|")
                    for row in rows[1:]:
                        # 길이 맞춤
                        padded = row + [""] * (len(headers) - len(row))
                        lines.append("| " + " | ".join(padded[: len(headers)]) + " |")
        elif s and "error" in s:
            lines.append(f"- summary 오류: {s['error']}")
        lines.append("")

    OUT_SUMMARY.parent.mkdir(exist_ok=True)
    OUT_SUMMARY.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"요약: {OUT_SUMMARY}")


if __name__ == "__main__":
    main()
