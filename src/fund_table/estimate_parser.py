"""내역서(견적서/산출내역서) 엑셀에서 공종 자동 추출

사용자가 업로드한 내역서 엑셀 파일을 분석하여
공정표 마스터 데이터(PROCESS_GROUPS)와 매칭되는 공종 목록을 반환한다.
"""
import logging
from difflib import SequenceMatcher

from openpyxl import load_workbook

from src.fund_table.process_map_master import get_all_trade_names

logger = logging.getLogger(__name__)

# 내역서 헤더로 자주 사용되는 키워드
_HEADER_KEYWORDS = {"공종", "품명", "항목", "내역", "공사명", "구분", "분류", "세부공종", "공종명"}

# 공종 매칭용 별칭 (내역서에서 자주 쓰는 다른 이름 → 마스터 공종명)
_ALIASES = {
    "금속공사": "METAL STUD",
    "메탈스터드": "METAL STUD",
    "경량철골": "METAL STUD",
    "경량칸막이": "METAL STUD",
    "철골구조": "ST'L PIPE 구조틀",
    "조적공사": "조적",
    "방수공사": "방수미장",
    "미장공사": "방수미장",
    "도장공사": "벽체 도장",
    "도배공사": "도배",
    "타일공사": "타일/석재 (바닥)",
    "석재공사": "벽체 타일/석재",
    "목공사": "목문틀",
    "천정공사": "천정골조 (T-Bar, M-Bar)",
    "천장공사": "천정골조 (T-Bar, M-Bar)",
    "석고보드": "석고보드/합판 취부",
    "단열공사": "단열재 취부",
    "가구공사": "붙박이가구",
    "가구설치": "붙박이가구",
    "창호공사": "창호",
    "유리공사": "유리",
    "주방설비": "주방기구 설치",
    "주방공사": "주방기구 설치",
    "위생설비": "위생기구/액세서리",
    "카펫공사": "카펫",
    "마루공사": "마루",
    "마루바닥": "마루",
    "OA플로어": "악세스플로어",
    "엑세스플로어": "악세스플로어",
    "비닐시트": "P-TILE/비닐쉬트",
    "비닐쉬트": "P-TILE/비닐쉬트",
    "데크공사": "목재데크",
    "그리스트랩": "그리스트랩/트랜치",
    "트랜치": "그리스트랩/트랜치",
    "클린룸": "크린룸 공사",
    "크린룸": "크린룸 공사",
    "노출콘크리트": "노출마감",
    "롤블라인드": "롤스크린 설치",
    "롤스크린": "롤스크린 설치",
    "준공청소": "준공청소",
    "큐비클": "큐비클/유리시공",
}


def parse_estimate_file(file_path: str) -> dict:
    """내역서 엑셀에서 공종 목록 자동 추출

    Args:
        file_path: 내역서 엑셀 파일 경로

    Returns:
        {
            "matched_trades": ["METAL STUD", "조적", ...],
            "unmatched": ["기타공사A", ...],
            "raw_items": ["금속공사", "조적공사", ...],
        }
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"내역서 파일 열기 실패: {e}")
        return {"matched_trades": [], "unmatched": [], "raw_items": [], "error": str(e)}

    raw_items = set()
    try:
        for ws in wb.worksheets:
            trade_col = _find_trade_column(ws)
            if trade_col is None:
                trade_col = 1

            # max_row 비정상 방지 (100만행 등)
            max_row = min(ws.max_row or 1000, 10000)
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=trade_col, max_col=trade_col):
                cell = row[0]
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if len(val) >= 2 and not val.startswith(("합계", "소계", "총", "부가", "VAT")):
                        raw_items.add(val)
    finally:
        wb.close()

    raw_list = sorted(raw_items)
    all_trades = get_all_trade_names()
    matched = set()
    unmatched = []

    for raw in raw_list:
        trade = _match_trade(raw, all_trades)
        if trade:
            matched.add(trade)
        else:
            unmatched.append(raw)

    return {
        "matched_trades": sorted(matched),
        "unmatched": unmatched,
        "raw_items": raw_list,
    }


def _find_trade_column(ws) -> int | None:
    """첫 번째 행에서 공종 관련 헤더 컬럼 인덱스 탐색"""
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip() in _HEADER_KEYWORDS:
                    return cell.column
    return None


def _match_trade(raw_name: str, all_trades: list[str]) -> str | None:
    """내역서 항목명 → 마스터 공종명 매칭"""
    normalized = raw_name.strip().replace(" ", "")

    # 1) 별칭 정확 매칭
    if normalized in _ALIASES:
        return _ALIASES[normalized]
    for alias, trade in _ALIASES.items():
        # 짧은 문자열의 부분 매칭 방지 (최소 3글자 이상만 부분매칭)
        if len(normalized) >= 3 and alias in normalized:
            return trade
        if len(alias) >= 3 and normalized in alias:
            return trade

    # 2) 마스터 공종명에 포함 관계 (최소 3글자)
    for trade in all_trades:
        trade_norm = trade.replace(" ", "")
        if len(normalized) >= 3 and trade_norm in normalized:
            return trade
        if len(trade_norm) >= 3 and normalized in trade_norm:
            return trade

    # 3) 유사도 매칭 (0.7 이상)
    best_score = 0.0
    best_trade = None
    for trade in all_trades:
        score = SequenceMatcher(None, normalized, trade.replace(" ", "")).ratio()
        if score > best_score:
            best_score = score
            best_trade = trade

    if best_score >= 0.7:
        return best_trade

    return None
