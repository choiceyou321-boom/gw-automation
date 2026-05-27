"""
아마란스 v6 전체 export 크롤러
─────────────────────────────────────────
v5의 두 가지 문제 수정:
  1. 모듈 진입: span.module-link 단일 셀렉터로는 6개만 진입 → 다중 클릭 전략
     - 1차: span.module-link.{code} 클릭
     - 2차: dock 펼친 상태에서 dock 아이콘 좌표 클릭
     - 3차: title 속성으로 클릭
  2. LNB 수집: 좌측 사이드바(module-item)를 LNB로 오인 → 사이드바 명시 제외
     - x > 60 영역만 LNB로 인정 (사이드바는 x<50)
     - 콘텐츠 영역의 트리 메뉴만 수집

leaf 메뉴 클릭 → 페이지 진입 → cel_save 아이콘 자동 다운로드
다운로드 파일 헤더+5행 자동 미리보기

결과:
  data/amaranth_v6.json
  docs/AMARANTH_EXPORT_v6.md
  data/amaranth_exports/v6_*.xlsx
"""
from __future__ import annotations
import argparse, json, logging, re, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("v6")

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "data" / "amaranth_v6.json"
OUT_MD = ROOT / "docs" / "AMARANTH_EXPORT_v6.md"
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)

HOME_MODULES = [
    ("SET", "시스템설정"), ("HR", "임직원업무관리"), ("EA", "전자결재"),
    ("ML", "메일"), ("CL", "일정"), ("RM", "자원"),
    ("BD", "게시판"), ("KS", "업무관리"), ("OF", "ONEFFICE"),
    ("OC", "ONECHAMBER"), ("BPM", "프로세스관리"), ("UT", "오피스케어"),
]

SKIP_LNB_TEXTS = {
    "해당 탭 닫기", "해당 탭 제외 다른 탭 닫기", "모든 탭 닫기",
    "탭 닫기", "닫기", "새로고침", "위로", "아래로",
}

# LNB 수집 — 콘텐츠 영역만 (x > 60)
JS_CONTENT_LNB = r"""
(skipTexts) => {
    const out = [];
    const seen = new Set();
    // 사이드바(x<60) 제외 + 좌측 LNB 패널(보통 60 < x < 300)
    const sels = [
        '[class*="OBTLeftMenu"] li',
        '[class*="LeftMenu"] li',
        '[class*="LNB"] li',
        '[class*="lnb"] li',
        '[class*="TreeMenu"] li',
        '[role="treeitem"]',
        '[role="menuitem"]',
    ];
    sels.forEach(sel => {
        document.querySelectorAll(sel).forEach(el => {
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) return;
            if (r.x < 60 || r.x > 350) return;  // 사이드바 제외, LNB 영역 한정
            const text = (el.innerText || el.textContent || '').trim().split('\n')[0];
            if (!text || text.length > 50) return;
            if (skipTexts.includes(text)) return;
            const key = text + '|' + Math.round(r.x);
            if (seen.has(key)) return;
            seen.add(key);
            // 자식 ul 존재 여부 = 카테고리 vs leaf
            const hasChildUl = !!el.querySelector('ul li');
            out.push({
                text, x: Math.round(r.x), y: Math.round(r.y),
                w: Math.round(r.width), h: Math.round(r.height),
                cls: (el.className || '').toString().slice(0, 80),
                has_child: hasChildUl,
            });
        });
    });
    return out;
}
"""

JS_PAGE_AFTER_CLICK = r"""
() => {
    const out = {
        title: '', has_excel_btn: false, has_grid: false,
        button_texts: [], excel_btn_positions: [],
    };
    const t = document.querySelector('[class*="PageTitle"], [class*="pageTitle"], h1, h2');
    out.title = t ? (t.innerText || '').trim().slice(0, 80) : '';
    // 엑셀 아이콘
    document.querySelectorAll('img[src*="cel_save"]').forEach(img => {
        const r = img.getBoundingClientRect();
        if (r.width === 0) return;
        out.excel_btn_positions.push({
            x: Math.round(r.x + r.width/2), y: Math.round(r.y + r.height/2),
            src: img.getAttribute('src') || '',
        });
    });
    out.has_excel_btn = out.excel_btn_positions.length > 0;
    out.has_grid = !!document.querySelector('[class*="OBTDataGrid"], [class*="RealGrid"]');
    // 버튼 텍스트 샘플
    const seen = new Set();
    document.querySelectorAll('button').forEach(b => {
        const txt = (b.innerText || '').trim();
        if (!txt || txt.length > 15 || seen.has(txt)) return;
        seen.add(txt);
        if (out.button_texts.length < 20) out.button_texts.push(txt);
    });
    return out;
}
"""


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "row_counts": {}, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
            out["row_counts"][name] = ws.max_row
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def enter_module(page: Page, code: str, mod_text: str) -> str:
    """모듈 진입 — 다중 전략. 진입한 URL 반환 (실패 시 빈 문자열)"""
    base_url = page.url.split("#/")[0]
    page.goto(base_url + "#/", wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(3000)

    # 전략 1: span.module-link.{code} 직접 클릭
    try:
        loc = page.locator(f"span.module-link.{code}").first
        if loc.count() > 0 and loc.is_visible(timeout=2000):
            loc.click(force=True, timeout=3000)
            page.wait_for_timeout(5500)
            url = max([fr.url for fr in page.frames if "#/" in (fr.url or "")], key=len, default=page.url)
            if "#/" in url and url.split("#/")[-1].strip("/") != "":
                logger.info(f"  ✓ 진입 방식1 (module-link.{code})")
                return url
    except Exception as e:
        logger.debug(f"  방식1 실패: {e}")

    # 전략 2: title 속성으로 클릭
    try:
        loc = page.locator(f"[title='{mod_text}']").first
        if loc.count() > 0 and loc.is_visible(timeout=1000):
            loc.click(force=True, timeout=3000)
            page.wait_for_timeout(5500)
            url = max([fr.url for fr in page.frames if "#/" in (fr.url or "")], key=len, default=page.url)
            if "#/" in url and url.split("#/")[-1].strip("/") != "":
                logger.info(f"  ✓ 진입 방식2 (title='{mod_text}')")
                return url
    except Exception:
        pass

    # 전략 3: text 일치
    try:
        loc = page.locator(f"text='{mod_text}'").first
        if loc.count() > 0 and loc.is_visible(timeout=1000):
            loc.click(force=True, timeout=3000)
            page.wait_for_timeout(5500)
            url = max([fr.url for fr in page.frames if "#/" in (fr.url or "")], key=len, default=page.url)
            if "#/" in url and url.split("#/")[-1].strip("/") != "":
                logger.info(f"  ✓ 진입 방식3 (text)")
                return url
    except Exception:
        pass

    # 전략 4: dock 펼친 상태의 module-link 모든 후보 다 시도
    try:
        all_links = page.locator(f"span.module-link.{code}").all()
        for i, ll in enumerate(all_links):
            try:
                if ll.is_visible(timeout=300):
                    ll.click(force=True, timeout=2000)
                    page.wait_for_timeout(5000)
                    url = max([fr.url for fr in page.frames if "#/" in (fr.url or "")], key=len, default=page.url)
                    if "#/" in url and url.split("#/")[-1].strip("/") != "":
                        logger.info(f"  ✓ 진입 방식4 ({i}번째 module-link.{code})")
                        return url
            except Exception:
                continue
    except Exception:
        pass

    logger.warning(f"  ✗ {code} 모든 진입 방식 실패")
    return ""


def click_lnb_text(page: Page, text: str, timeout_ms: int = 3000) -> bool:
    """LNB 텍스트 클릭 (page + iframe, 사이드바 제외)"""
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        for sel in [f"li:has-text('{text}')", f"text='{text}'"]:
            try:
                loc = fr.locator(sel).first
                if loc.count() == 0:
                    continue
                box = loc.bounding_box()
                if not box or box["x"] < 60:  # 사이드바 영역 클릭 방지
                    continue
                if not loc.is_visible(timeout=400):
                    continue
                loc.click(force=True, timeout=timeout_ms)
                return True
            except Exception:
                continue
    return False


def try_excel_download(page: Page, save_label: str, timeout_ms: int = 15000) -> dict | None:
    """cel_save 이미지의 부모 button 클릭 → 다운로드 캡처"""
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            loc = fr.locator("button:has(img[src*='cel_save'])").first
            if loc.count() == 0:
                loc = fr.locator("img[src*='cel_save']").first
                if loc.count() == 0:
                    continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                with page.expect_download(timeout=timeout_ms) as dl_info:
                    loc.click(force=True, timeout=3000)
                    # 확인 모달
                    try:
                        for ok in ["확인", "다운로드", "OK"]:
                            okl = fr.locator(f"button:has-text('{ok}')").last
                            if okl.count() > 0 and okl.is_visible(timeout=400):
                                okl.click(timeout=1500)
                                break
                    except Exception:
                        pass
                dl = dl_info.value
                suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                path = DL_DIR / ("v6_" + safe_filename(save_label) + suffix)
                dl.save_as(str(path))
                size = path.stat().st_size if path.exists() else 0
                return {"saved": str(path), "size": size,
                        "suggested_name": dl.suggested_filename,
                        "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
            except PWTimeout:
                return {"saved": None, "error": "timeout"}
            except Exception as e:
                return {"saved": None, "error": str(e)}
        except Exception:
            continue
    return None


def collect_from_all_frames(page: Page, js: str, *args) -> list:
    out = []
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            r = fr.evaluate(js, *args) if args else fr.evaluate(js)
            if isinstance(r, list):
                out.extend(r)
            elif isinstance(r, dict):
                out.append(r)
        except Exception:
            continue
    return out


def current_url(page: Page) -> str:
    cands = [fr.url for fr in page.frames if fr.url and "#/" in fr.url]
    return max(cands, key=len) if cands else page.url


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--modules", type=str, default=None)
    ap.add_argument("--leaf_limit", type=int, default=40)
    ap.add_argument("--download_limit", type=int, default=50)
    args = ap.parse_args()

    only = set(args.modules.split(",")) if args.modules else None
    result = {
        "gw_user": "tgjeon",
        "crawled_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "modules": [],
        "errors": [],
    }
    downloads_done = 0

    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw, headless=args.headless, user_id="tgjeon",
        )
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        for code, mod_text in HOME_MODULES:
            if only and code not in only:
                continue
            logger.info(f"\n{'='*60}\n[{code}] {mod_text}\n{'='*60}")

            entry_url = enter_module(page, code, mod_text)
            mod_entry = {"code": code, "text": mod_text, "entry_url": entry_url, "lnb": [], "leaf_visits": []}

            if not entry_url:
                result["modules"].append(mod_entry)
                continue

            try:
                page.screenshot(path=str(SCR_DIR / f"v6_mod_{code}.png"))
            except Exception:
                pass

            # LNB 수집 (콘텐츠 영역만, 사이드바 제외)
            try:
                lnb_items = collect_from_all_frames(page, JS_CONTENT_LNB, list(SKIP_LNB_TEXTS))
                dedup = {}
                for it in lnb_items:
                    t = it.get("text")
                    if not t or t in dedup:
                        continue
                    dedup[t] = it
                lnb_unique = list(dedup.values())
                mod_entry["lnb"] = lnb_unique
                logger.info(f"  LNB 항목 (콘텐츠 영역): {len(lnb_unique)}")
            except Exception as e:
                logger.warning(f"  LNB 수집 실패: {e}")
                lnb_unique = []

            # 모든 카테고리 한번 클릭으로 펼침 (toggle 트리)
            for it in lnb_unique[:10]:
                if it.get("has_child"):
                    try:
                        click_lnb_text(page, it["text"], timeout_ms=1500)
                        page.wait_for_timeout(500)
                    except Exception:
                        pass

            # 다시 수집 (펼친 후)
            try:
                lnb_after = collect_from_all_frames(page, JS_CONTENT_LNB, list(SKIP_LNB_TEXTS))
                dedup = {}
                for it in lnb_after:
                    t = it.get("text")
                    if not t or t in dedup:
                        continue
                    dedup[t] = it
                lnb_unique = list(dedup.values())
                logger.info(f"  펼침 후 LNB: {len(lnb_unique)}")
            except Exception:
                pass

            # 진입 직후 페이지에서도 엑셀 시도 (모듈 메인 페이지)
            try:
                main_info = collect_from_all_frames(page, JS_PAGE_AFTER_CLICK)
                has_main_excel = any(i.get("has_excel_btn") for i in main_info)
                if has_main_excel and downloads_done < args.download_limit:
                    logger.info(f"  ★ 모듈 메인 페이지에 엑셀 있음 → 다운로드 시도")
                    dl = try_excel_download(page, f"{code}_MAIN")
                    if dl and dl.get("saved"):
                        downloads_done += 1
                        logger.info(f"  ✓ 메인 다운로드: {Path(dl['saved']).name} ({dl['size']}B)")
                        mod_entry["main_download"] = dl
            except Exception:
                pass

            # leaf 클릭 순회
            for idx, leaf in enumerate(lnb_unique[: args.leaf_limit]):
                lnb_text = leaf["text"]
                logger.info(f"  [{idx+1}/{min(len(lnb_unique), args.leaf_limit)}] {lnb_text}")
                ok = click_lnb_text(page, lnb_text, timeout_ms=2500)
                if not ok:
                    continue
                page.wait_for_timeout(2500)
                # 페이지 정보 수집
                try:
                    info_list = collect_from_all_frames(page, JS_PAGE_AFTER_CLICK)
                except Exception:
                    continue
                title = ""
                has_excel = False
                has_grid = False
                for inf in info_list:
                    if not title and inf.get("title"):
                        title = inf["title"]
                    has_excel = has_excel or inf.get("has_excel_btn")
                    has_grid = has_grid or inf.get("has_grid")

                url = current_url(page)
                visit = {
                    "lnb_text": lnb_text, "url": url, "page_title": title,
                    "has_excel_btn": has_excel, "has_grid": has_grid,
                    "download": None,
                }

                if has_excel and downloads_done < args.download_limit:
                    logger.info(f"    ★ 엑셀 발견 → 다운로드 시도")
                    save_label = f"{code}_{safe_filename(lnb_text)}"
                    dl = try_excel_download(page, save_label)
                    visit["download"] = dl
                    if dl and dl.get("saved"):
                        downloads_done += 1
                        logger.info(f"    ✓ 저장: {Path(dl['saved']).name} ({dl['size']}B)")
                    else:
                        logger.warning(f"    ✗ {dl}")

                mod_entry["leaf_visits"].append(visit)

            result["modules"].append(mod_entry)

        context.close()
        browser.close()

    # 저장
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON 저장: {OUT_JSON}")

    # Markdown 요약
    lines = ["# 아마란스 v6 전체 Export 결과", ""]
    lines.append(f"- 계정: tgjeon / 시각: {result['crawled_at']}")
    lines.append(f"- 모듈: {len(result['modules'])} / 다운로드 성공: {downloads_done}")
    lines.append("")

    # 성공 목록
    successes = []
    for me in result["modules"]:
        # 메인 페이지 다운로드
        if me.get("main_download") and me["main_download"].get("saved"):
            successes.append((me["code"], me["text"], "(모듈 메인)", me["main_download"]))
        for v in me["leaf_visits"]:
            if v.get("download") and v["download"].get("saved"):
                successes.append((me["code"], me["text"], v["lnb_text"], v["download"]))
    lines.append(f"## ✅ 다운로드 성공 ({len(successes)}개)\n")
    lines.append("| 모듈 | 메뉴 | 파일 | 크기 | 시트 |")
    lines.append("|---|---|---|---|---|")
    for code, mod_text, leaf, d in successes:
        sheets = (d.get("summary") or {}).get("sheets", [])
        lines.append(f"| {code} {mod_text} | {leaf} | `{Path(d['saved']).name}` | {d['size']}B | {','.join(sheets)} |")
    lines.append("")

    # 모듈별 진입 결과
    lines.append("## 모듈별 진입 및 LNB")
    for me in result["modules"]:
        lines.append(f"\n### [{me['code']}] {me['text']}")
        lines.append(f"- 진입 URL: `{me.get('entry_url','(실패)')[-80:]}`")
        lines.append(f"- LNB 항목: {len(me['lnb'])} / leaf 방문: {len(me['leaf_visits'])}")
        if me["leaf_visits"]:
            lines.append("")
            lines.append("| Leaf | URL | 그리드 | 엑셀 | 다운 |")
            lines.append("|---|---|---|---|---|")
            for v in me["leaf_visits"]:
                g = "✓" if v["has_grid"] else ""
                e = "✓" if v["has_excel_btn"] else ""
                d = v.get("download") or {}
                ds = "✅" if d.get("saved") else ("⚠️" if d else "")
                lines.append(f"| {v['lnb_text']} | `{(v['url'] or '')[-50:]}` | {g} | {e} | {ds} |")

    # 미리보기
    if successes:
        lines.append("\n## 다운로드 파일 미리보기\n")
        for code, mod_text, leaf, d in successes:
            s = d.get("summary")
            if not s or "error" in s:
                continue
            lines.append(f"### {code} > {leaf}")
            for sheet, rows in s.get("preview", {}).items():
                lines.append(f"\n**시트** `{sheet}` ({s['row_counts'].get(sheet, '?')}행)")
                if rows:
                    headers = rows[0]
                    n = min(len(headers), 10)
                    lines.append("| " + " | ".join(str(h)[:25] for h in headers[:n]) + " |")
                    lines.append("|" + "|".join(["---"] * n) + "|")
                    for row in rows[1:6]:
                        padded = list(row) + [""] * (n - len(row)) if len(row) < n else row
                        lines.append("| " + " | ".join(str(c)[:25] for c in padded[:n]) + " |")
            lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"MD 저장: {OUT_MD}")
    logger.info(f"=== 모듈 {len(result['modules'])} / 다운로드 {downloads_done} ===")


if __name__ == "__main__":
    main()
