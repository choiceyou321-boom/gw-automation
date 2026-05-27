"""
아마란스 전체 구조 자동 크롤링 v5 — 모든 모듈 × 모든 leaf 메뉴 + export 일괄 수집
─────────────────────────────────────────────────────────────────────────
확보한 핵심 셀렉터: 엑셀 아이콘 = button:has(img[src*='cel_save'])

전략:
  1) 12 모듈 순회 (span.module-link.{code})
  2) 각 모듈 진입 후 LNB 모든 카테고리 펼치기 (1~3-depth 재귀)
  3) Leaf 메뉴(클릭하면 페이지 변경) 자동 식별
  4) 각 leaf 클릭 → 페이지 진입 → URL/제목/버튼 수집 + 엑셀 아이콘 있으면 다운로드
  5) 다운로드 파일 헤더+5행 자동 요약

결과:
  data/amaranth_full_v5.json
  docs/AMARANTH_FULL_STRUCTURE.md
  data/amaranth_exports/{module}_{label}.xlsx
"""
from __future__ import annotations
import argparse, json, logging, re, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("v5")

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "data" / "amaranth_full_v5.json"
OUT_MD = ROOT / "docs" / "AMARANTH_FULL_STRUCTURE.md"
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)

# 우클릭 메뉴 등 LNB 아닌 텍스트
SKIP_TEXTS = {
    "해당 탭 닫기", "해당 탭 제외 다른 탭 닫기", "모든 탭 닫기",
    "탭 닫기", "닫기", "새로고침",
}

# LNB 모든 노드 + 펼침 상태 + leaf 여부 수집
JS_LNB_TREE = r"""
(skipTexts) => {
    const out = [];
    const seen = new Set();
    document.querySelectorAll('li, [role="treeitem"]').forEach(el => {
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return;
        // 좌측 패널 영역만 (x < 300)
        if (r.x > 300) return;
        // 자체 텍스트 또는 첫 자식 텍스트
        const direct = Array.from(el.childNodes)
            .filter(n => n.nodeType === 3).map(n => n.textContent.trim())
            .filter(Boolean).join(' ');
        const full = (el.innerText || el.textContent || '').trim().split('\n')[0];
        const text = direct || full;
        if (!text || text.length > 60) return;
        if (skipTexts.includes(text)) return;
        const cls = (el.className || '').toString();
        // leaf 판단: 자식 ul/li가 없거나 + 아이콘이 없거나 → 휴리스틱
        const hasChildList = !!el.querySelector('ul li, [role="treeitem"]');
        const hasExpandIcon = cls.match(/expand|plus|toggle|chevron/) || !!el.querySelector('[class*="expand"], [class*="plus"], [class*="chevron"]');
        const key = text + '|' + Math.round(r.x) + ',' + Math.round(r.y);
        if (seen.has(key)) return;
        seen.add(key);
        out.push({
            text, cls: cls.slice(0, 100),
            x: Math.round(r.x), y: Math.round(r.y),
            w: Math.round(r.width), h: Math.round(r.height),
            has_child_list: hasChildList,
            has_expand_icon: !!hasExpandIcon,
        });
    });
    return out;
}
"""

JS_PAGE_INFO = r"""
() => {
    const out = { title: '', has_excel_btn: false, buttons: [], excel_btn_count: 0 };
    const t = document.querySelector('[class*="PageTitle"], [class*="pageTitle"], h1, h2');
    out.title = t ? (t.innerText || '').trim().slice(0, 80) : '';
    // 엑셀 아이콘 탐지 (확보한 셀렉터)
    const excelImgs = document.querySelectorAll('img[src*="cel_save"]');
    out.excel_btn_count = excelImgs.length;
    out.has_excel_btn = excelImgs.length > 0;
    // 그리드 존재 여부
    out.has_grid = !!document.querySelector('[class*="OBTDataGrid"], [class*="RealGrid"]');
    // 일반 버튼 텍스트 샘플
    const btns = document.querySelectorAll('button');
    const seen = new Set();
    btns.forEach(b => {
        const txt = (b.innerText || '').trim();
        if (!txt || txt.length > 20 || seen.has(txt)) return;
        const r = b.getBoundingClientRect();
        if (r.width === 0) return;
        seen.add(txt);
        if (out.buttons.length < 15) out.buttons.push(txt);
    });
    return out;
}
"""


def click_text_in_any_frame(page: Page, text: str, timeout_ms: int = 3000) -> bool:
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        for sel in [f"text='{text}'", f"li:has-text('{text}')"]:
            try:
                loc = fr.locator(sel).first
                if loc.count() > 0 and loc.is_visible(timeout=400):
                    loc.click(force=True, timeout=timeout_ms)
                    return True
            except Exception:
                continue
    return False


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "row_count": {}, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 6:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
            out["row_count"][name] = ws.max_row
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def try_excel_download(page: Page, save_label: str) -> dict | None:
    """cel_save 아이콘 클릭으로 다운로드 시도"""
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        try:
            loc = fr.locator("button:has(img[src*='cel_save'])").first
            if loc.count() == 0:
                # 폴백: img 직접
                loc = fr.locator("img[src*='cel_save']").first
                if loc.count() == 0:
                    continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                with page.expect_download(timeout=12000) as dl_info:
                    loc.click(force=True, timeout=3000)
                    # 확인 모달
                    try:
                        for ok in ["확인", "다운로드", "OK"]:
                            okl = fr.locator(f"button:has-text('{ok}')").last
                            if okl.count() > 0 and okl.is_visible(timeout=400):
                                okl.click(timeout=1500)
                                break
                    except Exception:
                        pass
                dl = dl_info.value
                suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                path = DL_DIR / (safe_filename(save_label) + suffix)
                dl.save_as(str(path))
                size = path.stat().st_size if path.exists() else 0
                return {"saved": str(path), "size": size,
                        "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
            except PWTimeout:
                return {"saved": None, "error": "timeout"}
            except Exception as e:
                return {"saved": None, "error": str(e)}
        except Exception:
            continue
    return None  # 엑셀 버튼 자체가 없음


def current_url(page: Page) -> str:
    """가장 깊은 #/ 경로 추출"""
    cands = [fr.url for fr in page.frames if fr.url and "#/" in fr.url]
    return max(cands, key=len) if cands else page.url


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--modules", type=str, default=None, help="모듈 코드 필터 (쉼표)")
    ap.add_argument("--leaf_limit", type=int, default=40)
    ap.add_argument("--download_limit", type=int, default=30)
    args = ap.parse_args()

    only = set(args.modules.split(",")) if args.modules else None

    result = {
        "gw_user": "tgjeon",
        "crawled_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "modules": [],
        "errors": [],
    }
    downloads_done = 0

    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw, headless=args.headless, user_id="tgjeon",
        )
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        # 홈 진입 → 모듈 12개 수집
        page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(4000)
        modules = page.evaluate(r"""
            () => {
                const out = [];
                const seen = new Set();
                document.querySelectorAll('span.module-link, [class*="module-link"]').forEach(el => {
                    const cls = (el.className || '').toString();
                    const m = cls.match(/module-link\s+([A-Z]{2,4})/);
                    if (!m || seen.has(m[1])) return;
                    seen.add(m[1]);
                    out.push({ code: m[1], text: (el.innerText || el.textContent || '').trim() });
                });
                return out;
            }
        """)
        logger.info(f"홈 모듈 {len(modules)}개")

        for mod in modules:
            code = mod["code"]
            if only and code not in only:
                continue
            mod_text = mod.get("text") or code
            logger.info(f"\n{'='*60}\n[{code}] {mod_text}\n{'='*60}")

            # 홈 → 모듈 클릭
            page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(2500)
            try:
                page.locator(f"span.module-link.{code}").first.click(force=True, timeout=5000)
                page.wait_for_timeout(5000)
            except Exception as e:
                logger.warning(f"  진입 실패: {e}")
                result["errors"].append({"phase": "module_click", "code": code})
                continue

            mod_entry = {"code": code, "text": mod_text, "lnb_items": [], "leaf_visits": []}

            # LNB 전체 펼침 시도 — 1-depth 카테고리 모두 한 번씩 클릭 (토글 펼침)
            try:
                tree = page.evaluate(JS_LNB_TREE, list(SKIP_TEXTS))
                logger.info(f"  초기 LNB 노드: {len(tree)}")
            except Exception:
                tree = []

            # 모든 카테고리 항목 클릭(펼침). 폴더처럼 보이는 항목(has_child_list 또는 has_expand_icon)을 우선
            for item in tree[:20]:
                if item.get("has_child_list") or item.get("has_expand_icon"):
                    try:
                        click_text_in_any_frame(page, item["text"], timeout_ms=1500)
                        page.wait_for_timeout(600)
                    except Exception:
                        pass

            # 펼친 후 다시 수집
            try:
                tree2 = page.evaluate(JS_LNB_TREE, list(SKIP_TEXTS))
                logger.info(f"  펼침 후 LNB 노드: {len(tree2)}")
            except Exception:
                tree2 = tree
            mod_entry["lnb_items"] = tree2

            # leaf 후보: has_child_list 없고 has_expand_icon 없는 항목
            leaves = [t for t in tree2 if not t.get("has_child_list") and not t.get("has_expand_icon")]
            # 길이 너무 짧은 텍스트는 카테고리일 수도 있음, 일단 시도
            logger.info(f"  Leaf 후보: {len(leaves)} (limit={args.leaf_limit})")

            visited_texts = set()
            for idx, leaf in enumerate(leaves[: args.leaf_limit]):
                if leaf["text"] in visited_texts:
                    continue
                visited_texts.add(leaf["text"])
                logger.info(f"  [{idx+1}/{min(len(leaves), args.leaf_limit)}] {leaf['text']}")
                ok = click_text_in_any_frame(page, leaf["text"], timeout_ms=2500)
                if not ok:
                    continue
                page.wait_for_timeout(1800)
                # 페이지 정보 + 엑셀 시도
                try:
                    info_per_frame = []
                    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
                        try:
                            info_per_frame.append(fr.evaluate(JS_PAGE_INFO))
                        except Exception:
                            continue
                    merged_title = ""
                    has_excel = False
                    has_grid = False
                    excel_count = 0
                    for inf in info_per_frame:
                        if not merged_title and inf.get("title"):
                            merged_title = inf["title"]
                        has_excel = has_excel or inf.get("has_excel_btn")
                        has_grid = has_grid or inf.get("has_grid")
                        excel_count += inf.get("excel_btn_count", 0)
                except Exception as e:
                    logger.warning(f"    page info 실패: {e}")
                    continue

                url = current_url(page)
                visit = {
                    "lnb_text": leaf["text"],
                    "page_title": merged_title,
                    "url": url,
                    "has_excel_btn": has_excel,
                    "has_grid": has_grid,
                    "excel_btn_count": excel_count,
                    "download": None,
                }

                # 엑셀 다운로드 시도 (한도 내)
                if has_excel and downloads_done < args.download_limit:
                    save_label = f"{code}_{safe_filename(leaf['text'])}"
                    logger.info(f"    ★ 엑셀 발견 → 다운로드 시도")
                    dl_result = try_excel_download(page, save_label)
                    visit["download"] = dl_result
                    if dl_result and dl_result.get("saved"):
                        downloads_done += 1
                        logger.info(f"    ✓ 저장: {Path(dl_result['saved']).name} ({dl_result['size']}B)")
                    else:
                        logger.warning(f"    ✗ 다운로드 실패: {dl_result}")

                mod_entry["leaf_visits"].append(visit)

            result["modules"].append(mod_entry)
            # 스크린샷
            try:
                page.screenshot(path=str(SCR_DIR / f"v5_mod_{code}_final.png"))
            except Exception:
                pass

        context.close()
        browser.close()

    # 저장
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON 저장: {OUT_JSON}")

    # Markdown 요약
    lines = ["# 아마란스 전체 구조 + Export 자동 수집 (v5)", ""]
    lines.append(f"- 계정: tgjeon / 시각: {result['crawled_at']}")
    lines.append(f"- 탐색 모듈: {len(result['modules'])} / 다운로드 성공: {downloads_done}")
    lines.append("")

    # Export 가능 + 다운로드 성공 목록
    successes = []
    for me in result["modules"]:
        for v in me["leaf_visits"]:
            if v.get("download") and v["download"].get("saved"):
                successes.append((me["code"], me["text"], v))
    lines.append(f"## ✅ 다운로드 성공 ({len(successes)}개)\n")
    lines.append("| 모듈 | 메뉴 | URL | 파일 | 크기 |")
    lines.append("|---|---|---|---|---|")
    for code, mod_text, v in successes:
        d = v["download"]
        lines.append(f"| {code} | {v['lnb_text']} | `{(v['url'] or '')[-50:]}` | `{Path(d['saved']).name}` | {d['size']}B |")
    lines.append("")

    # 모듈별 전체
    lines.append("## 모듈별 LNB 트리 + 진입 결과")
    for me in result["modules"]:
        lines.append(f"\n### [{me['code']}] {me['text']}")
        lines.append(f"- LNB 노드: {len(me['lnb_items'])}, leaf 방문: {len(me['leaf_visits'])}")
        lines.append("")
        lines.append("| LNB | 페이지 제목 | URL | 그리드 | 엑셀 | 다운로드 |")
        lines.append("|---|---|---|---|---|---|")
        for v in me["leaf_visits"]:
            g = "✓" if v["has_grid"] else ""
            e = "✓" if v["has_excel_btn"] else ""
            d = v.get("download") or {}
            ds = "✅" if d.get("saved") else ("❌" if d else "")
            lines.append(
                f"| {v['lnb_text']} | {v['page_title'][:30]} | `{(v['url'] or '')[-50:]}` | {g} | {e} | {ds} |"
            )

    # 미리보기 포함
    if successes:
        lines.append("\n## 다운로드 파일 미리보기")
        for code, mod_text, v in successes:
            d = v["download"]
            s = d.get("summary")
            if not s or "error" in s:
                continue
            lines.append(f"\n### {code} > {v['lnb_text']}")
            for sheet, rows in s.get("preview", {}).items():
                lines.append(f"\n**시트**: `{sheet}` (총 {s.get('row_count', {}).get(sheet, '?')}행)")
                if rows:
                    headers = rows[0]
                    lines.append("| " + " | ".join(str(h)[:25] for h in headers[:10]) + " |")
                    lines.append("|" + "|".join(["---"] * min(len(headers), 10)) + "|")
                    for row in rows[1:6]:
                        padded = row + [""] * (10 - len(row)) if len(row) < 10 else row
                        lines.append("| " + " | ".join(str(c)[:25] for c in padded[:10]) + " |")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"MD 저장: {OUT_MD}")

    logger.info("=" * 60)
    logger.info(f"모듈: {len(result['modules'])} / 다운로드: {downloads_done}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
