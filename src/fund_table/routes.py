"""
프로젝트 관리표 API 라우터
- /api/fund/* : 프로젝트, 공종, 하도급, 연락처, GW 데이터 CRUD
- /fund       : 프로젝트 관리 웹 페이지 서빙
"""

import json
import logging
import os
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

# .env 로드 (ANTHROPIC_API_KEY 등)
load_dotenv(Path(__file__).parent.parent.parent / "config" / ".env")

from src.fund_table import db

logger = logging.getLogger("fund_routes")

# 정적 파일 경로 (src/chatbot/static/fund.html)
STATIC_DIR = Path(__file__).parent.parent / "chatbot" / "static"

router = APIRouter()


# 공용 인증 미들웨어 (중복 제거)
from src.auth.middleware import require_auth

# 관리자 GW ID (소유자 검증 우회용)
ADMIN_GW_ID = os.environ.get("ADMIN_GW_ID", "")


def _check_owner(project_id: int, user: dict):
    """프로젝트 소유자 검증. 관리자는 통과."""
    if ADMIN_GW_ID and user["gw_id"] == ADMIN_GW_ID:
        return
    if not db.check_project_owner(project_id, user["gw_id"]):
        raise HTTPException(status_code=403, detail="이 프로젝트의 소유자만 수정할 수 있습니다.")


# ─────────────────────────────────────────
# Pydantic 요청 모델
# ─────────────────────────────────────────

class ProjectCreate(BaseModel):
    """프로젝트 생성 요청"""
    name: str
    description: Optional[str] = None
    design_amount: Optional[int] = None
    construction_amount: Optional[int] = None
    execution_budget: Optional[int] = None
    profit_amount: Optional[int] = None
    profit_rate: Optional[float] = None
    status: Optional[str] = None
    grade: Optional[str] = None
    project_code: Optional[str] = None  # GW 사업코드 (예: GS-25-0088)


class ProjectUpdate(BaseModel):
    """프로젝트 수정 요청"""
    name: Optional[str] = None
    description: Optional[str] = None
    design_amount: Optional[int] = None
    construction_amount: Optional[int] = None
    execution_budget: Optional[int] = None
    profit_amount: Optional[int] = None
    profit_rate: Optional[float] = None
    status: Optional[str] = None
    grade: Optional[str] = None
    project_code: Optional[str] = None  # GW 사업코드
    timeline_start_month: Optional[str] = None  # 타임라인 시작월 (YYYY-MM)
    timeline_end_month: Optional[str] = None    # 타임라인 종료월 (YYYY-MM)


class TradeCreate(BaseModel):
    """공종 생성 요청"""
    name: str
    sort_order: Optional[int] = 0


class TradeUpdate(BaseModel):
    """공종 수정 요청"""
    name: Optional[str] = None
    sort_order: Optional[int] = None


class SubcontractCreate(BaseModel):
    """하도급 업체 생성 요청"""
    company_name: str
    trade_id: Optional[int] = None
    account_category: Optional[str] = None
    has_estimate: Optional[int] = None
    has_contract: Optional[int] = None
    has_vendor_reg: Optional[int] = None
    estimate_amount: Optional[int] = None
    contract_amount: Optional[int] = None
    payment_1: Optional[int] = None
    payment_2: Optional[int] = None
    payment_3: Optional[int] = None
    payment_4: Optional[int] = None
    remaining_amount: Optional[int] = None
    payment_rate: Optional[float] = None
    payment_1_confirmed: Optional[int] = None
    payment_2_confirmed: Optional[int] = None
    payment_3_confirmed: Optional[int] = None
    payment_4_confirmed: Optional[int] = None
    sort_order: Optional[int] = None


class SubcontractUpdate(BaseModel):
    """하도급 업체 수정 요청"""
    company_name: Optional[str] = None
    trade_id: Optional[int] = None
    account_category: Optional[str] = None
    has_estimate: Optional[int] = None
    has_contract: Optional[int] = None
    has_vendor_reg: Optional[int] = None
    estimate_amount: Optional[int] = None
    contract_amount: Optional[int] = None
    payment_1: Optional[int] = None
    payment_2: Optional[int] = None
    payment_3: Optional[int] = None
    payment_4: Optional[int] = None
    remaining_amount: Optional[int] = None
    payment_rate: Optional[float] = None
    payment_1_confirmed: Optional[int] = None
    payment_2_confirmed: Optional[int] = None
    payment_3_confirmed: Optional[int] = None
    payment_4_confirmed: Optional[int] = None
    sort_order: Optional[int] = None


class ContactCreate(BaseModel):
    """연락처 생성 요청"""
    vendor_name: Optional[str] = None  # 프론트엔드 필드명
    company_name: Optional[str] = None  # 레거시 호환
    trade_name: Optional[str] = None
    trade_id: Optional[int] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    note: Optional[str] = None

    @property
    def resolved_company_name(self) -> str:
        return self.vendor_name or self.company_name or ""


class ContactUpdate(BaseModel):
    """연락처 수정 요청"""
    vendor_name: Optional[str] = None
    company_name: Optional[str] = None
    trade_name: Optional[str] = None
    trade_id: Optional[int] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    note: Optional[str] = None

    @property
    def resolved_company_name(self) -> Optional[str]:
        return self.vendor_name or self.company_name


class ReorderItem(BaseModel):
    """순서 변경 항목"""
    id: int

class ReorderRequest(BaseModel):
    """프로젝트 순서 일괄 변경"""
    order: list[ReorderItem]


class CollectionItem(BaseModel):
    """수금현황 항목"""
    id: Optional[int] = None
    category: Optional[str] = ""
    stage: Optional[str] = ""
    amount: Optional[int] = 0
    collected: Optional[int] = 0
    collection_date: Optional[str] = ""

class CollectionsBulkSave(BaseModel):
    """수금현황 일괄 저장"""
    collections: list[CollectionItem]


class MemberItem(BaseModel):
    role: str = ""
    name: str = ""

class MilestoneItem(BaseModel):
    id: Optional[int] = None
    name: str = ""
    completed: int = 0
    date: str = ""

class OverviewSave(BaseModel):
    """프로젝트 개요 저장"""
    project_category: Optional[str] = None
    location: Optional[str] = None
    usage: Optional[str] = None
    scale: Optional[str] = None
    area_pyeong: Optional[float] = None
    design_start: Optional[str] = None
    design_end: Optional[str] = None
    construction_start: Optional[str] = None
    construction_end: Optional[str] = None
    open_date: Optional[str] = None
    current_status: Optional[str] = None
    design_contract_date: Optional[str] = None
    design_contract_amount: Optional[int] = None
    construction_contract_date: Optional[str] = None
    construction_contract_amount: Optional[int] = None
    issue_design: Optional[str] = None
    issue_schedule: Optional[str] = None
    issue_budget: Optional[str] = None
    issue_operation: Optional[str] = None
    issue_defect: Optional[str] = None
    issue_other: Optional[str] = None
    members: Optional[list[MemberItem]] = None
    milestones: Optional[list[MilestoneItem]] = None


# ─────────────────────────────────────────
# 프로젝트 CRUD 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects")
async def list_projects(request: Request):
    """전체 프로젝트 목록 조회"""
    require_auth(request)
    projects = db.list_projects()
    return JSONResponse({"projects": projects})


@router.get("/api/fund/portfolio-summary")
async def get_portfolio_summary(request: Request):
    """포트폴리오 전체 요약 (프로젝트 비교 뷰용)"""
    require_auth(request)
    data = db.get_portfolio_summary()
    return JSONResponse({"projects": data})


@router.post("/api/fund/projects")
async def create_project(req: ProjectCreate, request: Request):
    """프로젝트 생성. project_code가 있으면 백그라운드에서 GW 크롤링 자동 실행."""
    import threading
    user = require_auth(request)
    kwargs = req.model_dump(exclude={"name"}, exclude_none=True)
    kwargs["owner_gw_id"] = user["gw_id"]
    result = db.create_project(name=req.name, **kwargs)
    if not result["success"]:
        logger.warning("프로젝트 생성 실패: %s", result["message"])
        raise HTTPException(status_code=409, detail=result["message"])

    # project_code가 있으면 백그라운드에서 GW 크롤링 자동 시작
    new_id = result.get("id")
    project_code = req.project_code
    if project_code and new_id:
        gw_id = user["gw_id"]
        def _auto_crawl():
            try:
                from src.fund_table.project_crawler import crawl_project_info
                crawl_project_info(gw_id, project_code, new_id)
                logger.info("자동 크롤링 완료 (project_info): project_id=%d, code=%s", new_id, project_code)
            except Exception as e:
                logger.warning("자동 크롤링 실패 (project_info): %s", e)
            # 사업별 크롤러 (메인 — 전기+당기)
            try:
                from src.fund_table.budget_crawler_by_project import crawl_budget_by_project
                crawl_budget_by_project(gw_id, new_id, project_code)
                logger.info("자동 크롤링 완료 (budget_by_project): project_id=%d, code=%s", new_id, project_code)
            except Exception as e:
                logger.warning("자동 크롤링 실패 (budget_by_project): %s", e)
            # 상세 합계 크롤러 (보충 — 수입/지출 합계)
            try:
                from src.fund_table.budget_crawler import crawl_budget_summary
                crawl_budget_summary(gw_id, new_id, project_code)
                logger.info("자동 크롤링 완료 (budget_summary): project_id=%d, code=%s", new_id, project_code)
            except Exception as e:
                logger.warning("자동 크롤링 실패 (budget_summary): %s", e)
        threading.Thread(target=_auto_crawl, daemon=True).start()
        result["crawling"] = True  # 프론트엔드에 크롤링 진행 중임을 알림

    return JSONResponse(result, status_code=201)


@router.put("/api/fund/projects/reorder")
async def reorder_projects(req: ReorderRequest, request: Request):
    """프로젝트 순서 일괄 업데이트"""
    require_auth(request)
    order = [{"id": item.id} for item in req.order]
    result = db.reorder_projects(order)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


@router.get("/api/fund/projects/{project_id}")
async def get_project(project_id: int, request: Request):
    """프로젝트 상세 조회 (공종, 하도급 요약 포함)"""
    require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    # 공종 목록
    trades = db.list_trades(project_id)
    # 하도급 요약 (업체 수, 계약총액)
    subcontracts = db.list_subcontracts(project_id)
    sub_summary = {
        "count": len(subcontracts),
        "total_contract": sum((s.get("contract_amount") or 0) for s in subcontracts),
        "total_paid": sum(
            (s.get("payment_1") or 0) + (s.get("payment_2") or 0)
            + (s.get("payment_3") or 0) + (s.get("payment_4") or 0)
            for s in subcontracts
        ),
    }

    return JSONResponse({
        "project": project,
        "trades": trades,
        "subcontracts_summary": sub_summary,
    })


@router.put("/api/fund/projects/{project_id}")
async def update_project(project_id: int, req: ProjectUpdate, request: Request):
    """프로젝트 정보 수정"""
    user = require_auth(request)
    # 프로젝트 존재 확인
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    _check_owner(project_id, user)

    kwargs = req.model_dump(exclude_none=True)
    if not kwargs:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")

    result = db.update_project(project_id, **kwargs)
    if not result["success"]:
        logger.warning("프로젝트 수정 실패 (id=%d): %s", project_id, result["message"])
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


@router.delete("/api/fund/projects/{project_id}")
async def delete_project(project_id: int, request: Request):
    """프로젝트 삭제 (하위 데이터 포함)"""
    user = require_auth(request)
    _check_owner(project_id, user)
    result = db.delete_project(project_id)
    if not result["success"]:
        logger.warning("프로젝트 삭제 실패 (id=%d): %s", project_id, result["message"])
        raise HTTPException(status_code=404, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# 공종 CRUD 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/trades")
async def list_trades(project_id: int, request: Request):
    """프로젝트의 공종 목록 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    trades = db.list_trades(project_id)
    return JSONResponse({"trades": trades})


@router.post("/api/fund/projects/{project_id}/trades")
async def add_trade(project_id: int, req: TradeCreate, request: Request):
    """공종 추가"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    result = db.add_trade(project_id, name=req.name, sort_order=req.sort_order or 0)
    if not result["success"]:
        raise HTTPException(status_code=409, detail=result["message"])
    return JSONResponse(result, status_code=201)


@router.put("/api/fund/projects/{project_id}/trades/{trade_id}")
async def update_trade(project_id: int, trade_id: int, req: TradeUpdate, request: Request):
    """공종 수정"""
    user = require_auth(request)
    _check_owner(project_id, user)
    kwargs = req.model_dump(exclude_none=True)
    if not kwargs:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")
    result = db.update_trade(trade_id, **kwargs)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


@router.delete("/api/fund/projects/{project_id}/trades/{trade_id}")
async def delete_trade(project_id: int, trade_id: int, request: Request):
    """공종 삭제"""
    user = require_auth(request)
    _check_owner(project_id, user)
    result = db.delete_trade(trade_id)
    if not result["success"]:
        raise HTTPException(status_code=404, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# 하도급 상세 CRUD 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/subcontracts")
async def list_subcontracts(project_id: int, request: Request):
    """프로젝트의 하도급 목록 조회 (공종명 포함)"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    subcontracts = db.list_subcontracts(project_id)
    return JSONResponse({"subcontracts": subcontracts})


@router.post("/api/fund/projects/{project_id}/subcontracts")
async def add_subcontract(project_id: int, req: SubcontractCreate, request: Request):
    """하도급 업체 추가"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    kwargs = req.model_dump(exclude={"company_name"}, exclude_none=True)
    result = db.add_subcontract(project_id, company_name=req.company_name, **kwargs)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result, status_code=201)


@router.put("/api/fund/projects/{project_id}/subcontracts")
async def save_subcontracts_bulk(project_id: int, request: Request):
    """하도급 일괄 저장 (프론트엔드 saveSubcontracts)"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    body = await request.json()
    rows = body.get("subcontracts", [])

    # 기존 하도급 목록 조회
    existing = {s["id"]: s for s in db.list_subcontracts(project_id)}
    incoming_ids = set()

    for i, row in enumerate(rows):
        row_id = row.get("id")
        row["sort_order"] = i
        # 불필요한 키 제거
        clean = {k: v for k, v in row.items() if k not in ("id",)}

        if row_id and row_id in existing:
            # 기존 항목 수정
            db.update_subcontract(row_id, **clean)
            incoming_ids.add(row_id)
        else:
            # 새 항목 추가
            company = clean.pop("company_name", "")
            result = db.add_subcontract(project_id, company_name=company or "미정", **clean)
            if result.get("id"):
                incoming_ids.add(result["id"])

    # 삭제된 항목 처리
    for eid in existing:
        if eid not in incoming_ids:
            db.delete_subcontract(eid)

    return JSONResponse({"success": True, "message": f"하도급 {len(rows)}건 저장 완료"})


@router.put("/api/fund/subcontracts/{sub_id}")
async def update_subcontract(sub_id: int, req: SubcontractUpdate, request: Request):
    """하도급 업체 정보 수정 (개별)"""
    user = require_auth(request)
    # 소유자 검증 — sub_id로 project_id 조회
    conn = db.get_db()
    try:
        row = conn.execute("SELECT project_id FROM subcontracts WHERE id = ?", (sub_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="하도급 항목을 찾을 수 없습니다.")
        _check_owner(row["project_id"], user)
    finally:
        conn.close()
    kwargs = req.model_dump(exclude_none=True)
    if not kwargs:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")
    result = db.update_subcontract(sub_id, **kwargs)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


@router.delete("/api/fund/subcontracts/{sub_id}")
async def delete_subcontract(sub_id: int, request: Request):
    """하도급 업체 삭제"""
    user = require_auth(request)
    # 소유자 검증 — sub_id로 project_id 조회
    conn = db.get_db()
    try:
        row = conn.execute("SELECT project_id FROM subcontracts WHERE id = ?", (sub_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="하도급 항목을 찾을 수 없습니다.")
        _check_owner(row["project_id"], user)
    finally:
        conn.close()
    result = db.delete_subcontract(sub_id)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# 연락처 CRUD 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/contacts")
async def list_contacts(project_id: int, request: Request):
    """프로젝트의 연락처 목록 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    contacts = db.list_contacts(project_id)
    return JSONResponse({"contacts": contacts})


@router.post("/api/fund/projects/{project_id}/contacts")
async def add_contact(project_id: int, req: ContactCreate, request: Request):
    """연락처 추가"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    company = req.resolved_company_name
    if not company:
        raise HTTPException(status_code=400, detail="업체명을 입력하세요.")
    kwargs = req.model_dump(exclude={"company_name", "vendor_name"}, exclude_none=True)
    result = db.add_contact(project_id, company_name=company, **kwargs)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result, status_code=201)


@router.put("/api/fund/projects/{project_id}/contacts/{contact_id}")
async def update_contact(project_id: int, contact_id: int, req: ContactUpdate, request: Request):
    """연락처 수정"""
    user = require_auth(request)
    _check_owner(project_id, user)
    # vendor_name → company_name으로 변환
    kwargs = req.model_dump(exclude={"vendor_name"}, exclude_none=True)
    resolved = req.resolved_company_name
    if resolved:
        kwargs["company_name"] = resolved
    if not kwargs:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")
    result = db.update_contact(contact_id, **kwargs)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


@router.delete("/api/fund/projects/{project_id}/contacts/{contact_id}")
async def delete_contact(project_id: int, contact_id: int, request: Request):
    """연락처 삭제"""
    user = require_auth(request)
    _check_owner(project_id, user)
    result = db.delete_contact(contact_id)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# 프로젝트 개요 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/overview")
async def get_overview(project_id: int, request: Request):
    """프로젝트 개요 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    overview = db.get_project_overview(project_id)
    return JSONResponse({"overview": overview})


@router.put("/api/fund/projects/{project_id}/overview")
async def save_overview(project_id: int, req: OverviewSave, request: Request):
    """프로젝트 개요 저장"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    data = req.model_dump(exclude_none=True)
    # members/milestones는 dict 리스트로 변환
    if "members" in data:
        data["members"] = [m.model_dump() if hasattr(m, "model_dump") else m for m in req.members]
    if "milestones" in data:
        data["milestones"] = [m.model_dump() if hasattr(m, "model_dump") else m for m in req.milestones]
    result = db.save_project_overview(project_id, data)
    if not result["success"]:
        logger.error("개요 저장 실패 (project_id=%d): %s", project_id, result["message"])
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# GW 크롤링 엔드포인트
# ─────────────────────────────────────────

@router.post("/api/fund/projects/{project_id}/crawl-gw")
async def crawl_gw_project(project_id: int, request: Request):
    """
    GW에서 프로젝트 등록정보 + 예실대비 크롤링.
    순서: 프로젝트정보 → 사업별(메인) → 상세 합계(보충)
    각 단계 독립 try/except (한쪽 실패해도 계속)
    """
    user = require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)

    pcode = project.get("project_code")
    if not pcode:
        raise HTTPException(status_code=400, detail="프로젝트 코드가 설정되지 않았습니다. 개요 탭에서 GW 사업코드를 먼저 입력해주세요.")

    gw_id = user["gw_id"]
    results = {"project_info": None, "budget_by_project": None, "budget_summary": None}

    def _crawl():
        # 1단계: 프로젝트 등록정보
        try:
            from src.fund_table.project_crawler import crawl_project_info
            results["project_info"] = crawl_project_info(gw_id, pcode, project_id)
        except Exception as e:
            results["project_info"] = {"success": False, "error": f"프로젝트정보: {e}"}

        # 2단계: 예실대비현황(사업별) — 메인 크롤러 (전기+당기)
        try:
            from src.fund_table.budget_crawler_by_project import crawl_budget_by_project
            results["budget_by_project"] = crawl_budget_by_project(gw_id, project_id, pcode)
        except Exception as e:
            results["budget_by_project"] = {"success": False, "error": f"예실대비(사업별): {e}"}

        # 3단계: 예실대비현황(상세) — 합계 보충 (수입합계/지출합계/총잔액)
        try:
            from src.fund_table.budget_crawler import crawl_budget_summary
            results["budget_summary"] = crawl_budget_summary(gw_id, project_id, pcode)
        except Exception as e:
            results["budget_summary"] = {"success": False, "error": f"예실대비(합계): {e}"}

    # asyncio.to_thread로 이벤트 루프 블로킹 방지 (Playwright는 동기 API이므로 스레드 풀에서 실행)
    import asyncio
    try:
        await asyncio.wait_for(asyncio.to_thread(_crawl), timeout=300)
    except asyncio.TimeoutError:
        return JSONResponse({"success": False, "error": "크롤링 시간 초과 (5분)"}, status_code=504)

    # 항목별 성공/실패 집계
    label_map = {
        "project_info": "프로젝트정보",
        "budget_by_project": "예실대비(사업별)",
        "budget_summary": "예실대비(합계)",
    }
    detail_list = []
    synced = 0
    failed = 0
    for key, r in results.items():
        label = label_map.get(key, key)
        if r and r.get("success"):
            synced += 1
            msg = "완료"
            if key == "budget_by_project" and r.get("record_count"):
                msg = f"{r['record_count']}건 저장"
            elif key == "budget_summary" and r.get("summary"):
                s = r["summary"]
                msg = f"수입합계={s.get('income_total',0):,}, 지출합계={s.get('expense_total',0):,}"
            detail_list.append({"item": label, "status": "success", "message": msg})
        else:
            failed += 1
            detail_list.append({"item": label, "status": "fail", "message": r.get("error", "실패") if r else "실행되지 않음"})

    total = synced + failed
    success = synced > 0
    return JSONResponse({
        "success": success,
        "message": f"{total}개 중 {synced}개 동기화 완료" if success else "크롤링 실패",
        "synced": synced,
        "failed": failed,
        "total": total,
        "details": detail_list,
    })


@router.post("/api/fund/crawl-gw-all")
async def crawl_gw_all(request: Request):
    """
    등록된 모든 프로젝트의 GW 정보 일괄 크롤링.
    순서: 프로젝트정보 → 사업별(메인) → 상세합계(보충)
    각 단계 독립 try/except (한쪽 실패해도 계속)
    """
    from src.fund_table.scheduler import sync_running

    # 스케줄러 또는 다른 수동 트리거가 이미 실행 중이면 거부
    if sync_running.is_set():
        return JSONResponse(
            {"success": False, "error": "동기화가 이미 진행 중입니다."},
            status_code=409,
        )

    user = require_auth(request)
    gw_id = user["gw_id"]

    result = {}

    def _crawl():
        errors = []
        stage_results = {}

        # 1단계: 프로젝트 등록정보 크롤링
        try:
            from src.fund_table.project_crawler import crawl_all_project_info
            proj_result = crawl_all_project_info(gw_id)
            stage_results["project_info"] = proj_result
            proj_success = 0
            if proj_result.get("success") and proj_result.get("results"):
                proj_success = sum(
                    1 for r in proj_result["results"] if r.get("status") == "success"
                )
            elif not proj_result.get("success"):
                errors.append(f"프로젝트정보: {proj_result.get('error', '실패')}")
        except Exception as e:
            proj_result = {"success": False, "error": str(e)}
            stage_results["project_info"] = proj_result
            proj_success = 0
            errors.append(f"프로젝트정보: {e}")

        # 2단계: 예실대비현황(사업별) 크롤링 — 메인 (전기+당기)
        try:
            from src.fund_table.budget_crawler_by_project import crawl_all_by_project
            byprj_result = crawl_all_by_project(gw_id)
            stage_results["budget_by_project"] = byprj_result
            byprj_success = 0
            if byprj_result.get("success") and byprj_result.get("results"):
                byprj_success = sum(
                    1 for r in byprj_result["results"] if r.get("status") == "success"
                )
            elif not byprj_result.get("success"):
                errors.append(f"예실대비(사업별): {byprj_result.get('error', '실패')}")
        except Exception as e:
            byprj_result = {"success": False, "error": str(e)}
            stage_results["budget_by_project"] = byprj_result
            byprj_success = 0
            errors.append(f"예실대비(사업별): {e}")

        # 3단계: 예실대비현황(상세) 합계 크롤링 — 보충 (수입합계/지출합계/총잔액)
        try:
            from src.fund_table.budget_crawler import crawl_all_summary
            summary_result = crawl_all_summary(gw_id)
            stage_results["budget_summary"] = summary_result
            summary_success = 0
            if summary_result.get("success") and summary_result.get("results"):
                summary_success = sum(
                    1 for r in summary_result["results"] if r.get("status") == "success"
                )
            elif not summary_result.get("success"):
                errors.append(f"예실대비(합계): {summary_result.get('error', '실패')}")
        except Exception as e:
            summary_result = {"success": False, "error": str(e)}
            stage_results["budget_summary"] = summary_result
            summary_success = 0
            errors.append(f"예실대비(합계): {e}")

        # 3개 크롤러 중 가장 많이 성공한 수 기준
        synced = max(proj_success, byprj_success, summary_success)
        overall_success = synced > 0

        # 대상 프로젝트 수
        total_targets = (
            len(proj_result.get("results", []))
            or len(byprj_result.get("results", []))
            or len(summary_result.get("results", []))
        )
        failed_count = total_targets - synced if total_targets > synced else 0

        # 프로젝트별 상세 결과 (사업별 결과 우선, 없으면 프로젝트정보 결과)
        detail_list = []
        all_results = (
            byprj_result.get("results", [])
            or proj_result.get("results", [])
            or summary_result.get("results", [])
        )
        for item in all_results:
            detail_list.append({
                "project_id": item.get("project_id"),
                "project_name": item.get("project_name", item.get("project_code", "")),
                "status": item.get("status", "unknown"),
                "message": item.get("message", item.get("error", "")),
            })

        # 단계별 요약
        stages = []
        stage_labels = {
            "project_info": "프로젝트정보",
            "budget_by_project": "예실대비(사업별)",
            "budget_summary": "예실대비(합계)",
        }
        stage_counts = {
            "project_info": proj_success,
            "budget_by_project": byprj_success,
            "budget_summary": summary_success,
        }
        for key, label in stage_labels.items():
            sr = stage_results.get(key, {})
            stages.append({
                "stage": label,
                "success": sr.get("success", False),
                "count": stage_counts.get(key, 0),
                "message": sr.get("message", sr.get("error", "")),
            })

        result["data"] = {
            "success": overall_success,
            "synced": synced,
            "failed": failed_count,
            "total": total_targets,
            "message": f"{total_targets}개 중 {synced}개 동기화 완료"
                       + (f" (오류: {'; '.join(errors)})" if errors and overall_success else ""),
            "error": "; ".join(errors) if not overall_success and errors else None,
            "details": detail_list,
            "stages": stages,
        }

    import asyncio
    try:
        await asyncio.wait_for(asyncio.to_thread(_crawl), timeout=900)
    except asyncio.TimeoutError:
        return JSONResponse({"success": False, "error": "크롤링 시간 초과 (15분)"}, status_code=504)

    return JSONResponse(result.get("data", {"success": False, "error": "알 수 없는 오류"}))


# ─────────────────────────────────────────
# 수금현황 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/collections")
async def list_collections(project_id: int, request: Request):
    """수금현황 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    collections = db.list_collections(project_id)
    return JSONResponse({"collections": collections})


@router.put("/api/fund/projects/{project_id}/collections")
async def save_collections(project_id: int, req: CollectionsBulkSave, request: Request):
    """수금현황 일괄 저장"""
    user = require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    _check_owner(project_id, user)
    items = [item.model_dump() for item in req.collections]
    result = db.save_collections_bulk(project_id, items)
    if not result["success"]:
        logger.error("수금현황 저장 실패 (project_id=%d): %s", project_id, result["message"])
        raise HTTPException(status_code=400, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# GW 프로젝트 검색 (사업코드 자동 조회)
# ─────────────────────────────────────────

@router.post("/api/fund/gw/search-projects")
async def search_gw_projects_api(request: Request):
    """GW 프로젝트 키워드 검색 (캐시 우선, 없으면 안내)"""
    user = require_auth(request)
    body = await request.json()
    keyword = body.get("search_name", "").strip()

    cache_info = db.get_gw_cache_info()
    if cache_info["count"] > 0:
        results = db.search_gw_projects_cache(keyword)
        return JSONResponse({
            "success": True,
            "projects": results,
            "total": len(results),
            "source": "cache",
            "cache_count": cache_info["count"],
            "cache_updated": cache_info["last_update"],
        })

    return JSONResponse({
        "success": False,
        "error": "GW 프로젝트 목록이 아직 없습니다.",
        "need_fetch": True,
    })


@router.post("/api/fund/gw/fetch-project-list")
async def fetch_gw_project_list(request: Request):
    """GW에서 전체 프로젝트 목록 크롤링 → 캐시 저장"""
    import traceback
    user = require_auth(request)
    gw_id = user["gw_id"]
    result = {"data": None, "error": None}

    def _fetch():
        try:
            from src.fund_table.project_crawler import search_gw_projects as _search_fn
            result["data"] = _search_fn(gw_id, "")
        except Exception as e:
            logger.error(f"GW 프로젝트 목록 크롤링 스레드 오류: {e}\n{traceback.format_exc()}")
            result["error"] = str(e)

    import asyncio
    try:
        await asyncio.wait_for(asyncio.to_thread(_fetch), timeout=180)
    except asyncio.TimeoutError:
        return JSONResponse({"success": False, "error": "GW 접속 시간 초과 (3분)"}, status_code=504)

    if result["error"]:
        return JSONResponse({"success": False, "error": f"GW 크롤링 오류: {result['error']}"})

    data = result["data"]
    if not data or not data.get("success"):
        return JSONResponse({"success": False, "error": data.get("error", "GW 프로젝트 목록 가져오기 실패") if data else "알 수 없는 오류"})

    projects = data.get("projects", [])
    if projects:
        db.save_gw_projects_cache_v2(projects)

    return JSONResponse({
        "success": True,
        "total": len(projects),
        "message": f"GW에서 {len(projects)}개 프로젝트를 가져왔습니다.",
    })


# ─────────────────────────────────────────
# PM 시트 임포트 엔드포인트
# ─────────────────────────────────────────

@router.post("/api/fund/import-pm-sheet")
async def import_pm_sheet(request: Request):
    """
    PM팀 Official 시트에서 전체 프로젝트 기본정보를 읽어 DB에 반영.
    mode: 'upsert' (기존 업데이트 + 신규 생성) 또는 'insert_only' (신규만)
    """
    user = require_auth(request)
    gw_id = user["gw_id"]

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass

    spreadsheet_id = body.get("spreadsheet_id", None)
    mode = body.get("mode", "upsert")

    result = {"data": None, "error": None}

    def _import():
        try:
            from src.fund_table.sheets_import import import_from_pm_sheet
            result["data"] = import_from_pm_sheet(
                spreadsheet_id=spreadsheet_id,
                owner_gw_id=gw_id,
                mode=mode,
            )
        except Exception as e:
            logger.error("PM 시트 임포트 오류: %s", e, exc_info=True)
            result["error"] = str(e)

    try:
        await asyncio.wait_for(asyncio.to_thread(_import), timeout=120)
    except asyncio.TimeoutError:
        return JSONResponse(
            {"success": False, "error": "Google Sheets 접속 시간 초과 (2분)"},
            status_code=504,
        )

    if result["error"]:
        return JSONResponse(
            {"success": False, "error": f"PM 시트 임포트 오류: {result['error']}"},
            status_code=500,
        )

    data = result["data"]
    if not data:
        return JSONResponse(
            {"success": False, "error": "알 수 없는 오류"},
            status_code=500,
        )

    return JSONResponse(data)


# ─────────────────────────────────────────
# GW 데이터 조회 엔드포인트 (읽기 전용)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/payments")
async def list_payments(project_id: int, request: Request, limit: int = 500):
    """이체완료 내역 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    # 보안: limit 상한 제한 (DoS 방지)
    limit = min(max(1, limit), 500)
    payments = db.list_payment_history(project_id=project_id, limit=limit)
    return JSONResponse({"payments": payments})


@router.post("/api/fund/projects/{project_id}/payments/import")
async def import_payment_excel(project_id: int, request: Request, file: UploadFile = File(...)):
    """엑셀 파일에서 이체내역 임포트 (헤더 자동 감지)"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    # 파일 확장자 검증
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 지원합니다.")

    import openpyxl, io, re

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        # 헤더 행 감지 (키워드 매칭)
        header_row = None
        col_map = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False), start=1):
            texts = [str(cell.value or "").replace(" ", "") for cell in row]
            joined = " ".join(texts)
            if any(kw in joined for kw in ("확정일", "거래처", "금액", "이체일", "예정일")):
                header_row = r_idx
                # 컬럼 매핑
                for c_idx, cell in enumerate(row):
                    h = str(cell.value or "").replace(" ", "")
                    if not h:
                        continue
                    if "회계" in h and "단위" in h:
                        col_map["accounting_unit"] = c_idx
                    elif "예정" in h and "일" in h:
                        col_map["scheduled_date"] = c_idx
                    elif "확정" in h and "일" in h:
                        col_map["confirmed_date"] = c_idx
                    elif "자금" in h and "과목" in h:
                        col_map["fund_category"] = c_idx
                    elif "거래처" in h and "코드" in h and "카드" not in h:
                        col_map["vendor_code"] = c_idx
                    elif "거래처" in h and ("명" in h or "이름" in h) and "카드" not in h:
                        col_map["vendor_name"] = c_idx
                    elif "사업자" in h or "주민" in h:
                        col_map["business_number"] = c_idx
                    elif h in ("은행", "은행명"):
                        col_map["bank_name"] = c_idx
                    elif "계좌" in h and "번호" in h:
                        col_map["account_number"] = c_idx
                    elif "예금주" in h and "실제" not in h:
                        col_map["account_holder"] = c_idx
                    elif h == "적요" or h == "비고":
                        col_map["description"] = c_idx
                    elif h in ("금액", "이체금액"):
                        col_map["amount"] = c_idx
                    elif "부서" in h or "사용부서" in h:
                        col_map["department"] = c_idx
                    elif "사원" in h and "명" in h:
                        col_map["employee_name"] = c_idx
                    elif "프로젝트" in h:
                        col_map["project_name"] = c_idx
                break

        if header_row is None:
            raise HTTPException(status_code=400, detail="헤더 행을 찾을 수 없습니다. (확정일/거래처/금액 컬럼 필요)")

        # 데이터 행 수집
        def _safe(val):
            if val is None:
                return ""
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()

        def _parse_amount(val):
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return int(val)
            s = re.sub(r"[^\d\-]", "", str(val))
            return int(s) if s else 0

        records = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
            cells = [cell.value for cell in row]
            # 금액 또는 거래처명 없으면 빈 행 — 건너뜀
            amt_val = cells[col_map["amount"]] if "amount" in col_map else None
            vendor_val = cells[col_map["vendor_name"]] if "vendor_name" in col_map else None
            if not amt_val and not vendor_val:
                continue

            # 합계 행 제외 (거래처 '-' 또는 자금과목 '합계')
            vendor_str = str(vendor_val or "").strip()
            fund_cat = str(cells[col_map["fund_category"]] or "").strip() if "fund_category" in col_map else ""
            if fund_cat == "합계" or (vendor_str in ("-", "") and fund_cat in ("-", "")):
                continue

            rec = {}
            for field, c_idx in col_map.items():
                if field == "amount":
                    rec[field] = _parse_amount(cells[c_idx])
                else:
                    rec[field] = _safe(cells[c_idx])
            records.append(rec)

        wb.close()

        if not records:
            return JSONResponse({"success": True, "count": 0, "message": "임포트할 데이터가 없습니다."})

        result = db.save_payment_history(records, project_id=project_id)
        return JSONResponse({"success": True, "count": len(records), "message": f"이체내역 {len(records)}건 임포트 완료"})

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 파싱 오류: {str(e)}")


@router.get("/api/fund/projects/{project_id}/budget")
async def list_budget(project_id: int, request: Request, year: Optional[int] = None):
    """예실대비현황 조회"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    budget = db.list_budget_actual(project_id=project_id, year=year)
    return JSONResponse({"budget": budget})


@router.get("/api/fund/projects/{project_id}/summary")
async def get_fund_summary(project_id: int, request: Request):
    """프로젝트 자금현황 요약"""
    require_auth(request)
    summary = db.get_fund_summary(project_id)
    if "error" in summary:
        raise HTTPException(status_code=404, detail=summary["error"])
    return JSONResponse({"summary": summary})


# ─────────────────────────────────────────
# TODO CRUD
# ─────────────────────────────────────────

@router.get("/api/fund/todos")
async def list_all_todos(request: Request):
    """전체 TODO 조회"""
    require_auth(request)
    todos = db.list_todos()
    return JSONResponse({"todos": todos})


@router.get("/api/fund/projects/{project_id}/todos")
async def list_project_todos(project_id: int, request: Request):
    """프로젝트별 TODO 조회"""
    require_auth(request)
    todos = db.list_todos(project_id=project_id)
    return JSONResponse({"todos": todos})


class TodoCreate(BaseModel):
    content: str
    priority: str = "medium"
    category: str = ""
    project_id: Optional[int] = None


@router.post("/api/fund/todos")
async def create_todo(request: Request, body: TodoCreate):
    """TODO 생성"""
    require_auth(request)
    result = db.create_todo(
        project_id=body.project_id,
        content=body.content,
        priority=body.priority,
        category=body.category,
    )
    return JSONResponse(result)


class TodoUpdate(BaseModel):
    content: Optional[str] = None
    completed: Optional[int] = None
    priority: Optional[str] = None
    category: Optional[str] = None
    project_id: Optional[int] = None


@router.put("/api/fund/todos/{todo_id}")
async def update_todo(todo_id: int, request: Request, body: TodoUpdate):
    """TODO 수정"""
    user = require_auth(request)
    # 소유자 검증 — todo_id로 project_id 조회 후 확인
    conn = db.get_db()
    try:
        row = conn.execute("SELECT project_id FROM project_todos WHERE id = ?", (todo_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="TODO를 찾을 수 없습니다.")
        if row["project_id"]:
            _check_owner(row["project_id"], user)
    finally:
        conn.close()
    kwargs = {k: v for k, v in body.model_dump().items() if v is not None}
    result = db.update_todo(todo_id, **kwargs)
    return JSONResponse(result)


@router.delete("/api/fund/todos/{todo_id}")
async def delete_todo(todo_id: int, request: Request):
    """TODO 삭제"""
    user = require_auth(request)
    # 소유자 검증 — todo_id로 project_id 조회 후 확인
    conn = db.get_db()
    try:
        row = conn.execute("SELECT project_id FROM project_todos WHERE id = ?", (todo_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="TODO를 찾을 수 없습니다.")
        if row["project_id"]:
            _check_owner(row["project_id"], user)
    finally:
        conn.close()
    result = db.delete_todo(todo_id)
    return JSONResponse(result)


# ─────────────────────────────────────────
# AI 인사이트 생성
# ─────────────────────────────────────────

@router.get("/api/fund/insights")
async def get_insights(request: Request):
    """저장된 인사이트 조회"""
    require_auth(request)
    insights = db.get_insights()
    return JSONResponse({"insights": insights})


@router.get("/api/fund/portfolio-analysis")
async def get_portfolio_analysis(request: Request):
    """포트폴리오 분석 결과 조회 (캐시된 인사이트)"""
    require_auth(request)
    conn = db.get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM project_insights WHERE project_id IS NULL ORDER BY generated_at DESC"
        ).fetchall()
        insights = [dict(r) for r in rows]
        # 프로젝트별 인사이트도 가져와서 요약 카드 구성
        proj_rows = conn.execute(
            "SELECT i.*, p.name as project_name FROM project_insights i "
            "JOIN projects p ON i.project_id = p.id "
            "ORDER BY p.sort_order ASC, i.generated_at DESC"
        ).fetchall()
        project_insights = {}
        for r in proj_rows:
            pid = r["project_id"]
            if pid not in project_insights:
                project_insights[pid] = {"project_name": r["project_name"], "items": []}
            project_insights[pid]["items"].append({
                "type": r["insight_type"],
                "content": r["content"],
                "generated_at": r["generated_at"]
            })
        return JSONResponse({
            "portfolio": insights,
            "projects": project_insights
        })
    finally:
        conn.close()


@router.post("/api/fund/insights/generate")
async def generate_insights(request: Request):
    """Claude Opus로 전체 프로젝트 인사이트 생성"""
    require_auth(request)

    import asyncio
    projects = db.get_all_projects_full_data()
    if not projects:
        return JSONResponse({"success": False, "error": "프로젝트가 없습니다."})

    try:
        result = await asyncio.to_thread(_call_claude_for_insights, projects)
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"인사이트 생성 실패: {e}", exc_info=True)
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


def _format_materials_for_prompt(materials: list[dict]) -> str:
    """자료 목록을 프롬프트용 텍스트로 변환"""
    if not materials:
        return ""
    lines = []
    for m in materials[:10]:  # 최대 10개
        if m.get("material_type") == "text" and m.get("content_text"):
            lines.append(f"\n  - [메모] {m.get('description', '')}: {m['content_text'][:200]}")
        elif m.get("material_type") == "file":
            lines.append(f"\n  - [파일] {m.get('file_name', m.get('description', ''))}")
    return "".join(lines)


def _build_insights_prompt(projects: list[dict]) -> tuple[str, list[dict]]:
    """프로젝트 데이터를 인사이트 프롬프트로 변환"""
    from datetime import date
    project_summaries = []
    for p in projects:
        ov = p.get("overview", {})
        milestones = p.get("milestones", [])
        subs = p.get("subcontracts", [])
        colls = p.get("collections", [])
        issues = p.get("issues", [])
        todos = p.get("todos", [])

        total_order = (p.get("design_amount", 0) or 0) + (p.get("construction_amount", 0) or 0)
        ms_total = len(milestones)
        ms_done = sum(1 for m in milestones if m.get("completed"))

        coll_total = sum(c.get("amount", 0) for c in colls)
        coll_done = sum(c.get("amount", 0) for c in colls if c.get("collected"))

        sub_total_contract = sum(s.get("contract_amount", 0) for s in subs)
        sub_total_paid = 0
        for s in subs:
            for n in range(1, 5):
                if s.get(f"payment_{n}_confirmed"):
                    sub_total_paid += s.get(f"payment_{n}", 0) or 0

        summary = f"""
## [프로젝트 ID: {p['id']}] {p['name']} (등급: {p.get('grade', '-')}, 상태: {p.get('status', '-')})
- 수주액: {total_order:,}원 | 실행예산: {(p.get('execution_budget', 0) or 0):,}원 | 이익율: {p.get('profit_rate', 0)}%
- 위치: {ov.get('location', '-')} | 용도: {ov.get('usage', '-')} | 면적: {ov.get('area_pyeong', '-')}평
- 설계기간: {ov.get('design_start', '-')} ~ {ov.get('design_end', '-')}
- 시공기간: {ov.get('construction_start', '-')} ~ {ov.get('construction_end', '-')}
- 진행률: {ms_done}/{ms_total} 단계 완료
- 수금현황: {coll_done:,}/{coll_total:,}원 ({(coll_done/coll_total*100 if coll_total else 0):.1f}%)
- 하도급: {len(subs)}개 업체, 계약 {sub_total_contract:,}원, 기지급 {sub_total_paid:,}원
- 배정인원: {', '.join(f"{m.get('role', '')}:{m.get('name', '')}" for m in p.get('members', []))}
- 이슈: {'; '.join(issues) if issues else '없음'}
- 기존 TODO: {len(todos)}건 ({sum(1 for t in todos if t.get('completed'))}건 완료)
- 사용자 제공 자료: {len(p.get('materials', []))}건{_format_materials_for_prompt(p.get('materials', []))}
"""
        project_summaries.append({"id": p["id"], "name": p["name"], "text": summary})

    all_text = "\n".join(s["text"] for s in project_summaries)

    prompt = f"""당신은 건설/인테리어 프로젝트 관리에 20년 경력의 전문 컨설턴트입니다.
글로우서울(음향/인테리어 시공 전문)의 프로젝트 현황을 분석해주세요.

## 분석 관점 — 사람들이 쉽게 놓치는 것 위주로
당신의 역할은 일반적인 요약이 아니라, **담당자가 바쁘게 일하면서 미처 챙기지 못하는 포인트**를 짚어주는 것입니다:

- 수금 일정과 하도급 지급 일정의 **타이밍 불일치** (돈이 나가는데 들어오는 건 늦는 구간)
- 이익률이 양호해 보여도 **실집행 기준으로 이미 초과**된 항목
- 진행률 대비 수금이 늦거나, 반대로 선수금을 받았는데 착공이 안 된 경우
- 하도급 계약은 했지만 **지급 확인이 안 된 건** (분쟁 소지)
- 여러 프로젝트가 동시에 자금이 필요한 **자금 병목 시점**
- 데이터가 비어있어서 **판단 자체가 불가능한 항목** (이것이 가장 위험)

## 현재 프로젝트 현황
{all_text}

## 요청사항
각 프로젝트별로 아래 JSON 형식으로 반환해주세요:

1. **현황 점검** (strategy): 현재 프로젝트 상태를 3~5줄로 요약. 잘 가고 있는 점과 우려되는 점을 균형있게. 숫자 근거를 반드시 포함.

2. **놓치기 쉬운 리스크** (risk): 담당자가 바쁘면 지나치기 쉬운 리스크 위주. 예시:
   - "수금 계약금 0원 미수령 상태에서 하도급 지급이 진행 중"
   - "시공 종료일이 다가오는데 수금 잔금 일정이 미확정"
   - "하도급 3차 지급 미확인 — 업체와 확인 필요"
   뻔한 리스크("예산 초과 주의")보다는 구체적 상황을 지적해주세요.

3. **자금 흐름 분석** (profitability): 이익률뿐 아니라 실제 현금 흐름 관점에서 분석.
   - 수금된 금액 vs 기지급 금액 → 순현금 포지션
   - 향후 수금 예정 vs 지급 예정 → 자금 여유/부족 예측
   - 하도급 비율이 수주액 대비 몇 %인지, 직영 마진은 적정한지

4. **지금 해야 할 일** (action): 즉시 실행 가능한 구체적 액션. "~를 검토하세요" 같은 애매한 표현 대신:
   - "A업체 3차 지급 50,000,000원 확인서 수령 필요"
   - "발주처에 2차 중도금 500,000,000원 수금 요청 발송"
   - "시공 종료일 확정 후 잔금 수금 일정 협의"
   처럼 금액, 대상, 행동을 명시해주세요.

5. **필요한 정보 & TODO** (missing_data_todos): 정확한 분석을 위해 시스템에 입력이 필요한 항목.
   값이 '-', 0, 빈칸인 필드를 찾아서 TODO로 작성. 형식: "할 일 (이유)"
   예시:
   - "시공 시작일/종료일 입력 (일정 지연 여부 판단 불가)"
   - "계약서 PDF를 자료실에 업로드 (계약 조건 검토 필요)"
   - "수금 계약금 금액 입력 (현금흐름 분석 누락)"
   - "하도급 잔여 지급 스케줄 등록 (자금 계획 수립에 필요)"

6. **전체 포트폴리오 인사이트** (portfolio):
   - 전체 프로젝트를 동시에 보았을 때 발견되는 패턴
   - 어떤 프로젝트에 자금/인력을 우선 집중해야 하는지
   - 자금 병목 시점 예측 (수금 일정 vs 지급 일정 겹침)
   - 회사 전체의 현금 포지션 (총 수금액 vs 총 기지급 vs 미지급)

오늘 날짜: {date.today().isoformat()}

**규칙**:
- project_id는 현황에 표시된 실제 프로젝트 ID를 그대로 사용. project_name도 정확히 일치.
- 한국어로 작성. 금액은 원 단위 + 천단위 콤마.
- 숫자 근거 없는 추상적 조언은 금지. 반드시 데이터에서 근거를 찾아 인용.
- missing_data_todos가 있으면 해당 인사이트 본문에도 "⚠ 일부 정보가 미입력 상태입니다. TODO 리스트를 확인해주세요." 포함.
- **간결하게 작성**: 각 항목(strategy/risk/profitability/action)은 2~4문장 이내. 데이터가 부족한 프로젝트는 1~2문장으로 축소.
- 데이터가 거의 없는 프로젝트(수주액 0, 하도급 0, 수금 0)는 missing_data_todos만 작성하고 나머지는 "데이터 부족으로 분석 불가"로 통일.

반환 형식 (JSON만 반환):
```json
{{
  "projects": [
    {{
      "project_id": 실제_프로젝트_ID,
      "project_name": "정확한 프로젝트명",
      "strategy": "현황 점검 내용...",
      "risk": "놓치기 쉬운 리스크...",
      "profitability": "자금 흐름 분석...",
      "action": "지금 해야 할 일...",
      "missing_data_todos": ["할 일 (이유)", "..."]
    }}
  ],
  "portfolio": "전체 포트폴리오 인사이트..."
}}
```"""
    return prompt, project_summaries


def _parse_and_save_insights(raw_text: str, project_summaries: list[dict] = None) -> dict:
    """AI 응답 텍스트를 파싱하고 DB에 저장
    project_summaries: [{"id": 실제DB_ID, "name": "프로젝트명"}, ...] — 이름→ID 매핑용
    """
    import json, re
    # JSON 블록 추출 (다양한 형식 대응)
    if "```json" in raw_text:
        raw_text = raw_text.split("```json")[1].split("```")[0].strip()
    elif "```" in raw_text:
        raw_text = raw_text.split("```")[1].split("```")[0].strip()

    # JSON 객체 경계 탐지 (앞뒤 불필요한 텍스트 제거)
    match = re.search(r'\{[\s\S]*\}', raw_text)
    if match:
        raw_text = match.group(0)

    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError:
        logger.warning("인사이트 JSON 파싱 실패. raw_text 앞 500자: %s", raw_text[:500])
        db.save_insight(project_id=0, content=raw_text, insight_type="portfolio")
        return {"success": True, "raw": raw_text[:1000], "parsed": False}

    # 실제 DB ID 매핑 (이름 기반 + 순서 기반 fallback)
    name_to_id = {}
    idx_to_id = {}
    valid_ids = set()
    if project_summaries:
        for i, ps in enumerate(project_summaries):
            name_to_id[ps["name"]] = ps["id"]
            idx_to_id[i + 1] = ps["id"]  # Gemini가 1부터 순번 매기는 경우 대비
            valid_ids.add(ps["id"])

    # 프로젝트별 인사이트 저장
    for pi in data.get("projects", []):
        raw_pid = pi.get("project_id")
        pname = pi.get("project_name", "")
        # 1) 이름으로 매핑 시도
        pid = name_to_id.get(pname)
        # 2) 이름 부분 매칭
        if not pid and pname:
            for db_name, db_id in name_to_id.items():
                if pname in db_name or db_name in pname:
                    pid = db_id
                    break
        # 3) raw_pid가 실제 DB에 존재하면 사용
        if not pid and raw_pid in valid_ids:
            pid = raw_pid
        # 4) raw_pid를 순번으로 간주하여 매핑
        if not pid and raw_pid:
            pid = idx_to_id.get(raw_pid)
        if not pid:
            continue
        for itype in ["strategy", "risk", "profitability", "action"]:
            content = pi.get(itype, "")
            if content:
                db.save_insight(project_id=pid, content=content, insight_type=itype)

        # 누락 정보 TODO 자동 생성
        missing_todos = pi.get("missing_data_todos", [])
        if missing_todos and pid:
            for todo_text in missing_todos:
                if not todo_text or not isinstance(todo_text, str):
                    continue
                # 이미 동일한 TODO가 있는지 확인 (중복 방지)
                existing = db.list_todos(project_id=pid)
                already_exists = any(
                    todo_text.strip() in t.get("content", "") or t.get("content", "") in todo_text.strip()
                    for t in existing if not t.get("completed")
                )
                if not already_exists:
                    db.create_todo(
                        project_id=pid,
                        content=f"[AI 권고] {todo_text.strip()}",
                        priority="medium",
                        category="누락정보"
                    )

    # 포트폴리오 인사이트 저장 (project_id=0)
    portfolio = data.get("portfolio", "")
    if portfolio:
        db.save_insight(project_id=0, content=portfolio, insight_type="portfolio")

    return {"success": True, "data": data, "parsed": True}


def _call_claude_for_insights(projects: list[dict]) -> dict:
    """Gemini API로 인사이트 생성 (Anthropic API 대안)"""
    from google import genai

    gemini_key = os.environ.get("GEMINI_API_KEY")
    if not gemini_key:
        raise RuntimeError("GEMINI_API_KEY가 설정되지 않았습니다.")

    client = genai.Client(api_key=gemini_key)

    # 활성 프로젝트만 필터 (완료/취소 제외 — 응답 토큰 초과 방지)
    active_projects = [p for p in projects if p.get("status", "") not in ("완료", "취소", "중단")]
    if not active_projects:
        active_projects = projects[:10]  # fallback
    prompt, project_summaries = _build_insights_prompt(active_projects)

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config={
            "response_mime_type": "application/json",
            "temperature": 0.7,
            "max_output_tokens": 65536,
        },
    )

    raw_text = response.text.strip()
    return _parse_and_save_insights(raw_text, project_summaries)


# ─────────────────────────────────────────
# 프로젝트 관리 웹 페이지 서빙
# ─────────────────────────────────────────

@router.get("/fund")
async def serve_fund_page(request: Request):
    """프로젝트 관리표 웹 페이지 서빙"""
    require_auth(request)
    fund_html = STATIC_DIR / "fund.html"
    if not fund_html.exists():
        raise HTTPException(status_code=404, detail="fund.html을 찾을 수 없습니다.")
    return FileResponse(str(fund_html))


@router.get("/guide")
async def serve_guide_page(request: Request):
    """사용방법 가이드 페이지 서빙"""
    require_auth(request)
    guide_html = STATIC_DIR / "guide.html"
    if not guide_html.exists():
        raise HTTPException(status_code=404, detail="guide.html을 찾을 수 없습니다.")
    return FileResponse(str(guide_html))


@router.get("/insights")
async def serve_insights_page(request: Request):
    """인사이트 페이지 → fund 페이지로 리다이렉트"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/fund")


# ─────────────────────────────────────────
# 자료실 (Materials) 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/materials")
async def list_materials(project_id: int, request: Request):
    """프로젝트 자료 목록"""
    require_auth(request)
    materials = db.list_materials(project_id)
    return JSONResponse({"materials": materials})


class MaterialCreate(BaseModel):
    material_type: str = "text"
    description: str = ""
    content_text: str = ""

@router.post("/api/fund/projects/{project_id}/materials")
async def add_material_text(project_id: int, request: Request, body: MaterialCreate):
    """텍스트/메모 자료 추가"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    result = db.add_material(
        project_id=project_id,
        material_type=body.material_type,
        description=body.description,
        content_text=body.content_text,
    )
    return JSONResponse(result)


@router.post("/api/fund/projects/{project_id}/materials/upload")
async def upload_material_file(project_id: int, request: Request):
    """파일 업로드 자료 추가"""
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="파일이 없습니다.")

    # 저장 디렉토리
    materials_dir = Path(__file__).parent.parent.parent / "data" / "materials" / str(project_id)
    materials_dir.mkdir(parents=True, exist_ok=True)

    # 안전한 파일명
    import re
    safe_name = re.sub(r'[^\w\-_\. ]', '_', file.filename)
    file_path = materials_dir / safe_name
    content = await file.read()
    file_path.write_bytes(content)

    relative_path = f"data/materials/{project_id}/{safe_name}"
    description = form.get("description", file.filename)

    result = db.add_material(
        project_id=project_id,
        material_type="file",
        file_name=file.filename,
        file_path=relative_path,
        mime_type=file.content_type or "",
        description=description,
    )
    return JSONResponse(result)


@router.delete("/api/fund/materials/{material_id}")
async def delete_material(material_id: int, request: Request):
    """자료 삭제"""
    require_auth(request)
    result = db.delete_material(material_id)
    if not result["success"]:
        raise HTTPException(status_code=404, detail=result["message"])
    return JSONResponse(result)


# ─────────────────────────────────────────
# 알림 엔드포인트
# ─────────────────────────────────────────

_last_notification_check: float = 0.0  # 마지막 알림 체크 시각 (epoch)

@router.get("/api/fund/notifications")
async def get_notifications(request: Request):
    """알림 목록"""
    import time
    global _last_notification_check
    require_auth(request)
    # 자동 알림 생성: 최대 5분에 1회만 실행 (매 GET마다 DB 쓰기 방지)
    if time.monotonic() - _last_notification_check > 300:
        db.check_and_generate_notifications()
        _last_notification_check = time.monotonic()
    notifications = db.list_notifications()
    return JSONResponse({"notifications": notifications})


@router.post("/api/fund/notifications/read-all")
async def mark_all_read(request: Request):
    """모든 알림 읽음"""
    require_auth(request)
    result = db.mark_notifications_read()
    return JSONResponse(result)


# ─────────────────────────────────────────
# 프로젝트 별칭 (alias) 관리 API
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/aliases")
async def get_aliases(project_id: int, request: Request):
    """프로젝트 별칭 목록"""
    require_auth(request)
    aliases = db.get_project_aliases(project_id)
    return JSONResponse({"aliases": aliases})


@router.post("/api/fund/projects/{project_id}/aliases")
async def add_alias(project_id: int, request: Request):
    """프로젝트 별칭 추가"""
    require_auth(request)
    body = await request.json()
    alias = body.get("alias", "").strip()
    if not alias:
        return JSONResponse({"error": "별칭을 입력해주세요"}, status_code=400)
    result = db.add_project_alias(project_id, alias, alias_type="manual")
    return JSONResponse(result)


@router.delete("/api/fund/projects/{project_id}/aliases")
async def delete_alias(project_id: int, request: Request):
    """프로젝트 별칭 삭제"""
    require_auth(request)
    body = await request.json()
    alias = body.get("alias", "").strip()
    if not alias:
        return JSONResponse({"error": "삭제할 별칭을 지정해주세요"}, status_code=400)
    result = db.remove_project_alias(project_id, alias)
    return JSONResponse(result)


@router.get("/api/fund/aliases")
async def get_all_project_aliases(request: Request):
    """전체 프로젝트 별칭 목록"""
    require_auth(request)
    aliases = db.get_all_aliases()
    return JSONResponse({"aliases": aliases})


# ─────────────────────────────────────────
# GW 예산 실적 (budget_actual) — 상세 조회 + 동기화
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/budget/detail")
async def list_budget_detail(project_id: int, request: Request,
                              gisu: Optional[int] = None,
                              leaf_only: bool = False):
    """예실대비현황(상세) — GW DataProvider 기반 실적 데이터 조회
    leaf_only=true이면 말단(최하위) 항목만 반환
    """
    require_auth(request)
    if not db.get_project(project_id):
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")
    conn = db.get_db()
    try:
        query = "SELECT * FROM budget_actual WHERE project_id = ?"
        params = [project_id]
        if gisu:
            query += " AND gisu = ?"
            params.append(gisu)
        if leaf_only:
            query += " AND is_leaf = 1"
        query += " ORDER BY budget_code ASC"
        rows = conn.execute(query, params).fetchall()
        data = [dict(r) for r in rows]
    finally:
        conn.close()

    # 집계 계산
    total_budget = sum(r.get("budget_amount", 0) for r in data if r.get("is_leaf", 1))
    total_actual = sum(r.get("actual_amount", 0) for r in data if r.get("is_leaf", 1))
    total_diff   = total_budget - total_actual
    overall_rate = round(total_actual / total_budget * 100, 2) if total_budget else 0

    return JSONResponse({
        "budget_detail": data,
        "summary": {
            "total_budget": total_budget,
            "total_actual": total_actual,
            "total_diff": total_diff,
            "overall_rate": overall_rate,
            "item_count": len(data),
        }
    })


@router.post("/api/fund/projects/{project_id}/budget/sync-actuals")
async def sync_budget_actuals(project_id: int, request: Request):
    """단일 프로젝트 예실대비현황 GW 동기화 (RealGrid DataProvider 방식)"""
    user = require_auth(request)
    # gw_id는 인증된 사용자에서 파생 (클라이언트 임의 지정 불가)
    gw_id = user.get("gw_id", "")
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    body = await request.json()
    project_code = body.get("project_code") or project.get("project_code", "")

    if not gw_id:
        raise HTTPException(status_code=400, detail="GW 계정 정보를 찾을 수 없습니다.")
    if not project_code:
        raise HTTPException(status_code=400, detail="프로젝트 코드(GS-XX-XXXX)가 필요합니다.")

    results = {"status": "started", "project_id": project_id, "project_code": project_code}

    def _sync():
        try:
            from src.fund_table.budget_crawler import crawl_budget_actual
            r = crawl_budget_actual(gw_id, project_id=project_id, project_code=project_code)
            logger.info(f"예실대비 동기화 완료: {r}")
        except Exception as e:
            logger.error(f"예실대비 동기화 실패: {e}", exc_info=True)

    threading.Thread(target=_sync, daemon=True).start()
    return JSONResponse({"message": "예실대비현황 GW 동기화 시작됨", **results})


@router.post("/api/fund/gw/sync-all-budget-actuals")
async def sync_all_budget_actuals(request: Request):
    """모든 프로젝트 예실대비현황 일괄 GW 동기화 (백그라운드)"""
    user = require_auth(request)
    # gw_id는 인증된 사용자에서 파생 (클라이언트 임의 지정 불가)
    gw_id = user.get("gw_id", "")
    if not gw_id:
        raise HTTPException(status_code=400, detail="GW 계정 정보를 찾을 수 없습니다.")

    def _sync_all():
        try:
            from src.fund_table.budget_crawler import crawl_all_projects
            r = crawl_all_projects(gw_id)
            logger.info(f"전체 예실대비 동기화 완료: {r.get('message')}")
        except Exception as e:
            logger.error(f"전체 예실대비 동기화 실패: {e}", exc_info=True)

    threading.Thread(target=_sync_all, daemon=True).start()
    return JSONResponse({"message": "전체 프로젝트 예실대비현황 동기화 시작됨"})


@router.get("/api/fund/gw/project-list")
async def get_gw_project_list(request: Request, keyword: str = ""):
    """GW 전체 프로젝트 목록 반환 (캐시 우선, 없으면 DB 프로젝트 목록)"""
    require_auth(request)
    # GW 캐시에서 먼저 조회
    cached = db.search_gw_projects_cache(keyword) if keyword else []
    if not cached:
        # DB 프로젝트 목록으로 fallback
        projects = db.list_projects()
        cached = [
            {"code": p.get("project_code",""), "name": p.get("name",""), "id": p.get("id")}
            for p in projects if p.get("project_code")
        ]
    return JSONResponse({"projects": cached, "count": len(cached)})


@router.get("/api/fund/budget/cross-project")
async def cross_project_budget(request: Request, gisu: Optional[int] = None, top_n: int = 20):
    """전체 프로젝트 예실대비 집계 (집행률 상위 N 반환)"""
    require_auth(request)
    top_n = min(max(1, top_n), 100)  # 1~100 범위로 제한
    conn = db.get_db()
    try:
        q = """
        SELECT
            gw_project_code,
            project_name,
            SUM(CASE WHEN is_leaf=1 THEN budget_amount ELSE 0 END) AS total_budget,
            SUM(CASE WHEN is_leaf=1 THEN actual_amount ELSE 0 END) AS total_actual,
            SUM(CASE WHEN is_leaf=1 THEN difference   ELSE 0 END) AS total_diff,
            MAX(scraped_at) AS last_synced
        FROM budget_actual
        WHERE gw_project_code != ''
        """
        params = []
        if gisu:
            q += " AND gisu = ?"
            params.append(gisu)
        q += " GROUP BY gw_project_code ORDER BY total_actual DESC LIMIT ?"
        params.append(top_n)
        rows = conn.execute(q, params).fetchall()
        data = []
        for r in rows:
            row = dict(r)
            row["execution_rate"] = (
                round(row["total_actual"] / row["total_budget"] * 100, 1)
                if row["total_budget"] else 0
            )
            data.append(row)
    finally:
        conn.close()
    return JSONResponse({"projects": data, "count": len(data)})


# ─────────────────────────────────────────
# Pydantic 요청 모델 — 신규 테이블용
# ─────────────────────────────────────────

class CollectionScheduleCreate(BaseModel):
    """수금 예정 직접 추가 요청 (수동 입력용)"""
    item_name: str
    scheduled_date: str
    amount: int
    status: Optional[str] = "pending"


class RiskCreate(BaseModel):
    """리스크 항목 추가 요청"""
    risk_type: str
    severity: str
    description: str
    source: Optional[str] = ""


class RiskUpdate(BaseModel):
    """리스크 항목 수정 요청"""
    is_resolved: Optional[bool] = None
    action_taken: Optional[str] = None


# ─────────────────────────────────────────
# 세금계산서 (gw_tax_invoices)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/tax-invoices")
async def get_tax_invoices(project_id: int, request: Request):
    """프로젝트별 세금계산서 목록 조회"""
    require_auth(request)
    try:
        data = db.list_tax_invoices(project_id=project_id)
        return JSONResponse({"tax_invoices": data, "count": len(data)})
    except Exception as e:
        logger.error(f"세금계산서 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="세금계산서 조회 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# 예산 변경 이력 (gw_budget_changes)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/budget-changes")
async def get_budget_changes(project_id: int, request: Request):
    """프로젝트별 예산 변경 이력 조회"""
    require_auth(request)
    try:
        data = db.list_budget_changes(project_id=project_id)
        return JSONResponse({"budget_changes": data, "count": len(data)})
    except Exception as e:
        logger.error(f"예산 변경 이력 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="예산 변경 이력 조회 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# 수금 예정 내역 (gw_collection_schedule)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/collection-schedule")
async def get_collection_schedule(project_id: int, request: Request, status: Optional[str] = None):
    """프로젝트별 수금 예정 내역 조회"""
    require_auth(request)
    try:
        data = db.list_collection_schedule(project_id=project_id, status=status)
        return JSONResponse({"collection_schedule": data, "count": len(data)})
    except Exception as e:
        logger.error(f"수금 예정 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="수금 예정 조회 중 오류가 발생했습니다.")


@router.post("/api/fund/projects/{project_id}/collection-schedule")
async def create_collection_schedule(project_id: int, req: CollectionScheduleCreate, request: Request):
    """수금 예정 항목 직접 추가 (수동 입력용)"""
    require_auth(request)
    try:
        result = db.add_collection_schedule(
            project_id=project_id,
            item_name=req.item_name,
            scheduled_date=req.scheduled_date,
            amount=req.amount,
            status=req.status,
        )
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"수금 예정 추가 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="수금 예정 추가 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# 자금집행 승인 현황 (gw_payment_approvals)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/payment-approvals")
async def get_payment_approvals(project_id: int, request: Request, status: Optional[str] = None):
    """프로젝트별 자금집행 승인 현황 조회"""
    require_auth(request)
    try:
        data = db.list_payment_approvals(project_id=project_id, status=status)
        return JSONResponse({"payment_approvals": data, "count": len(data)})
    except Exception as e:
        logger.error(f"자금집행 승인 현황 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="자금집행 승인 현황 조회 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# 리스크 관리 (project_risk_log)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/risks")
async def get_risks(project_id: int, request: Request, status: Optional[str] = None):
    """프로젝트별 리스크 이력 조회"""
    require_auth(request)
    try:
        data = db.list_risks(project_id=project_id, status=status)
        return JSONResponse({"risks": data, "count": len(data)})
    except Exception as e:
        logger.error(f"리스크 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="리스크 조회 중 오류가 발생했습니다.")


@router.post("/api/fund/projects/{project_id}/risks")
async def create_risk(project_id: int, req: RiskCreate, request: Request):
    """프로젝트별 리스크 항목 추가"""
    require_auth(request)
    try:
        result = db.add_risk(
            project_id=project_id,
            risk_type=req.risk_type,
            severity=req.severity,
            description=req.description,
            created_by=req.source,
        )
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"리스크 추가 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="리스크 추가 중 오류가 발생했습니다.")


@router.put("/api/fund/risks/{risk_id}")
async def update_risk(risk_id: int, req: RiskUpdate, request: Request):
    """리스크 항목 수정 (해결 여부, 조치 내용 업데이트)"""
    require_auth(request)
    try:
        kwargs = {}
        # is_resolved=True 이면 status를 "resolved"로 변경
        if req.is_resolved is not None:
            kwargs["status"] = "resolved" if req.is_resolved else "open"
        # action_taken은 mitigation 필드에 저장
        if req.action_taken is not None:
            kwargs["mitigation"] = req.action_taken
        result = db.update_risk(risk_id=risk_id, **kwargs)
        if not result["success"]:
            raise HTTPException(status_code=400, detail=result["message"])
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"리스크 수정 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="리스크 수정 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# GW 계약 현황 (gw_contracts)
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/gw-contracts")
async def get_gw_contracts(project_id: int, request: Request):
    """프로젝트별 GW 계약 현황 조회"""
    require_auth(request)
    try:
        data = db.list_gw_contracts(project_id=project_id)
        return JSONResponse({"gw_contracts": data, "count": len(data)})
    except Exception as e:
        logger.error(f"GW 계약 현황 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="GW 계약 현황 조회 중 오류가 발생했습니다.")


# ─────────────────────────────────────────
# 세금계산서 간편 동기화 (D4 요청)
# ─────────────────────────────────────────

@router.post("/api/fund/projects/{project_id}/tax-invoices/sync")
async def sync_tax_invoices_simple(project_id: int, request: Request):
    """
    단일 프로젝트 세금계산서 발행 내역 GW 동기화 (간편 경로).
    인증된 사용자의 gw_id를 자동 사용 — gw_id를 body에 별도 전달하지 않아도 됨.
    """
    import threading
    user = require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    gw_id = user["gw_id"]
    project_code = project.get("project_code", "")
    if not project_code:
        raise HTTPException(
            status_code=400,
            detail="프로젝트 코드(GS-XX-XXXX)가 설정되지 않았습니다. 개요 탭에서 GW 사업코드를 먼저 입력해주세요.",
        )

    result: dict = {}

    def _sync():
        try:
            from src.fund_table.tax_invoice_crawler import TaxInvoiceCrawler
            from playwright.sync_api import sync_playwright
            from src.auth.login import login_and_get_context, close_session
            from src.auth.user_db import get_decrypted_password

            gw_pw = get_decrypted_password(gw_id)
            if not gw_pw:
                result["error"] = "비밀번호 복호화 실패"
                return

            pw = sync_playwright().start()
            try:
                browser, context, page = login_and_get_context(
                    playwright_instance=pw,
                    headless=True,
                    user_id=gw_id,
                    user_pw=gw_pw,
                )
                crawler = TaxInvoiceCrawler(gw_id=gw_id, encrypted_pw="")
                invoices = crawler._navigate_and_extract(page, project_code, year=None)
                if invoices:
                    for inv in invoices:
                        inv["project_id"] = project_id
                    db.save_tax_invoices(invoices, project_id=project_id)
                result["count"] = len(invoices) if invoices else 0
                close_session(browser)
            finally:
                pw.stop()
        except Exception as e:
            logger.error(f"세금계산서 동기화 실패 (project_id={project_id}): {e}", exc_info=True)
            result["error"] = str(e)

    try:
        await asyncio.wait_for(asyncio.to_thread(_sync), timeout=300)
    except asyncio.TimeoutError:
        return JSONResponse({"success": False, "error": "GW 접속 시간 초과 (5분)"}, status_code=504)

    if result.get("error"):
        raise HTTPException(status_code=500, detail=result["error"])

    count = result.get("count", 0)
    return JSONResponse({
        "success": True,
        "count": count,
        "message": f"세금계산서 {count}건 동기화 완료",
    })


# ─────────────────────────────────────────
# GW 세금계산서 동기화
# ─────────────────────────────────────────

@router.post("/api/fund/projects/{project_id}/gw/sync-tax-invoices")
async def sync_tax_invoices(project_id: int, request: Request):
    """단일 프로젝트 세금계산서 발행 내역 GW 동기화 (백그라운드)"""
    require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    body = await request.json()
    gw_id = body.get("gw_id", "")
    project_code = body.get("project_code") or project.get("project_code", "")
    year = body.get("year")  # 없으면 None (전체 조회)

    if not gw_id:
        raise HTTPException(status_code=400, detail="gw_id가 필요합니다.")
    if not project_code:
        raise HTTPException(status_code=400, detail="프로젝트 코드(GS-XX-XXXX)가 필요합니다.")

    def _sync():
        import threading
        try:
            from src.fund_table.tax_invoice_crawler import TaxInvoiceCrawler
            from playwright.sync_api import sync_playwright
            from src.auth.login import login_and_get_context, close_session
            from src.auth.user_db import get_decrypted_password

            gw_pw = get_decrypted_password(gw_id)
            if not gw_pw:
                logger.error(f"세금계산서 동기화: 비밀번호 복호화 실패 gw_id={gw_id}")
                return

            pw = sync_playwright().start()
            try:
                browser, context, page = login_and_get_context(
                    playwright_instance=pw,
                    headless=True,
                    user_id=gw_id,
                    user_pw=gw_pw,
                )
                crawler = TaxInvoiceCrawler(gw_id=gw_id, encrypted_pw="")
                invoices = crawler._navigate_and_extract(page, project_code, year)
                if invoices:
                    for inv in invoices:
                        inv["project_id"] = project_id
                    result = db.save_tax_invoices(invoices, project_id=project_id)
                    logger.info(f"세금계산서 동기화 완료: {result}")
                else:
                    logger.info(f"세금계산서 동기화: 수집 결과 없음 ({project_code})")
                close_session(browser)
            finally:
                pw.stop()
        except Exception as e:
            logger.error(f"세금계산서 동기화 실패: {e}", exc_info=True)

    import threading
    threading.Thread(target=_sync, daemon=True).start()
    return JSONResponse({
        "message": "세금계산서 GW 동기화 시작됨",
        "project_id": project_id,
        "project_code": project_code,
        "year": year,
    })


# ─────────────────────────────────────────
# GW 수금 예정 내역 동기화
# ─────────────────────────────────────────

@router.post("/api/fund/projects/{project_id}/gw/sync-collection-schedule")
async def sync_collection_schedule(project_id: int, request: Request):
    """단일 프로젝트 수금 예정 내역 GW 동기화 (백그라운드)"""
    require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    body = await request.json()
    gw_id = body.get("gw_id", "")
    project_code = body.get("project_code") or project.get("project_code", "")

    if not gw_id:
        raise HTTPException(status_code=400, detail="gw_id가 필요합니다.")
    if not project_code:
        raise HTTPException(status_code=400, detail="프로젝트 코드(GS-XX-XXXX)가 필요합니다.")

    def _sync():
        try:
            from src.fund_table.collection_schedule_crawler import CollectionScheduleCrawler
            from playwright.sync_api import sync_playwright
            from src.auth.login import login_and_get_context, close_session
            from src.auth.user_db import get_decrypted_password

            gw_pw = get_decrypted_password(gw_id)
            if not gw_pw:
                logger.error(f"수금예정 동기화: 비밀번호 복호화 실패 gw_id={gw_id}")
                return

            pw = sync_playwright().start()
            try:
                browser, context, page = login_and_get_context(
                    playwright_instance=pw,
                    headless=True,
                    user_id=gw_id,
                    user_pw=gw_pw,
                )
                crawler = CollectionScheduleCrawler(gw_id=gw_id, encrypted_pw="")
                items = crawler._navigate_and_extract(page, project_code)
                if items:
                    for item in items:
                        item["project_id"] = project_id
                    result = db.save_collection_schedule(items, project_id=project_id)
                    logger.info(f"수금예정 동기화 완료: {result}")
                else:
                    logger.info(f"수금예정 동기화: 수집 결과 없음 ({project_code})")
                close_session(browser)
            finally:
                pw.stop()
        except Exception as e:
            logger.error(f"수금예정 동기화 실패: {e}", exc_info=True)

    import threading
    threading.Thread(target=_sync, daemon=True).start()
    return JSONResponse({
        "message": "수금 예정 내역 GW 동기화 시작됨",
        "project_id": project_id,
        "project_code": project_code,
    })


# ─────────────────────────────────────────
# GW 자금집행 승인 현황 동기화
# ─────────────────────────────────────────

@router.post("/api/fund/projects/{project_id}/gw/sync-payment-approvals")
async def sync_payment_approvals(project_id: int, request: Request):
    """단일 프로젝트 자금집행 승인 현황 GW 동기화 (백그라운드)"""
    require_auth(request)
    project = db.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다.")

    body = await request.json()
    gw_id = body.get("gw_id", "")
    project_code = body.get("project_code") or project.get("project_code", "")
    year = body.get("year")  # 없으면 None (전체 조회)

    if not gw_id:
        raise HTTPException(status_code=400, detail="gw_id가 필요합니다.")
    if not project_code:
        raise HTTPException(status_code=400, detail="프로젝트 코드(GS-XX-XXXX)가 필요합니다.")

    def _sync():
        try:
            from src.fund_table.payment_approval_crawler import PaymentApprovalCrawler
            from playwright.sync_api import sync_playwright
            from src.auth.login import login_and_get_context, close_session
            from src.auth.user_db import get_decrypted_password

            gw_pw = get_decrypted_password(gw_id)
            if not gw_pw:
                logger.error(f"자금집행승인 동기화: 비밀번호 복호화 실패 gw_id={gw_id}")
                return

            pw = sync_playwright().start()
            try:
                browser, context, page = login_and_get_context(
                    playwright_instance=pw,
                    headless=True,
                    user_id=gw_id,
                    user_pw=gw_pw,
                )
                crawler = PaymentApprovalCrawler(gw_id=gw_id, encrypted_pw="")
                approvals = crawler._navigate_and_extract(page, project_code, year)
                if approvals:
                    for approval in approvals:
                        approval["project_id"] = project_id
                    result = db.save_payment_approvals(approvals, project_id=project_id)
                    logger.info(f"자금집행승인 동기화 완료: {result}")
                else:
                    logger.info(f"자금집행승인 동기화: 수집 결과 없음 ({project_code})")
                close_session(browser)
            finally:
                pw.stop()
        except Exception as e:
            logger.error(f"자금집행승인 동기화 실패: {e}", exc_info=True)

    import threading
    threading.Thread(target=_sync, daemon=True).start()
    return JSONResponse({
        "message": "자금집행 승인 현황 GW 동기화 시작됨",
        "project_id": project_id,
        "project_code": project_code,
        "year": year,
    })


# ─────────────────────────────────────────
# 일정표 (공정 일정) 엔드포인트
# ─────────────────────────────────────────

@router.get("/api/fund/projects/{project_id}/schedule")
async def get_schedule(project_id: int, request: Request):
    """공정 일정 항목 조회"""
    require_auth(request)
    try:
        items = db.list_schedule_items(project_id)
        return JSONResponse({"items": items})
    except Exception as e:
        logger.error(f"일정 조회 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="일정 조회 실패")


@router.post("/api/fund/projects/{project_id}/schedule")
async def save_schedule(project_id: int, request: Request):
    """공정 일정 항목 저장 (전체 교체)"""
    user = require_auth(request)
    _check_owner(project_id, user)
    body = await request.json()
    items = body.get("items", [])
    result = db.save_schedule_items(project_id, items)
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("message", "저장 실패"))
    return JSONResponse(result)


@router.post("/api/fund/projects/{project_id}/archive")
async def archive_project(project_id: int, request: Request):
    """프로젝트 이전 프로젝트로 이동 (보관) / 복원"""
    user = require_auth(request)
    _check_owner(project_id, user)
    body = await request.json()
    is_archived = body.get("is_archived", True)
    result = db.set_project_archived(project_id, bool(is_archived))
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("message", "실패"))
    return JSONResponse(result)


@router.post("/api/fund/import-schedule-from-pm-sheet")
async def import_schedule_pm_sheet(request: Request):
    """PM Official 시트에서 프로젝트 일정 데이터를 가져와 일정표에 반영"""
    require_auth(request)
    body = await request.json()
    overwrite = body.get("overwrite", False)
    try:
        import asyncio
        from src.fund_table.sheets_import import import_schedule_from_pm_sheet
        result = await asyncio.to_thread(import_schedule_from_pm_sheet, overwrite=overwrite)
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"PM 시트 일정 임포트 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/fund/import-collections-from-pm-sheet")
async def import_collections_pm_sheet(request: Request):
    """PM Official 시트에서 모든 프로젝트 수금일정(수금현황) 가져오기"""
    require_auth(request)
    body = await request.json()
    overwrite = body.get("overwrite", True)
    try:
        import asyncio
        from src.fund_table.sheets_import import import_collections_from_pm_sheet
        result = await asyncio.to_thread(import_collections_from_pm_sheet, overwrite=overwrite)
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"수금일정 임포트 API 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ────────────────────────────────────────────
# 공정표 자동 생성 & 내보내기 API
# ────────────────────────────────────────────

@router.get("/api/fund/process-map/trades")
async def get_process_map_trades(request: Request, type: str = "오피스"):
    """공종 마스터 데이터 조회 (DB 우선, 폴백: 하드코딩)"""
    require_auth(request)
    trades = db.list_construction_trades()
    presets = db.list_construction_presets()

    if trades:
        # DB 데이터를 PROCESS_GROUPS 형태로 변환
        groups_map: dict[str, dict] = {}
        for t in trades:
            gn = t["group_name"]
            if gn not in groups_map:
                groups_map[gn] = {"group": gn, "color": t["group_color"], "items": []}
            groups_map[gn]["items"].append({
                "id": t["id"],
                "name": t["name"],
                "item_type": t.get("item_type", "bar"),
                "default_days": t.get("default_days", 0),
                "predecessors": json.loads(t["predecessors"]) if isinstance(t["predecessors"], str) else t.get("predecessors", []),
                "steps": json.loads(t["steps"]) if isinstance(t["steps"], str) else t.get("steps", []),
                "is_custom": t.get("is_custom", 0),
            })
        groups = list(groups_map.values())

        # 프리셋에서 해당 타입 찾기
        preset_trades = []
        for p in presets:
            if p["preset_name"] == type:
                preset_trades = p["trade_names"]
                break
        if not preset_trades:
            # 폴백: 전체 공종명
            preset_trades = [t["name"] for t in trades]
    else:
        # DB 비어있으면 하드코딩 폴백
        from src.fund_table.process_map_master import PROCESS_GROUPS, get_preset_trades
        groups = PROCESS_GROUPS
        preset_trades = get_preset_trades(type)

    return JSONResponse({
        "groups": groups,
        "preset_trades": preset_trades,
        "project_type": type,
        "presets": [{"name": p["preset_name"], "is_custom": p.get("is_custom", 0)} for p in presets],
    })


@router.post("/api/fund/process-map/trades")
async def add_process_map_trade(request: Request):
    """공종 추가"""
    require_auth(request)
    body = await request.json()
    name = body.get("name", "").strip()
    group_name = body.get("group_name", "").strip()
    if not name or not group_name:
        raise HTTPException(status_code=400, detail="공종명과 그룹명은 필수입니다.")
    result = db.add_construction_trade(group_name, name, **{
        k: body[k] for k in ("group_color", "item_type", "default_days",
                              "predecessors", "steps", "sort_order", "is_custom")
        if k in body
    })
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "추가 실패"))
    return JSONResponse(result)


@router.put("/api/fund/process-map/trades/{trade_id}")
async def update_process_map_trade(trade_id: int, request: Request):
    """공종 수정"""
    require_auth(request)
    body = await request.json()
    result = db.update_construction_trade(trade_id, **body)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "수정 실패"))
    return JSONResponse(result)


@router.delete("/api/fund/process-map/trades/{trade_id}")
async def delete_process_map_trade(trade_id: int, request: Request):
    """공종 삭제"""
    require_auth(request)
    result = db.delete_construction_trade(trade_id)
    if not result.get("success"):
        raise HTTPException(status_code=404, detail=result.get("message", "삭제 실패"))
    return JSONResponse(result)


@router.get("/api/fund/process-map/presets")
async def get_process_map_presets(request: Request):
    """프리셋 목록 조회"""
    require_auth(request)
    presets = db.list_construction_presets()
    return JSONResponse({"presets": presets})


@router.post("/api/fund/process-map/presets")
async def save_process_map_preset(request: Request):
    """프리셋 저장 (upsert)"""
    require_auth(request)
    body = await request.json()
    preset_name = body.get("preset_name", "").strip()
    trade_names = body.get("trade_names", [])
    if not preset_name:
        raise HTTPException(status_code=400, detail="프리셋명은 필수입니다.")
    result = db.save_construction_preset(
        preset_name, trade_names,
        is_custom=body.get("is_custom", 1)
    )
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "저장 실패"))
    return JSONResponse(result)


@router.post("/api/fund/process-map/parse-estimate")
async def parse_estimate(request: Request):
    """내역서 엑셀 파일에서 공종 자동 추출"""
    require_auth(request)
    import tempfile
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="파일이 첨부되지 않았습니다.")

    suffix = os.path.splitext(file.filename)[1] if file.filename else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        # 파일 크기 제한 (20MB)
        if len(content) > 20 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="파일 크기가 20MB를 초과합니다.")
        tmp.write(content)
        tmp_path = tmp.name

    try:
        import asyncio
        from src.fund_table.estimate_parser import parse_estimate_file
        result = await asyncio.to_thread(parse_estimate_file, tmp_path)
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"내역서 파싱 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@router.post("/api/fund/projects/{project_id}/generate-schedule")
async def generate_schedule(project_id: int, request: Request):
    """공정표 자동 생성 → DB 저장 + 엑셀/PDF 생성"""
    user = require_auth(request)
    _check_owner(project_id, user)
    body = await request.json()

    start_date = body.get("start_date", "")
    end_date = body.get("end_date", "")
    area = body.get("area_pyeong", 100)
    project_type = body.get("project_type", "오피스")
    selected_trades = body.get("selected_trades")
    has_import = body.get("has_import_materials", False)

    if not start_date or not end_date:
        raise HTTPException(status_code=400, detail="착공일과 준공일을 입력하세요.")

    try:
        import asyncio
        import tempfile
        from src.fund_table.schedule_generator import generate_construction_schedule
        from src.fund_table.schedule_export import export_schedule_xlsx, export_schedule_pdf

        # 1) 공정표 생성
        result = await asyncio.to_thread(
            generate_construction_schedule,
            start_date=start_date,
            end_date=end_date,
            area_pyeong=float(area),
            project_type=project_type,
            selected_trades=selected_trades,
            has_import_materials=has_import,
        )

        if result.get("summary", {}).get("error"):
            raise HTTPException(status_code=400, detail=result["summary"]["error"])

        # 2) DB 저장 (기존 일정 교체)
        items = result.get("schedule_items", [])
        db.save_schedule_items(project_id, items)

        # 3) 프로젝트명 조회
        project = db.get_project(project_id) if hasattr(db, 'get_project') else {}
        project_name = project.get("name", "") if project else ""

        # 4) 엑셀 파일 생성
        data_tmp = Path(__file__).parent.parent.parent / "data" / "tmp"
        data_tmp.mkdir(parents=True, exist_ok=True)
        safe_name = project_name.replace("/", "_").replace("\\", "_")[:30] if project_name else "schedule"
        xlsx_filename = f"공정표_{safe_name}_{start_date}.xlsx"
        xlsx_path = str(data_tmp / xlsx_filename)

        await asyncio.to_thread(export_schedule_xlsx, result, xlsx_path, project_name)

        # 5) PDF 변환 시도
        pdf_path = await asyncio.to_thread(export_schedule_pdf, xlsx_path)

        # 6) 다운로드 토큰 생성 (_download_registry 사용)
        from src.chatbot._download_registry import register as register_download
        gw_id = user.get("gw_id", "")
        xlsx_token = register_download(xlsx_path, gw_id)
        pdf_token = register_download(pdf_path, gw_id) if pdf_path else ""

        response = {
            "success": True,
            "schedule_items": items,
            "milestones": result.get("milestones", []),
            "summary": result.get("summary", {}),
            "xlsx_url": f"/download/{xlsx_token}" if xlsx_token else "",
            "pdf_url": f"/download/{pdf_token}" if pdf_token else "",
        }
        return JSONResponse(response)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"공정표 생성 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/fund/projects/{project_id}/export-schedule")
async def export_existing_schedule(project_id: int, request: Request):
    """기존 저장된 일정 데이터를 엑셀/PDF로 내보내기"""
    user = require_auth(request)
    _check_owner(project_id, user)
    body = await request.json()
    fmt = body.get("format", "xlsx")  # "xlsx" | "pdf" | "both"

    items = db.list_schedule_items(project_id)
    if not items:
        raise HTTPException(status_code=400, detail="저장된 일정 항목이 없습니다.")

    try:
        import asyncio
        import tempfile
        from src.fund_table.schedule_export import export_schedule_xlsx, export_schedule_pdf

        # 기존 데이터로 schedule_data 구성
        dates = []
        for it in items:
            for k in ("start_date", "end_date"):
                if it.get(k):
                    dates.append(it[k])
        sd = min(dates) if dates else ""
        ed = max(dates) if dates else ""
        bar_items = [it for it in items if it.get("item_type") != "milestone"]

        # 총 공사일수 계산
        total_cal_days = 0
        if sd and ed:
            from datetime import datetime as _dt
            try:
                total_cal_days = (_dt.strptime(ed, "%Y-%m-%d") - _dt.strptime(sd, "%Y-%m-%d")).days
            except ValueError:
                pass

        schedule_data = {
            "schedule_items": items,
            "milestones": [
                {"name": it["item_name"], "date": it["start_date"], "completed": it.get("status") == "done"}
                for it in items if it.get("item_type") == "milestone"
            ],
            "summary": {
                "start_date": sd, "end_date": ed,
                "total_calendar_days": total_cal_days,
                "total_trades": len(bar_items),
                "total_milestones": len([it for it in items if it.get("item_type") == "milestone"]),
                "area_pyeong": "", "project_type": "",
            },
        }

        project = db.get_project(project_id) if hasattr(db, 'get_project') else {}
        project_name = project.get("name", "") if project else ""

        data_tmp = Path(__file__).parent.parent.parent / "data" / "tmp"
        data_tmp.mkdir(parents=True, exist_ok=True)
        safe_name = project_name.replace("/", "_").replace("\\", "_")[:30] if project_name else "schedule"
        xlsx_filename = f"공정표_{safe_name}.xlsx"
        xlsx_path = str(data_tmp / xlsx_filename)

        await asyncio.to_thread(export_schedule_xlsx, schedule_data, xlsx_path, project_name)

        from src.chatbot._download_registry import register as register_download
        gw_id = user.get("gw_id", "")

        xlsx_token = register_download(xlsx_path, gw_id)
        pdf_token = ""
        pdf_path = ""

        if fmt in ("pdf", "both"):
            pdf_path = await asyncio.to_thread(export_schedule_pdf, xlsx_path)
            if pdf_path:
                pdf_token = register_download(pdf_path, gw_id)

        response = {
            "success": True,
            "xlsx_url": f"/download/{xlsx_token}",
        }
        if pdf_token:
            response["pdf_url"] = f"/download/{pdf_token}"
        return JSONResponse(response)

    except Exception as e:
        logger.error(f"공정표 내보내기 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
