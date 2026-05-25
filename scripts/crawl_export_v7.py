"""
아마란스 v7 전체 export 크롤러 — 사이드바 펼침 + 12개 모듈 모두 진입
─────────────────────────────────────────────────────────────────
v6의 결정적 오류 수정:
  - LNB는 콘텐츠 영역(x>60)에 없고, 사이드바(x<280) 안에 트리로 들어있다
  - "비활성" 6개는 단지 span.module-link 단일 셀렉터로 못 잡았던 것
  - 사이드바 펼친 상태에서 모듈명 텍스트로 클릭하면 진입 가능

v7 전략:
  1) 좌상단 햄버거(▦) 아이콘 클릭 → 사이드바 펼침
  2) 사이드바 안의 모듈명 텍스트(예: '게시판')로 클릭 → 진입
  3) 모듈 안의 sub-LNB는 사이드바 안의 트리 (x<280)
  4) leaf 클릭 → 콘텐츠 영역에 페이지 로드 → cel_save 다운로드
"""
from __future__ import annotations
import argparse, json, logging, re, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env", override=True)
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("v7")

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "data" / "amaranth_v7.json"
OUT_MD = ROOT / "docs" / "AMARANTH_EXPORT_v7.md"
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"
DL_DIR.mkdir(parents=True, exist_ok=True)

HOME_MODULES = [
    "시스템설정", "임직원업무관리", "전자결재", "메일", "일정", "자원",
    "게시판", "업무관리", "ONEFFICE", "ONECHAMBER", "프로세스관리", "오피스케어",
]

SKIP_LNB_TEXTS = {
    "해당 탭 닫기", "해당 탭 제외 다른 탭 닫기", "모든 탭 닫기",
    "탭 닫기", "닫기", "새로고침", "위로", "아래로",
}

# 사이드바 펼친 상태에서의 트리 (x < 280, h > 25)
JS_SIDEBAR_LNB = r"""
(skipTexts, currentModuleText, allModulesList) => {
    const out = [];
    const seen = new Set();
    document.querySelectorAll('li, [role="treeitem"], a, span, div').forEach(el => {
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return;
        // 사이드바 내부 (x: 60~280)
        if (r.x < 60 || r.x > 300) return;
        if (r.height < 20 || r.height > 50) return;
        // 직접 텍스트 (자식 텍스트가 아닌 본인 텍스트)
        const direct = Array.from(el.childNodes)
            .filter(n => n.nodeType === 3).map(n => n.textContent.trim())
            .filter(Boolean).join(' ');
        const full = (el.innerText || el.textContent || '').trim().split('\n')[0];
        const text = direct || full;
        if (!text || text.length > 50 || text.length < 2) return;
        if (skipTexts.includes(text)) return;
        // 12 모듈명 자체는 제외 (이미 진입한 모듈 아니면 진입용)
        // 그러나 진입한 모듈은 제외하지 않음 (LNB로 사용 가능)
        if (allModulesList.includes(text) && text !== currentModuleText) return;
        const key = text + '|' + Math.round(r.x);
        if (seen.has(key)) return;
        seen.add(key);
        out.push({
            text,
            x: Math.round(r.x), y: Math.round(r.y),
            cls: (el.className || '').toString().slice(0, 80),
        });
    });
    return out;
}
"""

JS_PAGE_AFTER_CLICK = r"""
() => {
    const out = { title: '', has_excel_btn: false, has_grid: false, excel_btn_positions: [] };
    const t = document.querySelector('[class*="PageTitle"], [class*="pageTitle"], h1, h2');
    out.title = t ? (t.innerText || '').trim().slice(0, 80) : '';
    document.querySelectorAll('img[src*="cel_save"]').forEach(img => {
        const r = img.getBoundingClientRect();
        if (r.width === 0) return;
        out.excel_btn_positions.push({
            x: Math.round(r.x + r.width/2), y: Math.round(r.y + r.height/2),
        });
    });
    out.has_excel_btn = out.excel_btn_positions.length > 0;
    out.has_grid = !!document.querySelector('[class*="OBTDataGrid"], [class*="RealGrid"]');
    return out;
}
"""


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "row_counts": {}, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
            out["row_counts"][name] = ws.max_row
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def expand_sidebar(page: Page) -> bool:
    """사이드바 펼침 — 좌상단 햄버거(≡) 아이콘 클릭"""
    # 좌상단 영역의 햄버거 후보 셀렉터
    candidates = [
        "button[aria-label*='메뉴']",
        "button[title*='메뉴']",
        "[class*='Hamburger']",
        "[class*='hamburger']",
        "[class*='ham']",
        # 좌상단 위치 직접 셀렉터
        "header button:nth-child(1)",
    ]
    for sel in candidates:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0 and loc.is_visible(timeout=500):
                loc.click(force=True, timeout=2000)
                page.wait_for_timeout(1000)
                logger.info(f"  사이드바 펼침 시도 ({sel})")
                return True
        except Exception:
            continue
    # 좌표 폴백: 햄버거는 보통 (80, 20)
    try:
        page.mouse.click(82, 22)
        page.wait_for_timeout(1000)
        logger.info("  사이드바 펼침 좌표 클릭 (82,22)")
        return True
    except Exception:
        return False


def click_sidebar_text(page: Page, text: str, timeout_ms: int = 3000) -> bool:
    """사이드바 안의 텍스트 클릭"""
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        for sel in [f"text='{text}'", f"li:has-text('{text}')", f"a:has-text('{text}')"]:
            try:
                items = fr.locator(sel)
                count = items.count()
                if count == 0:
                    continue
                # 사이드바 영역(x<300) 안에 있는 요소만
                for i in range(min(count, 5)):
                    item = items.nth(i)
                    try:
                        box = item.bounding_box()
                        if not box or box["x"] >= 300:
                            continue
                        if not item.is_visible(timeout=400):
                            continue
                        item.click(force=True, timeout=timeout_ms)
                        return True
                    except Exception:
                        continue
            except Exception:
                continue
    return False


def try_excel_download(page: Page, save_label: str) -> dict | None:
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            loc = fr.locator("button:has(img[src*='cel_save'])").first
            if loc.count() == 0:
                loc = fr.locator("img[src*='cel_save']").first
                if loc.count() == 0:
                    continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                with page.expect_download(timeout=15000) as dl_info:
                    loc.click(force=True, timeout=3000)
                    try:
                        for ok in ["확인", "다운로드", "OK"]:
                            okl = fr.locator(f"button:has-text('{ok}')").last
                            if okl.count() > 0 and okl.is_visible(timeout=400):
                                okl.click(timeout=1500)
                                break
                    except Exception:
                        pass
                dl = dl_info.value
                suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                path = DL_DIR / ("v7_" + safe_filename(save_label) + suffix)
                dl.save_as(str(path))
                size = path.stat().st_size if path.exists() else 0
                return {"saved": str(path), "size": size,
                        "suggested_name": dl.suggested_filename,
                        "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
            except PWTimeout:
                return {"saved": None, "error": "timeout"}
            except Exception as e:
                return {"saved": None, "error": str(e)}
        except Exception:
            continue
    return None


def collect_all_frames(page: Page, js: str, *args) -> list:
    out = []
    for fr in [page.main_frame] + [f for f in page.frames if f != page.main_frame]:
        try:
            r = fr.evaluate(js, *args) if args else fr.evaluate(js)
            if isinstance(r, list):
                out.extend(r)
            elif isinstance(r, dict):
                out.append(r)
        except Exception:
            continue
    return out


def current_url(page: Page) -> str:
    cands = [fr.url for fr in page.frames if fr.url and "#/" in fr.url]
    return max(cands, key=len) if cands else page.url


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--modules", type=str, default=None)
    ap.add_argument("--leaf_limit", type=int, default=50)
    ap.add_argument("--download_limit", type=int, default=80)
    args = ap.parse_args()

    only = set(args.modules.split(",")) if args.modules else None
    result = {
        "gw_user": "tgjeon",
        "crawled_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "modules": [],
        "errors": [],
    }
    downloads_done = 0

    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw, headless=args.headless, user_id="tgjeon",
        )
        try:
            page.set_viewport_size({"width": 1920, "height": 1080})
        except Exception:
            pass
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        # 시작: 홈 → 사이드바 펼침
        page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(4000)
        expand_sidebar(page)
        page.wait_for_timeout(2000)
        try:
            page.screenshot(path=str(SCR_DIR / "v7_sidebar_expanded.png"))
        except Exception:
            pass

        for mod_text in HOME_MODULES:
            if only and mod_text not in only:
                continue
            logger.info(f"\n{'='*60}\n[{mod_text}]\n{'='*60}")

            # 사이드바 펼친 상태 유지하면서 모듈 텍스트 클릭
            ok = click_sidebar_text(page, mod_text, timeout_ms=3000)
            if not ok:
                # 사이드바 펼치고 재시도
                expand_sidebar(page)
                page.wait_for_timeout(1500)
                ok = click_sidebar_text(page, mod_text, timeout_ms=3000)
            if not ok:
                logger.warning(f"  {mod_text} 클릭 실패")
                result["errors"].append({"phase": "module_click", "text": mod_text})
                continue
            page.wait_for_timeout(5000)
            try:
                page.screenshot(path=str(SCR_DIR / f"v7_mod_{safe_filename(mod_text)}.png"))
            except Exception:
                pass

            entry_url = current_url(page)
            mod_entry = {
                "text": mod_text, "entry_url": entry_url,
                "lnb": [], "leaf_visits": [], "main_download": None,
            }
            logger.info(f"  진입 URL: {entry_url[-80:]}")

            # 모듈 메인 페이지에서 엑셀 다운로드 시도
            try:
                info_list = collect_all_frames(page, JS_PAGE_AFTER_CLICK)
                has_excel = any(i.get("has_excel_btn") for i in info_list)
                if has_excel and downloads_done < args.download_limit:
                    logger.info(f"  ★ 모듈 메인에 엑셀 발견 → 다운로드 시도")
                    dl = try_excel_download(page, f"{mod_text}_MAIN")
                    if dl and dl.get("saved"):
                        downloads_done += 1
                        logger.info(f"  ✓ 메인 다운: {Path(dl['saved']).name} ({dl['size']}B)")
                        mod_entry["main_download"] = dl
                    else:
                        logger.warning(f"  ✗ 메인 다운 실패: {dl}")
            except Exception as e:
                logger.warning(f"  메인 다운로드 시도 실패: {e}")

            # 사이드바 안의 LNB 항목 수집 (이 모듈의 카테고리/leaf)
            try:
                lnb_items = collect_all_frames(page, JS_SIDEBAR_LNB,
                                               list(SKIP_LNB_TEXTS), mod_text, HOME_MODULES)
                # 중복 제거
                dedup = {}
                for it in lnb_items:
                    t = it.get("text")
                    if not t or t in dedup:
                        continue
                    dedup[t] = it
                lnb_unique = list(dedup.values())
                mod_entry["lnb"] = lnb_unique
                logger.info(f"  사이드바 LNB: {len(lnb_unique)}개")
            except Exception as e:
                logger.warning(f"  LNB 수집 실패: {e}")
                lnb_unique = []

            # LNB 항목 모두 클릭 (펼침 + leaf 진입)
            visited = set()
            for idx, leaf in enumerate(lnb_unique[: args.leaf_limit]):
                lt = leaf["text"]
                if lt in visited:
                    continue
                visited.add(lt)
                logger.info(f"  [{idx+1}/{min(len(lnb_unique), args.leaf_limit)}] {lt}")
                ok = click_sidebar_text(page, lt, timeout_ms=2500)
                if not ok:
                    continue
                page.wait_for_timeout(2500)
                # 펼친 후 새 항목이 있을 수 있음 — 다시 수집해서 추가
                try:
                    new_items = collect_all_frames(page, JS_SIDEBAR_LNB,
                                                   list(SKIP_LNB_TEXTS), mod_text, HOME_MODULES)
                    for ni in new_items:
                        if ni["text"] not in dedup:
                            dedup[ni["text"]] = ni
                            lnb_unique.append(ni)
                except Exception:
                    pass

                # 페이지 정보 수집
                try:
                    info_list = collect_all_frames(page, JS_PAGE_AFTER_CLICK)
                except Exception:
                    continue
                title = ""
                has_excel = False
                has_grid = False
                for inf in info_list:
                    if not title and inf.get("title"):
                        title = inf["title"]
                    has_excel = has_excel or inf.get("has_excel_btn")
                    has_grid = has_grid or inf.get("has_grid")
                url = current_url(page)
                visit = {
                    "lnb_text": lt, "url": url, "page_title": title,
                    "has_excel_btn": has_excel, "has_grid": has_grid, "download": None,
                }
                if has_excel and downloads_done < args.download_limit:
                    logger.info(f"    ★ 엑셀 발견 → 다운로드")
                    save_label = f"{mod_text}_{lt}"
                    dl = try_excel_download(page, save_label)
                    visit["download"] = dl
                    if dl and dl.get("saved"):
                        downloads_done += 1
                        logger.info(f"    ✓ 저장: {Path(dl['saved']).name} ({dl['size']}B)")
                mod_entry["leaf_visits"].append(visit)

            result["modules"].append(mod_entry)
            logger.info(f"  → 모듈 완료: leaf 방문 {len(mod_entry['leaf_visits'])}, 다운 누적 {downloads_done}")

        context.close()
        browser.close()

    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON 저장: {OUT_JSON}")

    # Markdown 보고서
    successes = []
    for me in result["modules"]:
        if me.get("main_download") and me["main_download"].get("saved"):
            successes.append((me["text"], "(모듈 메인)", me["main_download"]))
        for v in me["leaf_visits"]:
            if v.get("download") and v["download"].get("saved"):
                successes.append((me["text"], v["lnb_text"], v["download"]))

    lines = ["# 아마란스 v7 전체 Export 결과", ""]
    lines.append(f"- 계정: tgjeon / 시각: {result['crawled_at']}")
    lines.append(f"- 모듈: {len(result['modules'])} / 다운로드 성공: {len(successes)}")
    lines.append("")
    lines.append(f"## ✅ 다운로드 성공 ({len(successes)}개)\n")
    lines.append("| 모듈 | 메뉴 | 파일 | 크기 | 시트(행) |")
    lines.append("|---|---|---|---|---|")
    for mod_text, leaf, d in successes:
        s = d.get("summary") or {}
        sheet_info = ", ".join(f"{name}({s.get('row_counts',{}).get(name,'?')})" for name in s.get("sheets", []))
        lines.append(f"| {mod_text} | {leaf} | `{Path(d['saved']).name}` | {d['size']}B | {sheet_info} |")
    lines.append("")

    # 모듈별 진입
    for me in result["modules"]:
        lines.append(f"\n### [{me['text']}]")
        lines.append(f"- 진입 URL: `{(me.get('entry_url') or '')[-80:]}`")
        lines.append(f"- 사이드바 LNB: {len(me['lnb'])} / leaf 방문: {len(me['leaf_visits'])}")
        if me["leaf_visits"]:
            lines.append("\n| Leaf | URL | 그리드 | 엑셀 | 다운 |")
            lines.append("|---|---|---|---|---|")
            for v in me["leaf_visits"]:
                g = "✓" if v["has_grid"] else ""
                e = "✓" if v["has_excel_btn"] else ""
                d = v.get("download") or {}
                ds = "✅" if d.get("saved") else ("⚠️" if d else "")
                lines.append(f"| {v['lnb_text']} | `{(v['url'] or '')[-50:]}` | {g} | {e} | {ds} |")

    # 미리보기
    if successes:
        lines.append("\n## 다운로드 파일 미리보기\n")
        for mod_text, leaf, d in successes:
            s = d.get("summary")
            if not s or "error" in s:
                continue
            lines.append(f"### {mod_text} > {leaf}")
            for sheet, rows in s.get("preview", {}).items():
                lines.append(f"\n**시트** `{sheet}` ({s['row_counts'].get(sheet, '?')}행)")
                if rows:
                    headers = rows[0]
                    n = min(len(headers), 10)
                    lines.append("| " + " | ".join(str(h)[:25] for h in headers[:n]) + " |")
                    lines.append("|" + "|".join(["---"] * n) + "|")
                    for row in rows[1:6]:
                        padded = list(row) + [""] * (n - len(row)) if len(row) < n else row
                        lines.append("| " + " | ".join(str(c)[:25] for c in padded[:n]) + " |")
            lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"MD 저장: {OUT_MD}")
    logger.info(f"=== 모듈 {len(result['modules'])} / 다운로드 {len(successes)} ===")


if __name__ == "__main__":
    main()
