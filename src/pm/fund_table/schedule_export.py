"""공정표 엑셀(간트차트+리스트) 및 PDF 출력 모듈"""
import logging
import os
import subprocess
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 스타일 상수
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="2D3748")
_HEADER_FONT = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
_GROUP_FILL = PatternFill("solid", fgColor="374151")
_GROUP_FONT = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="맑은 고딕", size=8)
_TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
_MS_FONT = Font(name="맑은 고딕", size=8, color="EF4444")
_CP_FONT = Font(name="맑은 고딕", size=8, bold=True, color="DC2626")  # 임계경로 공종명 (빨간 굵게)
_CP_SIDE = Side(style="medium", color="DC2626")  # 임계경로 간트바 테두리
_CP_BORDER = Border(left=_CP_SIDE, right=_CP_SIDE, top=_CP_SIDE, bottom=_CP_SIDE)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _hex_to_argb(hex_color: str) -> str:
    """#RRGGBB → AARRGGBB"""
    h = hex_color.lstrip("#")
    return f"FF{h.upper()}"


def _week_mondays(start: datetime, end: datetime) -> list[datetime]:
    """시작~종료 사이 모든 월요일 리스트"""
    d = start - timedelta(days=start.weekday())
    result = []
    while d <= end:
        result.append(d)
        d += timedelta(days=7)
    return result


def export_schedule_xlsx(schedule_data: dict, output_path: str, project_name: str = "") -> str:
    """공정표 엑셀 생성 (시트1: 간트차트, 시트2: 리스트)

    Args:
        schedule_data: schedule_generator.generate_construction_schedule() 결과
        output_path: 저장 경로 (.xlsx)
        project_name: 프로젝트명 (제목에 표시)

    Returns:
        저장된 파일 경로
    """
    items = schedule_data.get("schedule_items", [])
    summary = schedule_data.get("summary", {})
    milestones = schedule_data.get("milestones", [])

    if not items:
        raise ValueError("일정 항목이 없습니다.")

    wb = Workbook()

    # ── 시트 1: 간트차트 ──
    ws_gantt = wb.active
    ws_gantt.title = "공정표 (간트차트)"
    _build_gantt_sheet(ws_gantt, items, summary, project_name)

    # ── 시트 2: 리스트 ──
    ws_list = wb.create_sheet("공정표 (리스트)")
    _build_list_sheet(ws_list, items, milestones, summary, project_name)

    wb.save(output_path)
    logger.info(f"공정표 엑셀 생성 완료: {output_path}")
    return output_path


def _build_gantt_sheet(ws, items: list, summary: dict, project_name: str):
    """간트차트 시트 구성"""
    # 날짜 범위 계산
    dates = []
    for it in items:
        for key in ("start_date", "end_date"):
            if it.get(key):
                dates.append(datetime.strptime(it[key], "%Y-%m-%d"))
    if not dates:
        return
    min_date = min(dates)
    max_date = max(dates)
    weeks = _week_mondays(min_date, max_date)
    if not weeks:
        return

    # 좌측 고정 열: 그룹명, 공종명, 소요일수
    left_cols = 3
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8

    # 제목 행
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=left_cols + len(weeks))
    title_cell = ws.cell(row=1, column=1)
    title_text = f"시공 공정표 — {project_name}" if project_name else "시공 공정표"
    title_cell.value = title_text
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 요약 행
    ws.cell(row=2, column=1, value=f"공사기간: {summary.get('start_date', '')} ~ {summary.get('end_date', '')}")
    ws.cell(row=2, column=1).font = Font(name="맑은 고딕", size=9, color="6B7280")
    area_txt = f"면적: {summary.get('area_pyeong', '')}평 | 유형: {summary.get('project_type', '')}"
    ws.cell(row=2, column=left_cols + 1, value=area_txt)
    ws.cell(row=2, column=left_cols + 1).font = Font(name="맑은 고딕", size=9, color="6B7280")

    # 헤더 행 (행 4): 그룹 / 공종명 / 일수 / 주차별 날짜
    header_row = 4
    for col, label in enumerate(["그룹", "공종명", "일수"], 1):
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    for wi, monday in enumerate(weeks):
        col = left_cols + 1 + wi
        ws.column_dimensions[get_column_letter(col)].width = 4
        c = ws.cell(row=header_row, column=col, value=monday.strftime("%m/%d"))
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    # 데이터 행
    row = header_row + 1
    current_group = None

    for it in items:
        grp = it.get("group_name", "")

        # 그룹 헤더 행
        if grp != current_group:
            current_group = grp
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=left_cols + len(weeks))
            gc = ws.cell(row=row, column=1, value=grp)
            gc.font = _GROUP_FONT
            gc.fill = _GROUP_FILL
            gc.alignment = _LEFT
            row += 1

        # 공종 행 — 임계경로 공종은 빨간 굵은 글씨
        is_cp = it.get("is_critical", False)
        name_font = _CP_FONT if is_cp else _BODY_FONT
        ws.cell(row=row, column=1, value=grp).font = Font(name="맑은 고딕", size=7, color="9CA3AF")
        ws.cell(row=row, column=1).alignment = _CENTER
        item_name = it.get("item_name", "")
        if is_cp:
            item_name = f"★ {item_name}"
        ws.cell(row=row, column=2, value=item_name).font = name_font
        ws.cell(row=row, column=2).alignment = _LEFT

        # 소요일수
        if it.get("start_date") and it.get("end_date"):
            sd = datetime.strptime(it["start_date"], "%Y-%m-%d")
            ed = datetime.strptime(it["end_date"], "%Y-%m-%d")
            days = (ed - sd).days
        else:
            sd = ed = None
            days = 0

        ws.cell(row=row, column=3, value=days if days > 0 else "").font = name_font
        ws.cell(row=row, column=3).alignment = _CENTER

        # 간트 바 채우기
        if sd and ed:
            bar_color = it.get("bar_color", "#3b82f6").lstrip("#").upper()
            fill = PatternFill("solid", fgColor=bar_color)
            is_milestone = it.get("item_type") == "milestone"
            bar_border = _CP_BORDER if is_cp else _BORDER

            for wi, monday in enumerate(weeks):
                col = left_cols + 1 + wi
                week_end = monday + timedelta(days=6)
                cell = ws.cell(row=row, column=col)
                cell.border = _BORDER

                if is_milestone and monday <= sd <= week_end:
                    cell.value = "\u25C6"
                    cell.font = _MS_FONT
                    cell.alignment = _CENTER
                elif not is_milestone and sd <= week_end and ed >= monday:
                    cell.fill = fill
                    cell.border = bar_border

        for ci in range(1, left_cols + 1):
            ws.cell(row=row, column=ci).border = _BORDER

        row += 1

    # 오늘 마커 (주차 열에 빨간 테두리)
    today = datetime.now()
    for wi, monday in enumerate(weeks):
        week_end = monday + timedelta(days=6)
        if monday <= today <= week_end:
            red_side = Side(style="medium", color="EF4444")
            col = left_cols + 1 + wi
            for r in range(header_row, row):
                cell = ws.cell(row=r, column=col)
                cell.border = Border(left=red_side, right=red_side, top=cell.border.top, bottom=cell.border.bottom)

    # 인쇄 설정: A3 가로
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.freeze_panes = f"D{header_row + 1}"


def _build_list_sheet(ws, items: list, milestones: list, summary: dict, project_name: str):
    """리스트 시트 구성"""
    # 제목
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1)
    title_text = f"시공 공정표 (리스트) — {project_name}" if project_name else "시공 공정표 (리스트)"
    title_cell.value = title_text
    title_cell.font = _TITLE_FONT

    # 요약
    ws.cell(row=2, column=1, value=f"공사기간: {summary.get('start_date', '')} ~ {summary.get('end_date', '')} | "
                                   f"면적: {summary.get('area_pyeong', '')}평 | "
                                   f"유형: {summary.get('project_type', '')} | "
                                   f"공종: {summary.get('total_trades', 0)}개")
    ws.cell(row=2, column=1).font = Font(name="맑은 고딕", size=9, color="6B7280")

    # 요약에 CPM 정보 추가
    cp_count = summary.get("critical_path_count", 0)
    if cp_count:
        ws.cell(row=3, column=1, value=f"임계경로(CP): {cp_count}개 공종 | "
                                       f"면적 보정: {summary.get('area_factor', 1.0)}x | "
                                       f"스케일: {summary.get('scale_factor', 1.0)}x")
        ws.cell(row=3, column=1).font = Font(name="맑은 고딕", size=9, color="DC2626")

    # 헤더
    headers = ["#", "그룹", "공종명", "시작일", "종료일", "소요일수", "CP", "여유", "상태", "세부 단계"]
    widths = [5, 12, 24, 12, 12, 8, 4, 6, 8, 36]
    header_row = 4

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = w
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    # 데이터
    row = header_row + 1
    current_group = None
    num = 0

    for it in items:
        grp = it.get("group_name", "")

        # 그룹 구분 행
        if grp != current_group:
            current_group = grp
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            gc = ws.cell(row=row, column=1, value=grp)
            gc.font = _GROUP_FONT
            gc.fill = _GROUP_FILL
            gc.alignment = _LEFT
            row += 1

        num += 1
        sd = it.get("start_date", "")
        ed = it.get("end_date", "")
        days = 0
        if sd and ed:
            try:
                days = (datetime.strptime(ed, "%Y-%m-%d") - datetime.strptime(sd, "%Y-%m-%d")).days
            except ValueError:
                pass

        status_map = {"planned": "예정", "ongoing": "진행중", "done": "완료", "hold": "보류"}
        status = status_map.get(it.get("status", "planned"), it.get("status", ""))
        item_type = it.get("item_type", "bar")
        is_cp = it.get("is_critical", False)
        display_name = it.get("item_name", "")
        if item_type == "milestone":
            display_name = f"\u25C6 {display_name}"
        elif is_cp:
            display_name = f"★ {display_name}"

        cp_mark = "★" if is_cp else ""
        float_val = it.get("total_float", 0)
        float_display = f"{float_val}일" if float_val > 0 else ("-" if is_cp else "0")

        values = [num, grp, display_name, sd, ed, days if days > 0 else "-",
                  cp_mark, float_display, status, it.get("notes", "")]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = _CP_FONT if is_cp and ci in (3, 7) else _BODY_FONT
            c.alignment = _CENTER if ci in (1, 4, 5, 6, 7, 8, 9) else _LEFT
            c.border = _BORDER

            # 마일스톤 빨간 글자
            if item_type == "milestone" and ci == 3:
                c.font = _MS_FONT

        row += 1

    # 하단 요약
    row += 1
    ws.cell(row=row, column=1, value="요약").font = Font(name="맑은 고딕", size=10, bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"총 공사일수: {summary.get('total_calendar_days', 0)}일").font = _BODY_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"공종 수: {summary.get('total_trades', 0)}개").font = _BODY_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"마일스톤: {summary.get('total_milestones', 0)}개").font = _BODY_FONT

    # 마일스톤 목록
    if milestones:
        row += 2
        ws.cell(row=row, column=1, value="마일스톤 체크리스트").font = Font(name="맑은 고딕", size=10, bold=True)
        row += 1
        for mi, ms in enumerate(milestones, 1):
            ws.cell(row=row, column=1, value=mi).font = _BODY_FONT
            ws.cell(row=row, column=2, value=ms.get("name", "")).font = _BODY_FONT
            ws.cell(row=row, column=3, value=ms.get("date", "")).font = _BODY_FONT
            row += 1

    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1


def export_schedule_pdf(xlsx_path: str, output_dir: str | None = None) -> str:
    """엑셀 → PDF 변환 (LibreOffice headless)

    Args:
        xlsx_path: 입력 엑셀 파일 경로
        output_dir: PDF 저장 디렉토리 (None이면 같은 디렉토리)

    Returns:
        생성된 PDF 파일 경로
    """
    if output_dir is None:
        output_dir = os.path.dirname(xlsx_path) or "."

    soffice_paths = [
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for soffice in soffice_paths:
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_path],
                capture_output=True, text=True, timeout=60,
            )
            if result.returncode == 0:
                break
            logger.warning(f"LibreOffice ({soffice}) 변환 실패: {result.stderr}")
        except FileNotFoundError:
            continue
        except subprocess.TimeoutExpired:
            logger.error("LibreOffice 변환 시간 초과 (60초)")
            return ""
    else:
        logger.error("LibreOffice가 설치되어 있지 않습니다. PDF 변환을 건너뜁니다.")
        return ""

    pdf_name = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
    pdf_path = os.path.join(output_dir, pdf_name)
    if os.path.exists(pdf_path):
        logger.info(f"PDF 변환 완료: {pdf_path}")
        return pdf_path
    logger.error(f"PDF 파일 미생성: {pdf_path}")
    return ""
