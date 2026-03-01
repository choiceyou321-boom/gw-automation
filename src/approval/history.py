"""
Task #2: 결재 이력 조회/정리
- 과거 상신한 전자결재 전체 목록 조회
- 양식별 분류 및 상세 정보 수집
- Excel 파일로 정리/출력
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from playwright.sync_api import Page, BrowserContext

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from src.auth.login import login_and_get_context, close_session, GW_URL, DATA_DIR

logger = logging.getLogger("approval")


def fetch_approval_history(page: Page, max_pages: int = 50) -> list[dict]:
    """
    전자결재 상신 이력 전체를 수집.
    Playwright로 결재함 페이지를 탐색하고, 네트워크 인터셉트로 API 데이터도 캡처.
    """
    approvals = []
    api_responses = []

    # 네트워크 인터셉트 - API 응답 캡처
    def handle_response(response):
        url = response.url.lower()
        # 결재 관련 API 응답 캡처
        if any(kw in url for kw in ["approval", "eap", "draft", "sanction", "document"]):
            try:
                body = response.json()
                api_responses.append({"url": response.url, "data": body})
                logger.info(f"API 캡처: {response.url[:100]}")
            except Exception:
                pass

    page.on("response", handle_response)

    # 전자결재 메뉴로 이동
    logger.info("전자결재 메뉴 진입 중...")
    _navigate_to_approval(page)

    # 상신함(기안함) 이동
    logger.info("상신함으로 이동 중...")
    _navigate_to_sent_box(page)

    # 목록 수집
    page_num = 1
    while page_num <= max_pages:
        logger.info(f"페이지 {page_num} 수집 중...")
        items = _extract_list_items(page)

        if not items:
            logger.info("더 이상 항목이 없습니다.")
            break

        approvals.extend(items)
        logger.info(f"  → {len(items)}건 수집 (누적: {len(approvals)}건)")

        # 다음 페이지
        if not _go_next_page(page):
            break
        page_num += 1

    # 각 결재건 상세 정보 수집 (선택적)
    logger.info(f"총 {len(approvals)}건 수집 완료")

    # API 캡처 데이터도 저장
    if api_responses:
        api_file = DATA_DIR / "captured_apis.json"
        api_file.write_text(json.dumps(api_responses, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(f"캡처된 API {len(api_responses)}개 저장: {api_file}")

    return approvals


def _navigate_to_approval(page: Page):
    """전자결재 메뉴로 이동"""
    # 더존 그룹웨어의 전자결재 메뉴 셀렉터들
    selectors = [
        'a:has-text("전자결재")',
        'span:has-text("전자결재")',
        '[data-menu*="approval"]',
        '[href*="approval"]',
        '[href*="eap"]',
        '.menu-item:has-text("전자결재")',
        'li:has-text("전자결재")',
    ]

    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                page.wait_for_timeout(3000)
                logger.info(f"전자결재 메뉴 클릭: {sel}")
                return
        except Exception:
            continue

    # URL 직접 이동 시도
    approval_urls = [
        f"{GW_URL}/#/approval",
        f"{GW_URL}/#/eap",
        f"{GW_URL}/#/app/approval",
    ]
    for url in approval_urls:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(3000)
            if "approval" in page.url.lower() or "eap" in page.url.lower():
                logger.info(f"URL로 전자결재 진입: {url}")
                return
        except Exception:
            continue

    # 스크린샷 저장
    page.screenshot(path=str(DATA_DIR / "approval_nav_failed.png"))
    logger.warning("전자결재 메뉴 자동 진입 실패 - 스크린샷 확인 필요")


def _navigate_to_sent_box(page: Page):
    """상신함(기안함/발신함) 이동"""
    selectors = [
        'a:has-text("기안함")',
        'a:has-text("상신함")',
        'a:has-text("발신함")',
        'span:has-text("기안함")',
        'span:has-text("상신함")',
        '[class*="draft"]',
        '[class*="sent"]',
        'li:has-text("기안함")',
        'li:has-text("상신함")',
    ]

    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                el.click()
                page.wait_for_timeout(3000)
                logger.info(f"상신함 이동: {sel}")
                return
        except Exception:
            continue

    logger.warning("상신함 메뉴를 찾지 못했습니다 - 현재 페이지에서 목록 추출 시도")


def _extract_list_items(page: Page) -> list[dict]:
    """현재 페이지의 결재 목록 추출"""
    items = []

    # 테이블 기반 목록 추출
    try:
        rows = page.locator("table tbody tr, .list-item, .approval-item, .document-item").all()
        for row in rows:
            try:
                text = row.inner_text(timeout=2000)
                cells = [c.strip() for c in text.split("\t") if c.strip()]
                if not cells:
                    cells = [c.strip() for c in text.split("\n") if c.strip()]

                if len(cells) >= 2:
                    item = _parse_row_cells(cells)
                    if item:
                        items.append(item)
            except Exception:
                continue
    except Exception as e:
        logger.warning(f"목록 추출 실패: {e}")

    # 목록이 비어있으면 전체 텍스트에서 패턴 추출 시도
    if not items:
        try:
            content = page.content()
            # 스크린샷으로 현재 상태 기록
            page.screenshot(path=str(DATA_DIR / "approval_list_page.png"))
            logger.info("목록 페이지 스크린샷 저장: approval_list_page.png")
        except Exception:
            pass

    return items


def _parse_row_cells(cells: list[str]) -> dict | None:
    """테이블 행 데이터를 딕셔너리로 변환"""
    if len(cells) < 2:
        return None

    item = {"raw_cells": cells}

    # 일반적인 결재 목록 컬럼: 번호, 제목, 양식, 기안일, 상태, 결재자 등
    for i, cell in enumerate(cells):
        # 날짜 패턴 감지
        if any(c in cell for c in ["-", ".", "/"]) and any(c.isdigit() for c in cell):
            if len(cell) >= 8:
                item.setdefault("date", cell)
        # 상태 감지
        elif cell in ["결재완료", "진행중", "반려", "완료", "승인", "대기", "상신"]:
            item["status"] = cell
        # 긴 텍스트는 제목으로
        elif len(cell) > 10 and "title" not in item:
            item["title"] = cell

    # 최소한 제목은 있어야 함
    if "title" not in item:
        item["title"] = cells[1] if len(cells) > 1 else cells[0]

    return item


def _go_next_page(page: Page) -> bool:
    """다음 페이지로 이동, 성공 여부 반환"""
    next_selectors = [
        'button:has-text("다음")',
        'a:has-text("다음")',
        '.pagination .next',
        'button.next',
        '[aria-label="Next"]',
        '.paging-next',
        'a:has-text(">")',
    ]

    for sel in next_selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1000) and el.is_enabled(timeout=1000):
                el.click()
                page.wait_for_timeout(2000)
                return True
        except Exception:
            continue

    return False


def save_to_excel(approvals: list[dict], filename: str = None):
    """결재 이력을 Excel 파일로 저장"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"결재이력_{timestamp}.xlsx"

    filepath = DATA_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "결재 이력"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더 작성
    headers = ["번호", "제목", "양식", "기안일", "상태", "기타"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 데이터 작성
    for i, item in enumerate(approvals, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=item.get("title", "")).border = thin_border
        ws.cell(row=row, column=3, value=item.get("form_type", "")).border = thin_border
        ws.cell(row=row, column=4, value=item.get("date", "")).border = thin_border
        ws.cell(row=row, column=5, value=item.get("status", "")).border = thin_border
        ws.cell(row=row, column=6, value=str(item.get("raw_cells", ""))).border = thin_border

    # 열 너비 조정
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40

    wb.save(str(filepath))
    logger.info(f"Excel 저장 완료: {filepath}")
    return filepath


def run():
    """결재 이력 조회 메인 실행"""
    logger.info("=" * 50)
    logger.info("Task #2: 결재 이력 조회/정리 시작")
    logger.info("=" * 50)

    browser, context, page = login_and_get_context(headless=False)

    try:
        # 결재 이력 수집
        approvals = fetch_approval_history(page)

        if approvals:
            # Excel로 저장
            filepath = save_to_excel(approvals)
            logger.info(f"결과: {len(approvals)}건 → {filepath}")

            # JSON 백업도 저장
            json_path = DATA_DIR / "approval_history.json"
            json_path.write_text(
                json.dumps(approvals, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            logger.info(f"JSON 백업: {json_path}")
        else:
            logger.warning("수집된 결재 이력이 없습니다. 스크린샷을 확인해주세요.")
            page.screenshot(path=str(DATA_DIR / "approval_empty.png"))

    finally:
        close_session(browser)

    logger.info("Task #2 완료")


if __name__ == "__main__":
    run()
