"""
HR 모듈 sub-LNB 깊이 탐색 + Export 실제 다운로드 (v4)
─────────────────────────────────────────────────
v3 발견: HR(임직원업무) LNB가 진짜 통합 메뉴이고 13개 1-depth 카테고리 보유
v4 전략:
  1) 홈 → HR(임직원업무관리) 클릭 (span.module-link.HR)
  2) HR의 1-depth 카테고리(근태관리/인사관리/지출결의/예산관리/법정의무교육 ...) 각각 클릭
  3) 2-depth LNB 수집 (우클릭 컨텍스트 메뉴 "해당 탭 닫기" 류 블랙리스트)
  4) 2-depth leaf 클릭 → 페이지 진입 → URL + 버튼 + export 탐지
  5) export 후보 발견 시 실제 클릭 → 다운로드 시도 → openpyxl 로 헤더+5행 미리보기

결과:
  data/amaranth_hr_deep.json
  docs/AMARANTH_HR_DEEP.md
  data/amaranth_exports/*.xlsx
"""
from __future__ import annotations
import argparse, json, logging, re, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from playwright.sync_api import sync_playwright, Page, Download, TimeoutError as PWTimeout
from src.shared.auth.login import login_and_get_context

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("hr_deep")

ROOT = Path(__file__).resolve().parent.parent
OUT_JSON = ROOT / "data" / "amaranth_hr_deep.json"
OUT_MD = ROOT / "docs" / "AMARANTH_HR_DEEP.md"
DL_DIR = ROOT / "data" / "amaranth_exports"
SCR_DIR = ROOT / "data" / "amaranth_screens"

GW_USER = "tgjeon"

# 우클릭 컨텍스트 메뉴 등 LNB가 아닌 항목 블랙리스트
LNB_BLACKLIST = {
    "해당 탭 닫기", "해당 탭 제외 다른 탭 닫기", "모든 탭 닫기",
    "탭 닫기", "닫기", "새로고침", "프로세스갤러리",
}

# HR 1-depth 카테고리 (v3에서 발견된 텍스트 그대로 사용)
HR_TOP_CATEGORIES = [
    "마이페이지", "근태관리", "인사관리", "급여관리", "경비청구",
    "지출결의/계산서", "개인지출결의서", "예산관리", "법정의무교육",
]

EXPORT_KEYWORDS = [
    "엑셀 다운로드", "엑셀다운로드", "엑셀", "엑셀저장", "엑셀변환",
    "CSV", "PDF", "다운로드", "내보내기", "Excel", "Export", "Download",
]

JS_VISIBLE_LIST_ITEMS = r"""
(blacklist) => {
    const out = [];
    const seen = new Set();
    const sels = [
        '[class*="OBTLeftMenu"] li',
        '[class*="LeftMenu"] li',
        '[class*="LNB"] li',
        '[class*="lnb"] li',
        '[class*="TreeMenu"] li',
        '[class*="Tree"] li',
        '[role="treeitem"]',
        '[role="menuitem"]',
        'aside li',
        'nav li',
        'ul.menu li',
    ];
    sels.forEach(sel => {
        document.querySelectorAll(sel).forEach(el => {
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) return;
            const txt = (el.innerText || el.textContent || '').trim().split('\n')[0];
            if (!txt || txt.length > 60) return;
            if (blacklist.includes(txt)) return;
            if (seen.has(txt)) return;
            seen.add(txt);
            out.push({
                text: txt,
                cls: (el.className || '').toString().slice(0, 100),
                x: Math.round(r.x), y: Math.round(r.y),
                w: Math.round(r.width), h: Math.round(r.height),
            });
        });
    });
    return out;
}
"""

JS_BUTTONS = r"""
(keywords) => {
    const out = { title: '', buttons: [], export_candidates: [] };
    const t = document.querySelector('[class*="PageTitle"], [class*="pageTitle"], h1, h2');
    out.title = t ? (t.innerText || '').trim().slice(0, 80) : '';
    const seen = new Set();
    document.querySelectorAll('button, a[role="button"], [class*="Button"], [role="button"]').forEach(el => {
        const txt = (el.innerText || el.textContent || el.title || el.getAttribute('aria-label') || '').trim();
        if (!txt || txt.length > 30) return;
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return;
        if (seen.has(txt)) return;
        seen.add(txt);
        const btn = { text: txt, title: el.title || '', aria: el.getAttribute('aria-label') || '' };
        out.buttons.push(btn);
        for (const kw of keywords) {
            if (txt.includes(kw)) {
                out.export_candidates.push({ ...btn, matched: kw });
                break;
            }
        }
    });
    return out;
}
"""


def collect_all_frames(page: Page, js: str, *args) -> list:
    out = []
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        try:
            r = fr.evaluate(js, *args) if args else fr.evaluate(js)
            if isinstance(r, list):
                for it in r:
                    if isinstance(it, dict):
                        it["_frame"] = fr.url[:80]
                out.extend(r)
            elif isinstance(r, dict):
                r["_frame"] = fr.url[:80]
                out.append(r)
        except Exception:
            continue
    return out


def current_content_url(page: Page) -> str:
    """가장 의미 있는 iframe URL 선택 (가장 깊은 #/ 경로)"""
    candidates = []
    for fr in page.frames:
        if fr.url and "#/" in fr.url:
            candidates.append(fr.url)
    if not candidates:
        return page.url
    return max(candidates, key=len)


def click_text_in_any_frame(page: Page, text: str, timeout_ms: int = 2500) -> bool:
    """page + iframe 모두에서 텍스트 매칭 항목 클릭 시도"""
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        try:
            # 정확 일치
            loc = fr.locator(f"text='{text}'").first
            if loc.count() > 0 and loc.is_visible():
                loc.click(force=True, timeout=timeout_ms)
                return True
        except Exception:
            pass
        try:
            # has-text 폴백
            loc = fr.locator(f"li:has-text('{text}')").first
            if loc.count() > 0 and loc.is_visible():
                loc.click(force=True, timeout=timeout_ms)
                return True
        except Exception:
            pass
    return False


def safe_filename(s: str) -> str:
    s = re.sub(r"[^\w가-힣\-]+", "_", s).strip("_")
    return s[:60] or "unnamed"


def summarize_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {"sheets": wb.sheetnames, "preview": {}}
        for name in wb.sheetnames[:2]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            out["preview"][name] = rows
        wb.close()
        return out
    except Exception as e:
        return {"error": str(e)}


def try_export(page: Page, button_text: str, save_label: str) -> dict:
    """export 버튼 클릭 → 다운로드 캡처. 모달이 뜨면 확인 클릭."""
    targets = [page.main_frame] + [f for f in page.frames if f != page.main_frame]
    for fr in targets:
        try:
            loc = fr.locator(f"button:has-text('{button_text}')").first
            if loc.count() == 0:
                continue
            if not loc.is_visible():
                continue
            try:
                with page.expect_download(timeout=15000) as dl_info:
                    loc.click(force=True, timeout=3000)
                    # 확인 모달 대응
                    try:
                        for ok in ["확인", "다운로드", "OK", "저장"]:
                            ok_loc = fr.locator(f"button:has-text('{ok}')").last
                            if ok_loc.count() > 0 and ok_loc.is_visible():
                                ok_loc.click(timeout=1500)
                                break
                    except Exception:
                        pass
                dl: Download = dl_info.value
                suffix = Path(dl.suggested_filename).suffix or ".xlsx"
                path = DL_DIR / (safe_filename(save_label) + suffix)
                dl.save_as(str(path))
                size = path.stat().st_size if path.exists() else 0
                logger.info(f"      ★ 저장: {path.name} ({size}B)")
                return {"saved": str(path), "size": size, "summary": summarize_xlsx(path) if suffix.lower() in (".xlsx", ".xls") else None}
            except PWTimeout:
                logger.warning(f"      다운로드 타임아웃 (버튼 '{button_text}')")
                return {"saved": None, "error": "timeout"}
            except Exception as e:
                logger.warning(f"      다운로드 실패: {e}")
                return {"saved": None, "error": str(e)}
        except Exception:
            continue
    return {"saved": None, "error": "button_not_found"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--top_only", type=str, default=None, help="HR 카테고리 필터 (쉼표)")
    ap.add_argument("--leaf_limit", type=int, default=15, help="카테고리당 leaf 클릭 한도")
    ap.add_argument("--download_limit", type=int, default=5, help="실제 다운로드 시도 한도")
    args = ap.parse_args()

    DL_DIR.mkdir(parents=True, exist_ok=True)
    SCR_DIR.mkdir(parents=True, exist_ok=True)

    top_filter = set(args.top_only.split(",")) if args.top_only else None
    target_tops = [t for t in HR_TOP_CATEGORIES if (not top_filter) or t in top_filter]

    result = {
        "gw_user": GW_USER,
        "crawled_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "hr_top_categories": target_tops,
        "categories": [],   # {top, sub_lnb: [{text, url, buttons, exports, downloads}]}
        "errors": [],
    }
    downloads_done = 0

    with sync_playwright() as pw:
        browser, context, page = login_and_get_context(
            playwright_instance=pw,
            headless=args.headless,
            user_id=GW_USER,
        )
        m = re.match(r"(https://[^/]+)", page.url)
        base = m.group(1) if m else "https://gw.glowseoul.co.kr"

        # 홈 진입 → HR 클릭
        page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(4000)
        try:
            page.locator("span.module-link.HR").first.click(force=True, timeout=5000)
            page.wait_for_timeout(4000)
            logger.info("HR 모듈 진입 완료")
        except Exception as e:
            logger.error(f"HR 클릭 실패: {e}")
            return
        try:
            page.screenshot(path=str(SCR_DIR / "hr_after_enter.png"))
        except Exception:
            pass

        for top in target_tops:
            logger.info(f"\n=== HR > {top} ===")
            # 매번 HR 홈으로 복귀하지 않고, 직접 클릭만 시도 (LNB가 펼쳐진 상태일 가능성)
            clicked = click_text_in_any_frame(page, top, timeout_ms=3000)
            if not clicked:
                # 폴백: HR 홈 재진입 후 다시 클릭
                logger.warning(f"  '{top}' 1차 클릭 실패 — HR 재진입")
                page.goto(f"{base}/#/", wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(2000)
                try:
                    page.locator("span.module-link.HR").first.click(force=True, timeout=3000)
                    page.wait_for_timeout(3000)
                except Exception:
                    pass
                clicked = click_text_in_any_frame(page, top, timeout_ms=3000)
            if not clicked:
                result["errors"].append({"phase": "top_click", "top": top})
                continue
            page.wait_for_timeout(2500)
            try:
                page.screenshot(path=str(SCR_DIR / f"hr_{safe_filename(top)}.png"))
            except Exception:
                pass

            # 2-depth LNB 수집
            sub_items = collect_all_frames(page, JS_VISIBLE_LIST_ITEMS, list(LNB_BLACKLIST | set(HR_TOP_CATEGORIES)))
            # 텍스트 중복 제거
            dedup = {}
            for it in sub_items:
                t = it.get("text")
                if not t or t in dedup:
                    continue
                dedup[t] = it
            sub_unique = list(dedup.values())
            logger.info(f"  2-depth LNB 후보 {len(sub_unique)}개")

            cat_entry = {"top": top, "sub_count": len(sub_unique), "sub_lnb": []}

            for idx, sub in enumerate(sub_unique[: args.leaf_limit]):
                sub_text = sub["text"]
                logger.info(f"  [{idx+1}/{min(len(sub_unique), args.leaf_limit)}] '{sub_text}'")
                ok = click_text_in_any_frame(page, sub_text, timeout_ms=3000)
                page.wait_for_timeout(2200)
                # 페이지 정보 수집
                url = current_content_url(page)
                btn_data = collect_all_frames(page, JS_BUTTONS, EXPORT_KEYWORDS)
                merged_buttons = []
                merged_exports = []
                title = ""
                seen_b = set()
                for d in btn_data:
                    if not title and d.get("title"):
                        title = d["title"]
                    for b in d.get("buttons", []):
                        if b["text"] in seen_b:
                            continue
                        seen_b.add(b["text"])
                        merged_buttons.append(b)
                    merged_exports.extend(d.get("export_candidates", []))

                # 실제 다운로드 시도 (한도 내)
                dl_result = None
                if merged_exports and downloads_done < args.download_limit:
                    # 우선순위: "엑셀 다운로드" > "다운로드" > "엑셀" > 기타
                    priorities = ["엑셀 다운로드", "엑셀다운로드", "다운로드", "엑셀", "내보내기"]
                    sorted_exp = sorted(merged_exports, key=lambda e: next((i for i,p in enumerate(priorities) if p in e["text"]), 99))
                    for ex in sorted_exp[:2]:
                        save_label = f"HR_{safe_filename(top)}_{safe_filename(sub_text)}"
                        logger.info(f"    → 다운로드 시도: '{ex['text']}'")
                        dl_result = try_export(page, ex["text"], save_label)
                        if dl_result and dl_result.get("saved"):
                            downloads_done += 1
                            break

                cat_entry["sub_lnb"].append({
                    "text": sub_text,
                    "url": url,
                    "page_title": title,
                    "buttons_count": len(merged_buttons),
                    "buttons_sample": [b["text"] for b in merged_buttons[:20]],
                    "export_candidates": [e["text"] for e in merged_exports],
                    "download": dl_result,
                })
                if merged_exports:
                    logger.info(f"    ★ Export: {[e['text'] for e in merged_exports]}")

            result["categories"].append(cat_entry)

        context.close()
        browser.close()

    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"JSON 저장: {OUT_JSON}")

    # Markdown 요약
    lines = ["# HR 모듈 깊이 탐색 + Export 다운로드 (v4)", ""]
    lines.append(f"- 계정: `{GW_USER}` / 시각: {result['crawled_at']}")
    lines.append(f"- 탐색 카테고리: {len(result['categories'])} / 다운로드 성공: {downloads_done}")
    lines.append("")
    for cat in result["categories"]:
        lines.append(f"## {cat['top']} ({cat['sub_count']} sub)")
        lines.append("| Sub | URL | 버튼 수 | Export | 다운로드 |")
        lines.append("|---|---|---|---|---|")
        for s in cat["sub_lnb"]:
            ex = " / ".join(s["export_candidates"]) if s["export_candidates"] else ""
            dl = s.get("download") or {}
            saved = dl.get("saved") if isinstance(dl, dict) else None
            dl_show = "✅ " + Path(saved).name if saved else ("❌ " + (dl.get("error", "") if isinstance(dl, dict) else ""))
            lines.append(f"| {s['text']} | `{(s['url'] or '')[-60:]}` | {s['buttons_count']} | {ex} | {dl_show} |")
        lines.append("")
        # 다운로드 미리보기
        for s in cat["sub_lnb"]:
            dl = s.get("download")
            if not dl or not dl.get("saved"):
                continue
            summary = dl.get("summary")
            if not summary or "error" in summary:
                continue
            lines.append(f"### 📄 {s['text']} — 미리보기")
            for sheet, rows in summary.get("preview", {}).items():
                lines.append(f"**시트**: `{sheet}`")
                if rows:
                    headers = rows[0]
                    lines.append("| " + " | ".join(str(h)[:30] for h in headers) + " |")
                    lines.append("|" + "|".join(["---"] * len(headers)) + "|")
                    for row in rows[1:6]:
                        padded = row + [""] * (len(headers) - len(row))
                        lines.append("| " + " | ".join(str(c)[:30] for c in padded[:len(headers)]) + " |")
            lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"MD 저장: {OUT_MD}")

    logger.info("=" * 60)
    logger.info(f"카테고리: {len(result['categories'])} / 다운로드 성공: {downloads_done}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
